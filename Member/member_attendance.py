import customtkinter as ctk
from database import Database
from datetime import datetime, timedelta
from datetime import time as dt_time
from tkinter import filedialog, messagebox, ttk
from xml.sax.saxutils import escape
from tkcalendar import Calendar
import tkinter as tk
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch


class DatePickerButton(ctk.CTkFrame):
    def __init__(self, master, initial_date=None, **kwargs):
        super().__init__(master, fg_color="transparent")

        self._date = initial_date or datetime.today().date()
        self._open = False
        self._callback = None

        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Custom.Calendar",
            background="#1A1A2E",
            foreground="white",
            headersbackground="#16213E",
            headersforeground="#4FC3F7",
            selectbackground="#3498DB",
            selectforeground="white",
            normalbackground="#1A1A2E",
            normalforeground="#CCCCCC",
            weekendbackground="#1A1A2E",
            weekendforeground="#F39C12",
            othermonthforeground="#555555",
            bordercolor="#2A2A4A",
            relief="flat",
        )

        self.btn = ctk.CTkButton(
            self,
            text=self._fmt(),
            width=140,
            height=36,
            corner_radius=8,
            fg_color=("#F2F2F2", "#1E1E1E"),
            text_color=("#000000", "#FFFFFF"),
            hover_color=("#E5E5E5", "#2A2A2A"),
            border_width=1,
            border_color=("#CCCCCC", "#2C2C2C"),
            anchor="w",
            command=self.toggle,
        )
        self.btn.pack()

        self.panel = ctk.CTkFrame(
            self.winfo_toplevel(),
            fg_color=("#FFFFFF", "#141E2B"),
            corner_radius=12,
            border_width=1,
            border_color=("#CCCCCC", "#2A3A4A"),
        )

        self.cal = Calendar(
            self.panel,
            style="Custom.Calendar",
            selectmode="day",
            date_pattern="yyyy-mm-dd",
            year=self._date.year,
            month=self._date.month,
            day=self._date.day,
            showweeknumbers=False,
            firstweekday="monday",
            font=("Arial", 11),
            cursor="hand2",
        )
        self.cal.pack(padx=8, pady=8)
        self.cal.bind("<<CalendarSelected>>", self._select)

    def toggle(self):
        if self._open:
            self.panel.place_forget()
        else:
            self.panel.lift()
            self.panel.place(in_=self, x=0, y=self.btn.winfo_height() + 2)
        self._open = not self._open

    def _select(self, _event=None):
        selected = self.cal.get_date()
        self._date = datetime.strptime(selected, "%Y-%m-%d").date()
        self.btn.configure(text=self._fmt())
        if self._callback:
            self._callback(self._date)
        self.toggle()

    def on_change(self, callback):
        self._callback = callback

    def get_date(self):
        return self._date

    def set_date(self, d):
        self._date = d
        self.cal.selection_set(d)
        self.btn.configure(text=self._fmt())

    def _fmt(self):
        return f"📅 {self._date.strftime('%Y-%m-%d')}"

class MemberAttendance(ctk.CTkFrame):
    def __init__(self, master, user):
        super().__init__(master, fg_color=("white", "#1A1A1A"))
        self.db = Database()
        self.user = user
        self.LATE_TIME = "07:45:00"  # align late threshold with Admin/Leader views

        self.month_rows = []
        self.current_rows = []
        self.date_filter_enabled = False
        self.month_total_workdays = "-"
        self.metric_colors = {
            "work": "#27AE60",
            "leave": "#3498DB",
            "late": "#E67E22",
            "ot": "#8B5CF6",
        }
        self.default_from_date, self.default_to_date = self._default_payroll_range()

        self.setup_ui()

    # ---------------- UI ----------------
    def setup_ui(self):
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=80, pady=14)

        ctk.CTkLabel(
            header,
            text="📅 Attendance & Work Summary",
            font=("Arial", 20, "bold"),
            text_color=("#1A1A1A", "#FFFFFF")
        ).pack(side="left")

        # ---------------- FILTER ----------------
        filter_frame = ctk.CTkFrame(
            self,
            fg_color=("#F6F8FC", "#1F2833"),
            corner_radius=16,
            border_width=1,
            border_color=("#D6DEEB", "#3A4656"),
        )
        filter_frame.pack(fill="x", padx=80, pady=(0, 18))
        filter_frame.grid_columnconfigure(8, weight=1)

        ctk.CTkLabel(filter_frame, text="From").grid(row=0, column=0, padx=(20, 6), pady=8, sticky="w")

        self.date_picker = DatePickerButton(filter_frame, initial_date=self.default_from_date, width=140, height=36, corner_radius=8)
        self.date_picker.grid(row=0, column=1, padx=6, pady=8)

        ctk.CTkLabel(filter_frame, text="To").grid(row=0, column=2, padx=(18, 6), pady=8, sticky="w")
        self.to_date_picker = DatePickerButton(filter_frame, initial_date=self.default_to_date, width=140, height=36, corner_radius=8)
        self.to_date_picker.grid(row=0, column=3, padx=6, pady=8)

        ctk.CTkButton(
            filter_frame,
            text="🔍 Filter",
            width=65,
            height=36,
            fg_color=("#2563EB", "#1D4ED8"),
            hover_color=("#1D4ED8", "#1E40AF"),
            text_color=("#1A1A1A", "#FFFFFF"),
            corner_radius=8,
            command=self.load_data
        ).grid(row=0, column=4, padx=6, pady=8)

        ctk.CTkButton(
            filter_frame,
            text="✖ Clear",
            width=65,
            height=36,
            fg_color=("#BABDC1", "#6B7280"),
            hover_color=("#AEB3B8", "#4B5563"),
            text_color=("#1A1A1A", "#FFFFFF"),
            corner_radius=8,
            command=self.reset_filters
        ).grid(row=0, column=5, padx=6, pady=8)

        ctk.CTkButton(
            filter_frame,
            text="📄 Export",
            width=65,
            height=36,
            fg_color="#DC2626",
            hover_color="#B91C1C",
            text_color=("#1A1A1A", "#FFFFFF"),
            corner_radius=8,
            command=self.export_pdf
        ).grid(row=0, column=6, padx=6, pady=8)

        ctk.CTkButton(
            filter_frame,
            text="CSV",
            width=65,
            height=36,
            fg_color="#26DCB5",
            text_color=("#1A1A1A", "#FFFFFF"),
            corner_radius=8,
            command=self.export_excel_file
        ).grid(row=0, column=7, padx=6, pady=8)

        # ---------------- STATS ----------------
        self.stats_container = ctk.CTkFrame(self, fg_color="transparent")
        self.stats_container.pack(fill="x", padx=80, pady=5)

        self.month_work_lbl = self.create_stat_card("Month Work", "0", self.metric_colors["work"])
        self.month_hours_lbl = self.create_stat_card("Monthly Hours", "0", self.metric_colors["work"])
        self.actual_hours_lbl = self.create_stat_card("Actual Hours", "0", self.metric_colors["work"])
        self.working_lbl = self.create_stat_card("Working Days", "0", self.metric_colors["work"])
        self.leave_lbl = self.create_stat_card("Leave Days", "0", self.metric_colors["leave"])
        self.late_lbl = self.create_stat_card("Late Days", "0", self.metric_colors["late"])
        self.ot_lbl = self.create_stat_card("OT Hours", "0", self.metric_colors["ot"])

        # ---------------- TABLE ----------------
        self.table_card = ctk.CTkFrame(
            self,
            fg_color=("#FFFFFF", "#1B1F24"),
            corner_radius=16,
            border_width=1,
            border_color=("#D6DEEB", "#2A3442"),
        )
        self.table_card.pack(fill="both", expand=True, padx=80, pady=(0, 20))

        header_row = ctk.CTkFrame(self.table_card, fg_color=("#BABDC1", "#6B7280"))
        header_row.pack(fill="x", padx=20, pady=(10, 0))

        def h(text, w):
            return ctk.CTkLabel(header_row, text=text, width=w,anchor="w",
                                text_color=("#000000", "#FFFFFF"), font=("Arial", 12, "bold"))

        h("Date", 170).grid(row=0, column=0, padx=(20, 0), pady=8, sticky="w")
        h("Check-In", 140).grid(row=0, column=1, padx=0, pady=8, sticky="w")
        h("Check-Out", 140).grid(row=0, column=2, padx=0, pady=8, sticky="w")
        h("Total Hours", 120).grid(row=0, column=3, padx=0, pady=8, sticky="w")
        h("Stacked Hours", 130).grid(row=0, column=4, padx=0, pady=8, sticky="w")
        h("OT Hours", 120).grid(row=0, column=5, padx=10, pady=8, sticky="w")
        h("Remark", 120).grid(row=0, column=6, padx=(10, 0), pady=8, sticky="w")
        header_row.grid_columnconfigure(7, weight=1)

        self.scroll = ctk.CTkScrollableFrame(self.table_card, fg_color="transparent")
        self.scroll.pack(fill="both", expand=True, padx=10, pady=10)

        self.load_data()

    # ---------------- CARD ----------------
    def create_stat_card(self, title, value, color):
        card = ctk.CTkFrame(self.stats_container, corner_radius=10, width=180)
        card.pack(side="left", padx=(25, 25))

        ctk.CTkLabel(card, text=title).pack(padx=15, pady=(5, 0))

        val_lbl = ctk.CTkLabel(
            card,
            text=value,
            font=("Arial", 18, "bold"),
            text_color=color
        )
        val_lbl.pack(pady=(0, 5))

        return val_lbl

    @staticmethod
    def _current_payroll_month_key():
        today = datetime.now()
        if today.day >= 26:
            if today.month == 12:
                return f"{today.year + 1}-01"
            return f"{today.year}-{today.month + 1:02d}"
        return today.strftime("%Y-%m")

    @staticmethod
    def _calculate_month_total_workdays(month_key):
        if not month_key:
            return "-"

        payroll_month = datetime.strptime(month_key, "%Y-%m")
        end_date = payroll_month.replace(day=25).date()
        if payroll_month.month == 1:
            start_date = payroll_month.replace(year=payroll_month.year - 1, month=12, day=26).date()
        else:
            start_date = payroll_month.replace(month=payroll_month.month - 1, day=26).date()

        count = 0
        current = start_date
        while current <= end_date:
            if current.weekday() < 5:
                count += 1
            current += timedelta(days=1)

        return str(count)

    @staticmethod
    def _dash(value):
        return "-" if value in (None, "") else str(value)

    @staticmethod
    def _month_hours_from_workdays(workdays):
        try:
            return f"{float(workdays) * 8:g}"
        except (TypeError, ValueError):
            return "-"

    @staticmethod
    def _default_payroll_range():
        today = datetime.now().date()
        if today.month == 1:
            start = today.replace(year=today.year - 1, month=12, day=26)
        else:
            start = today.replace(month=today.month - 1, day=26)
        end = today.replace(day=25)
        return start, end

    @staticmethod
    def _to_time(value):
        if value in (None, ""):
            return None
        if isinstance(value, dt_time):
            return value
        try:
            return datetime.strptime(str(value), "%H:%M:%S").time()
        except Exception:
            try:
                return datetime.strptime(str(value), "%H:%M").time()
            except Exception:
                return None

    def _actual_hours_from_row(self, row):
        check_in = self._to_time(row.get("check_in"))
        check_out = self._effective_checkout_time(row)
        if check_in is None:
            return 0.0

        leave_type = str(row.get("leave_type") or "").lower()
        work_start = max(check_in, dt_time(7, 45))

        if leave_type == "full_day":
            return 0.0
        if leave_type == "morning_half":
            work_end = check_out if check_out is not None else dt_time(11, 45)
        elif leave_type == "evening_half":
            work_end = check_out if check_out is not None else dt_time(12, 30)
        else:
            work_end = check_out

        start_dt = datetime.combine(datetime.today(), work_start)
        end_dt = datetime.combine(datetime.today(), work_end)
        return max((end_dt - start_dt).total_seconds() / 3600.0, 0.0)

    def _effective_checkout_time(self, row):
        check_out = self._to_time(row.get("check_out"))
        try:
            has_accepted_ot_request = float(row.get("ot_hours") or 0) > 0
        except (TypeError, ValueError):
            has_accepted_ot_request = False
        if has_accepted_ot_request and check_out is not None:
            return check_out
        return dt_time(16, 30)

    def _total_hours_between_checkin_checkout(self, row):
        check_in = self._to_time(row.get("check_in"))
        check_out = self._effective_checkout_time(row)
        if check_in is None:
            return 0.0
        start_dt = datetime.combine(datetime.today(), check_in)
        end_dt = datetime.combine(datetime.today(), check_out)
        return max((end_dt - start_dt).total_seconds() / 3600.0, 0.0)

    @staticmethod
    def _fmt_hours(value):
        total_minutes = int(round(max(float(value or 0), 0.0) * 60))
        hh = total_minutes // 60
        mm = total_minutes % 60
        return f"{hh}:{mm:02d}"

    def _stacked_hours_from_row(self, row):
        total_hours = self._total_hours_between_checkin_checkout(row)
        return max(total_hours - 8.0, 0.0)

    def _ot_hours_from_row(self, row):
        work_date = row.get("attendance_date")
        try:
            work_date_obj = datetime.strptime(str(work_date), "%Y-%m-%d").date() if work_date else None
        except Exception:
            work_date_obj = None
        is_weekend = bool(work_date_obj and work_date_obj.weekday() >= 5)
        has_accepted_ot_request = float(row.get("ot_hours") or 0) > 0
        if is_weekend and has_accepted_ot_request:
            return self._total_hours_between_checkin_checkout(row)

        stacked_hours = self._stacked_hours_from_row(row)
        return max(stacked_hours - 1.0, 0.0)

    def _late_remark_text(self, row):
        check_in = self._to_time(row.get("check_in"))
        late_cutoff = self._to_time("07:45:00")
        if check_in is None or late_cutoff is None or check_in <= late_cutoff:
            return "ON TIME", self.metric_colors["work"]
        late_mins = int((datetime.combine(datetime.today(), check_in) - datetime.combine(datetime.today(), late_cutoff)).total_seconds() // 60)
        if late_mins >= 60:
            h = late_mins // 60
            m = late_mins % 60
            if m == 0:
                late_txt = f"LATE ({h}h)"
            else:
                late_txt = f"LATE ({h}h {m}m)"
        else:
            late_txt = f"LATE ({late_mins}m)"
        return late_txt, self.metric_colors["late"]

    # ---------------- DATA ----------------
    def load_data(self):
        for w in self.scroll.winfo_children():
            w.destroy()

        user_id = self.user["id"]
        from_date = self.date_picker.get_date()
        to_date = self.to_date_picker.get_date()
        if from_date > to_date:
            from_date, to_date = to_date, from_date
        self.month_total_workdays = str(sum(1 for i in range((to_date - from_date).days + 1) if (from_date + timedelta(days=i)).weekday() < 5))
        self.month_work_lbl.configure(text=self.month_total_workdays)
        self.month_hours_lbl.configure(text=self._month_hours_from_workdays(self.month_total_workdays))

        sql = """
        SELECT
            d.work_date AS attendance_date,
            a.check_in,
            a.check_out,
            IFNULL(o.ot_hours, 0) AS ot_hours,
            CASE
                WHEN EXISTS (
                    SELECT 1
                    FROM leave_requests lr
                    WHERE lr.user_id = %s
                      AND lr.status = 'Approved'
                      AND d.work_date BETWEEN lr.start_date AND lr.end_date
                ) THEN 1
                WHEN a.check_in IS NOT NULL OR a.check_out IS NOT NULL THEN 1
                WHEN IFNULL(o.ot_hours, 0) > 0 THEN 1
                ELSE 0
            END AS has_attendance,
            CASE
                WHEN EXISTS (
                    SELECT 1
                    FROM leave_requests lr
                    WHERE lr.user_id = %s
                      AND lr.status = 'Approved'
                      AND d.work_date BETWEEN lr.start_date AND lr.end_date
                      AND (
                          (lr.start_date = d.work_date AND lr.end_date = d.work_date AND lr.start_shift = 'Morning' AND lr.end_shift = 'Morning')
                          OR (lr.end_date = d.work_date AND lr.end_date > lr.start_date AND lr.end_shift = 'Morning')
                      )
                ) THEN 'morning_half'
                WHEN EXISTS (
                    SELECT 1
                    FROM leave_requests lr
                    WHERE lr.user_id = %s
                      AND lr.status = 'Approved'
                      AND d.work_date BETWEEN lr.start_date AND lr.end_date
                      AND (
                          (lr.start_date = d.work_date AND lr.end_date = d.work_date AND lr.start_shift = 'Evening' AND lr.end_shift = 'Evening')
                          OR (lr.start_date = d.work_date AND lr.end_date > lr.start_date AND lr.start_shift = 'Evening')
                      )
                ) THEN 'evening_half'
                WHEN EXISTS (
                    SELECT 1
                    FROM leave_requests lr
                    WHERE lr.user_id = %s
                      AND lr.status = 'Approved'
                      AND d.work_date BETWEEN lr.start_date AND lr.end_date
                ) THEN 'full_day'
                ELSE ''
            END AS leave_type
        FROM (
            SELECT attendance_date AS work_date
            FROM attendance
            WHERE user_id = %s
            UNION
            SELECT DATE(ot_date) AS work_date
            FROM overtime_requests
            WHERE member_id = %s
              AND status IN ('Accepted', 'Approved')
        ) d
        LEFT JOIN attendance a
               ON a.user_id = %s
              AND a.attendance_date = d.work_date
        LEFT JOIN (
            SELECT DATE(ot_date) AS ot_date, ROUND(SUM(COALESCE(hours, 0)), 2) AS ot_hours
            FROM overtime_requests
            WHERE member_id = %s
              AND status IN ('Accepted', 'Approved')
            GROUP BY DATE(ot_date)
        ) o ON d.work_date = o.ot_date
        WHERE d.work_date IS NOT NULL
          AND d.work_date BETWEEN %s AND %s
        """
        params = [user_id, user_id, user_id, user_id, user_id, user_id, user_id, user_id, from_date, to_date]

        sql += " ORDER BY d.work_date DESC"

        self.db.cursor.execute(sql, tuple(params))
        month_rows = self.db.cursor.fetchall()
        self.month_rows = month_rows

        leave_sql = """
            SELECT IFNULL(SUM(
                CASE
                    WHEN total_days IS NOT NULL THEN total_days
                    ELSE CASE
                        WHEN DATEDIFF(end_date, start_date) = 0 THEN
                            CASE
                                WHEN start_shift = 'Full Day' OR end_shift = 'Full Day' THEN 1.0
                                WHEN start_shift = end_shift THEN 0.5
                                ELSE 1.0
                            END
                        ELSE (DATEDIFF(end_date, start_date) + 1)
                             - (CASE WHEN start_shift = 'Full Day' THEN 0 ELSE 0.5 END)
                             - (CASE WHEN end_shift = 'Full Day' THEN 0 ELSE 0.5 END)
                        END
                END
            ), 0) leave_days
            FROM leave_requests
            WHERE user_id=%s
              AND status='Approved'
        """
        leave_sql += " AND start_date <= %s AND end_date >= %s "
        leave_params = [user_id, to_date, from_date]

        self.db.cursor.execute(leave_sql, tuple(leave_params))
        leave_row = self.db.cursor.fetchone() or {"leave_days": 0}
        leave_days = float(leave_row["leave_days"] or 0)

        # store leave and ot totals for export/reporting
        late = 0
        ot_total = 0.0
        self._export_leave_days = leave_days
        self._export_ot_hours = 0.0

        for r in month_rows:
            if r["check_in"] and str(r["check_in"]) > self.LATE_TIME:
                late += 1
            ot_total += self._ot_hours_from_row(r)

        # save computed ot_total for export
        self._export_ot_hours = ot_total

        working_days = sum(int(r.get("has_attendance", 0) or 0) for r in month_rows)
        actual_hours = sum(self._actual_hours_from_row(r) for r in month_rows)
        actual_hours_with_ot = actual_hours + ot_total
        self.actual_hours_lbl.configure(text=f"{actual_hours_with_ot:g}")
        self.working_lbl.configure(text=f"{working_days:g}")
        self.leave_lbl.configure(text=f"{leave_days:g}")
        self.late_lbl.configure(text=str(late))
        self.ot_lbl.configure(text=self._fmt_hours(ot_total))

        display = []

        for r in month_rows:
            if not r["attendance_date"] and self._ot_hours_from_row(r) <= 0:
                continue

            stacked_hours = self._stacked_hours_from_row(r)
            derived_ot = self._ot_hours_from_row(r)
            display.append(r)

        self.current_rows = display

        # ================= EMPTY MESSAGE =================
        if not display:

            empty_frame = ctk.CTkFrame(
                self.scroll,
                fg_color="transparent"
            )

            empty_frame.pack(
                fill="both",
                expand=True,
                pady=40
            )

            ctk.CTkLabel(
                empty_frame,
                text="There is no Records",
                font=("Arial", 18, "bold"),
                text_color=("#6B7280", "#9CA3AF")
            ).pack(pady=(20, 5))

            ctk.CTkLabel(
                empty_frame,
                text="Try changing the filter date range.",
                font=("Arial", 12),
                text_color=("#9CA3AF", "#6B7280")
            ).pack()

            return
        for r in display:
            row = ctk.CTkFrame(self.scroll, fg_color=("#EEF2F7", "#1F2933"), corner_radius=10)
            row.pack(fill="x", pady=4)
            total_hours = self._total_hours_between_checkin_checkout(r)
            stacked_hours = self._stacked_hours_from_row(r)
            derived_ot = self._ot_hours_from_row(r)

            ctk.CTkLabel(row, text=self._dash(r["attendance_date"]), width=170, anchor="w").grid(row=0, column=0, padx=(10, 0), pady=8, sticky="w")
            ctk.CTkLabel(row, text=self._dash(r["check_in"]), width=140, anchor="w").grid(row=0, column=1, padx=0, pady=8, sticky="w")
            ctk.CTkLabel(row, text=self._dash(r["check_out"]), width=140, anchor="w").grid(row=0, column=2, padx=0, pady=8, sticky="w")
            ctk.CTkLabel(row, text=self._fmt_hours(total_hours), width=120, anchor="w").grid(row=0, column=3, padx=0, pady=8, sticky="w")
            ctk.CTkLabel(
                row,
                text=self._fmt_hours(stacked_hours),
                width=130,
                anchor="w",
                text_color=self.metric_colors["work"],
            ).grid(row=0, column=4, padx=0, pady=8, sticky="w")
            ctk.CTkLabel(row, text=self._fmt_hours(derived_ot), width=120, text_color=self.metric_colors["ot"]).grid(row=0, column=5, padx=0, pady=8, sticky="w")

            remark, color = self._late_remark_text(r)

            ctk.CTkLabel(row, text=remark, text_color=color, width=120).grid(row=0, column=6, padx=(0, 0), pady=8, sticky="w")
            row.grid_columnconfigure(7, weight=1)

    # ---------------- RESET ----------------
    def reset_filters(self):
        # ===== CURRENT PAYROLL MONTH =====
        today = datetime.now()
        if today.day >= 26:
            if today.month == 12:
                current_ym = f"{today.year + 1}-01"
            else:
                current_ym = f"{today.year}-{today.month + 1:02d}"
        else:
            current_ym = today.strftime("%Y-%m")

        from_date, to_date = self._default_payroll_range()
        self.date_picker.set_date(from_date)
        self.to_date_picker.set_date(to_date)

        self.load_data()

    def export_excel_file(self):
        if not self.current_rows:
            messagebox.showwarning("No Data", "No records to export.")
            return

        selected_month = f"{self.date_picker.get_date()}_{self.to_date_picker.get_date()}"
        selected_member = str(self.user.get("full_name", "Member")).replace(" ", "_")
        file_name = f"Attendance_{selected_member}_{selected_month}.xlsx"

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=file_name,
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not file_path:
            return
        if not file_path.lower().endswith(".xlsx"):
            file_path = f"{file_path}.xlsx"

        self._export_excel(file_path)

    # ---------------- PDF ----------------
    def export_pdf(self):
        if not self.current_rows:
            messagebox.showwarning("No Data", "No records to export.")
            return

        selected_month = f"{self.date_picker.get_date()}_{self.to_date_picker.get_date()}"
        selected_member = str(self.user.get("full_name", "Member")).replace(" ", "_")
        file_name = f"Attendance_{selected_member}_{selected_month}.pdf"

        file = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            initialfile=file_name,
            filetypes=[("PDF files", "*.pdf"), ("Excel files", "*.xlsx")],
        )
        if not file:
            return
        if file.lower().endswith(".xlsx"):
            self._export_excel(file)
            return

        doc = SimpleDocTemplate(
            file,
            pagesize=letter,
            leftMargin=28,
            rightMargin=28,
            topMargin=28,
            bottomMargin=28,
        )
        styles = getSampleStyleSheet()
        elements = []
        safe_member = escape(str(self.user.get("full_name", "Member")))
        safe_member_id = escape(str(self.user.get("employee_id") or self.user.get("id") or "-"))
        safe_team = escape(str(self.user.get("team_name") or f"Team-{self.user.get('team_id', '-')}" ))
        safe_month = escape(f"{self.date_picker.get_date()} to {self.to_date_picker.get_date()}")
        month_workdays = int(self.month_total_workdays or 0) if str(self.month_total_workdays).isdigit() else 0
        actual_workdays = sum(int(r.get("has_attendance", 0) or 0) for r in self.current_rows)
        # prefer leave/ot values computed during load_data when available
        leave_days = getattr(self, "_export_leave_days", max(month_workdays - actual_workdays, 0))
        total_ot_hours = getattr(self, "_export_ot_hours", sum(self._ot_hours_from_row(r) for r in self.current_rows))
        month_hours = float(self._month_hours_from_workdays(month_workdays))
        actual_hours = sum(self._actual_hours_from_row(r) for r in self.current_rows)
        actual_hours_with_ot = actual_hours + total_ot_hours
        safe_month_work = escape(str(month_workdays))
        safe_actual_work = escape(f"{actual_workdays:g}")
        safe_leave = escape(f"{leave_days:g}")
        safe_month_hours = escape(f"{month_hours:g}")
        safe_actual_hours = escape(f"{actual_hours_with_ot:g}")
        safe_ot_hours = escape(self._fmt_hours(total_ot_hours))

        title_style = ParagraphStyle(
            "MemberTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=22,
            leading=26,
            textColor=colors.HexColor("#0B1220"),
            alignment=1,
            spaceAfter=6,
        )
        meta_style = ParagraphStyle(
            "MemberMeta",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=10,
            leading=13,
            textColor=colors.HexColor("#334155"),
        )

        elements.append(Paragraph("Employee Attendance Report", title_style))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", meta_style))
        elements.append(Spacer(1, 10))

        info_table = Table(
            [[
                Paragraph(
                    f"<b><font color='#1D4ED8'>Month:</font></b> {safe_month}<br/>"
                    f"<b><font color='#1D4ED8'>Team:</font></b> {safe_team}<br/>"
                    f"<b><font color='#1D4ED8'>Employee:</font></b> {safe_member}<br/>"
                    f"<b><font color='#1D4ED8'>Employee ID:</font></b> {safe_member_id}",
                    meta_style
                ),
                Paragraph(
                    f"<b><font color='#1D4ED8'>Month Work:</font></b> {safe_month_work}<br/>"
                    f"<b><font color='#1D4ED8'>Actual Work:</font></b> {safe_actual_work}<br/>"
                    f"<b><font color='#1D4ED8'>Leave:</font></b> {safe_leave}",
                    meta_style
                ),
                Paragraph(
                    f"<b><font color='#1D4ED8'>Month Hours:</font></b> {safe_month_hours}<br/>"
                    f"<b><font color='#1D4ED8'>Actual Hours:</font></b> {safe_actual_hours}<br/>"
                    f"<b><font color='#1D4ED8'>OT Hours:</font></b> {safe_ot_hours}",
                    meta_style
                ),
            ]],
            colWidths=[2.45 * inch, 2.35 * inch, 2.2 * inch],
        )
        info_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
            ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#CBD5E1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#E2E8F0")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 18))

        data = [["Date", "Check-In", "Check-Out", "OT Hours", "Remark"]]
        for r in self.current_rows:
            remark, remark_color = self._late_remark_text(r)
            data.append([
                self._dash(r["attendance_date"]),
                self._dash(r["check_in"]),
                self._dash(r["check_out"]),
                self._fmt_hours(self._ot_hours_from_row(r)),
                Paragraph(f"<font color='{remark_color}'>{remark}</font>", meta_style)
            ])

        table = Table(
            data,
            repeatRows=1,
            colWidths=[1.7 * inch, 1.45 * inch, 1.45 * inch, 1.0 * inch, 1.6 * inch],
        )
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#FFFFFF"), colors.HexColor("#F8FAFC")]),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("TEXTCOLOR", (3, 1), (3, -1), colors.HexColor("#7C3AED")),
        ]))

        elements.append(table)
        doc.build(elements)
        messagebox.showinfo("Export Successful", "Attendance PDF exported successfully.")

    def _export_excel(self, file_path):
        rows = self.current_rows or []

        if not rows:
            messagebox.showwarning(
                "No Data",
                "No records to export."
            )
            return
        ot_hours = sum(self._ot_hours_from_row(r) for r in rows)
        working_days = sum(int(r.get("workdays", 0) or 0) for r in rows)
        working_hours = sum(self._actual_hours_from_row(r) for r in rows)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except Exception as e:
            messagebox.showerror("Missing Dependency", f"Cannot export Excel: {e}")
            return

        try:
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance Report"

            # ================= EMPLOYEE INFO =================
            ws["A1"] = "Team"
            ws["B1"] = self.user.get("team_name") or f"Team-{self.user.get('team_id', '-') }"
            ws["A2"] = "Employee ID"
            ws["B2"] = str(self.user.get("employee_id") or self.user.get("id") or "-")
            ws["A3"] = "Employee Name"
            ws["B3"] = str(self.user.get("full_name", "Member"))

            ws["D1"] = "Working Days"
            ws["E1"] = working_days
            ws["D2"] = "Working Hours"
            ws["E2"] = working_hours
            ws["D3"] = "OT Hours"
            ws["E3"] = ot_hours

            # ================= TABLE HEADER =================
            header_row = 5
            headers = ["Date", "Check-In", "Check-Out", "OT Hours", "Remark"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # ================= DATA ROWS =================
            thin = Side(border_style="thin", color="D1D5DB")
            data_start = header_row + 1
            for idx, r in enumerate(rows, start=data_start):
                remark, _ = self._late_remark_text(r)
                values = [
                    self._dash(r["attendance_date"]),
                    self._dash(r["check_in"]),
                    self._dash(r["check_out"]),
                    self._fmt_hours(self._ot_hours_from_row(r)),
                    remark,
                ]
                for col_num, value in enumerate(values, 1):
                    cell = ws.cell(row=idx, column=col_num)
                    cell.value = value
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

            # ================= COLUMN WIDTH =================
            widths = {1: 18, 2: 16, 3: 16, 4: 14, 5: 14}
            for col_num, width in widths.items():
                ws.column_dimensions[get_column_letter(col_num)].width = width

            # ================= ROW HEIGHT =================
            ws.row_dimensions[1].height = 30

            wb.save(file_path)
            messagebox.showinfo("Export Successful", "Attendance Excel exported successfully.")
        except Exception as e:
            messagebox.showerror("Export Error", f"Excel export failed: {e}")


