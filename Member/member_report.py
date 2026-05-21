import customtkinter as ctk
import calendar
import os
from datetime import datetime
from database import Database
from tkinter import messagebox, filedialog, ttk
from tkcalendar import Calendar
import mysql.connector
import csv

class DatePickerButton(ctk.CTkFrame):
    def __init__(self, master, initial_date=None):
        super().__init__(master, fg_color="transparent")
        self._date = initial_date or datetime.today().date()
        self._open = False
        
        # Calendar style setup
        self.update_calendar_style()
        
        self.btn = ctk.CTkButton(
            self,
            text=self._fmt(),
            width=140,
            height=36,
            corner_radius=8,
            hover_color=("#EBEBEB", "#2A2A2A"),
            border_width=1,
            fg_color=("#F9F9F9", "#1E1E1E"),
            border_color=("#DBDBDB", "#2C2C2C"),
            text_color=("black", "white"),
            anchor="w",
            command=self.toggle
        )
        self.btn.pack()
        
        self.panel = ctk.CTkFrame(
            self.winfo_toplevel(),
            fg_color=("#FFFFFF", "#141E2B"),
            corner_radius=12,
            border_width=1,
            border_color=("#DBDBDB", "#2A3A4A")
        )
        
        self.cal = Calendar(
            self.panel,
            style="Custom.Calendar",
            selectmode="day",
            date_pattern="yyyy-mm-dd",
            year=self._date.year,
            month=self._date.month,
            day=self._date.day,
            maxdate=datetime.today().date(),
            showweeknumbers=False,
            firstweekday="monday",
            font=("Arial", 11),
            cursor="hand2"
        )
        self.cal.pack(padx=8, pady=8)
        self.cal.bind("<<CalendarSelected>>", self._select)

    def update_calendar_style(self):
        # TTK styles are global and don't support tuples, 
        # so we refresh this specifically when the picker is toggled.
        is_dark = ctk.get_appearance_mode() == "Dark"
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Custom.Calendar",
            background="#1A1A2E" if is_dark else "#FFFFFF",
            foreground="white" if is_dark else "black",
            headersbackground="#16213E" if is_dark else "#F0F0F0",
            headersforeground="#4FC3F7" if is_dark else "#1f538d",
            selectbackground="#3498DB",
            selectforeground="white",
            normalbackground="#1A1A2E" if is_dark else "#FFFFFF",
            normalforeground="#CCCCCC" if is_dark else "#333333",
            weekendbackground="#1A1A2E" if is_dark else "#F9F9F9",
            weekendforeground="#F39C12",
            othermonthforeground="#555555" if is_dark else "#AAAAAA",
            bordercolor="#2A2A4A" if is_dark else "#DBDBDB",
            relief="flat"
        )

    def toggle(self):
        if self._open:
            self.panel.place_forget()
        else:
            self.update_calendar_style()
            self.panel.lift()
            self.panel.place(in_=self, x=0, y=self.btn.winfo_height() + 2)
        self._open = not self._open

    def _select(self, event):
        selected = self.cal.get_date()
        self._date = datetime.strptime(selected, "%Y-%m-%d").date()
        self.btn.configure(text=self._fmt())
        self.toggle()

    def get_date(self): return self._date

    def set_date(self, d):
        self._date = d
        self.cal.selection_set(d)
        self.btn.configure(text=self._fmt())

    def _fmt(self): return f"  📅  {self._date.strftime('%Y-%m-%d')}"

class MemberReportFrame(ctk.CTkFrame):
    def __init__(self, master, user, **kwargs):
        super().__init__(master, **kwargs)
        self.user = user
        self.db = Database()
        self.ensure_report_columns()
        self.configure(fg_color="transparent")
        
        self.now = datetime.now()
        self.report_rows = [] 

        today = datetime.today().date()
        self.start_date = datetime(today.year, today.month, 1).date()
        self.end_date = today

        # Automatic Theme Constants (Tuples)
        self.COLOR_CARD_BG = ("#FFFFFF", "#1E1E1E")
        self.COLOR_BORDER = ("#DBDBDB", "#2C2C2C")
        self.COLOR_TEXT_MAIN = ("#1A1A1A", "#E8EDF2")
        self.COLOR_TEXT_SEC = ("#555555", "#AAB7C4")
        self.COLOR_TEXT_TER = ("#777777", "#718096")
        self.COLOR_SCROLL_BG = ("#F5F5F5", "#1A1A1A")
        self.COLOR_CONTAINER_BG = ("#F0F0F0", "#252525")

        self.show_history_view()

    def _show_message(self, message, message_type="info", duration=3000):
        if message_type == "error":
            bg_color = "#E74C3C"
        elif message_type == "warning":
            bg_color = "#F39C12"
        elif message_type == "success":
            bg_color = "#27AE60"
        else:
            bg_color = "#3498DB"

        message_frame = ctk.CTkFrame(
            self.winfo_toplevel(),
            fg_color=bg_color,
            corner_radius=8
        )
        message_frame.place(relx=1.0, rely=0, x=-20, y=20, anchor="ne")

        ctk.CTkLabel(
            message_frame,
            text=message,
            text_color="white",
            font=("Arial", 12, "bold"),
            wraplength=250
        ).pack(padx=15, pady=10)

        self.after(duration, message_frame.destroy)

    def clear_view(self):
        for widget in self.winfo_children(): widget.destroy()

    def ensure_report_columns(self):
        try:
            self.db.cursor.execute("SELECT today_work, tomorrow_work, problems_issues, shared_matters FROM daily_reports LIMIT 1")
        except mysql.connector.Error as e:
            msg = str(e).lower()
            if "unknown column" in msg:
                try:
                    self.db.cursor.execute(
                        "ALTER TABLE daily_reports "
                        "ADD COLUMN IF NOT EXISTS today_work TEXT, "
                        "ADD COLUMN IF NOT EXISTS tomorrow_work TEXT, "
                        "ADD COLUMN IF NOT EXISTS problems_issues TEXT, "
                        "ADD COLUMN IF NOT EXISTS shared_matters TEXT"
                    )
                    self.db.conn.commit()
                except Exception as alter_err:
                    print("Schema migration error:", alter_err)
            else:
                print("Report schema check error:", e)

    def get_db_categories(self):
        try:
            self.db.cursor.execute("SELECT name FROM report_categories ORDER BY name ASC")
            results = self.db.cursor.fetchall()
            return [row['name'] for row in results] if results else ["General"]
        except: return ["General"]

    def refresh_view(self):
        for w in self.scroll.winfo_children(): w.destroy()
        start, end = self.start_cal.get_date(), self.end_cal.get_date()
        
        if start > end:
            self._show_message("Start date cannot be later than End date.", "error")
            self.count_lbl.configure(text="(0 records)")
            return

        try:
            query = """SELECT report_date, today_work, tomorrow_work, problems_issues, shared_matters
                       FROM daily_reports
                       WHERE user_id = %s AND report_date BETWEEN %s AND %s
                       ORDER BY report_date DESC, created_at ASC"""
            self.db.cursor.execute(query, (self.user['id'], start, end))
            reports = self.db.cursor.fetchall()
        except Exception as e:
            self._show_message(str(e, "error"))
            return

        grouped = {}
        for r in reports:
            date_key = str(r['report_date'])
            if date_key not in grouped:
                grouped[date_key] = r

        self.count_lbl.configure(text=f"({len(grouped)} records)")
        if not grouped:
            ctk.CTkLabel(
                self.scroll,
                text="No records found.",
                font=("Arial", 15, "bold"),
                text_color=self.COLOR_TEXT_TER
            ).pack(
                pady=20
            )

            return

        for date, row in sorted(grouped.items(), reverse=True):
            card = ctk.CTkFrame(self.scroll, corner_radius=12, fg_color=self.COLOR_CARD_BG,
                                border_width=1, border_color=self.COLOR_BORDER)
            card.pack(fill="x", pady=6, padx=10)

            info = ctk.CTkFrame(card, fg_color="transparent")
            info.pack(side="left", fill="both", expand=True, padx=14, pady=10)

            ctk.CTkLabel(info, text=date, font=("Arial", 13, "bold"), text_color=self.COLOR_TEXT_MAIN).pack(anchor="w")
            preview = row.get('today_work', "") or "No details yet."

            lines = preview.splitlines()

            if len(lines) > 2:
                preview = "\n".join(lines[:2]) + " ..."
            elif len(preview) > 120:
                preview = preview[:120] + "..."

            ctk.CTkLabel(
                info,
                text=preview,
                wraplength=650,
                justify="left",
                anchor="w",
                text_color=self.COLOR_TEXT_SEC
            ).pack(anchor="w", fill="x", pady=(4, 0))

            btn_frame = ctk.CTkFrame(card, fg_color="transparent")
            btn_frame.pack(side="right", padx=12, pady=10)

            def _btn(text, color, hover, cmd):
                return ctk.CTkButton(
                    btn_frame,
                    text=text,
                    width=60,
                    height=30,
                    corner_radius=8,
                    font=("Arial", 12, "bold"),
                    fg_color=color,
                    hover_color=hover,
                    command=cmd
                )

            _btn("View", "#1f538d", "#174a7a", lambda d=date: (self.save_filter_state(), self.show_detail_view(d))).pack(side="left", padx=4)
            _btn("Edit", "#F39C12", "#D68910",lambda d=date: (self.save_filter_state(), self.edit_report(d))).pack(side="left", padx=4)
            _btn("Delete", "#E74C3C", "#C0392B", lambda d=date: self.delete_report(d)).pack(side="left", padx=4)

    def show_history_view(self):  
        self.clear_view()
        top = ctk.CTkFrame(self, fg_color="transparent")
        top.pack(fill="x", padx=80, pady=(10, 20))

        ctk.CTkLabel(top, text="Daily Reports", font=("Arial", 20, "bold"), text_color=("black", "white")).pack(side="left")
        ctk.CTkButton(top, text="+ New Report", width=120, height=36, corner_radius=8, fg_color="#10B981", command=self.show_form_view).pack(side="right")

        filter_card = ctk.CTkFrame(self, corner_radius=14, fg_color=self.COLOR_CARD_BG, border_width=1, border_color=self.COLOR_BORDER)
        filter_card.pack(fill="x", padx=80, pady=(4, 10))

        inner = ctk.CTkFrame(filter_card, fg_color="transparent")
        inner.pack(fill="x", padx=18, pady=14)

        ctk.CTkLabel(inner, text="From", text_color=("black", "white"), font=("Arial", 11, "bold")).pack(side="left", padx=(0, 4))
        self.start_cal = DatePickerButton(inner, initial_date=self.start_date)
        self.start_cal.pack(side="left", padx=(0, 16))

        ctk.CTkLabel(inner, text="To", text_color=("black", "white"), font=("Arial", 11, "bold")).pack(side="left", padx=(0, 4))
        self.end_cal = DatePickerButton(inner, initial_date=self.end_date)
        self.end_cal.pack(side="left", padx=(0, 24))

        ctk.CTkFrame(inner, width=1, height=32, fg_color=("#DBDBDB", "#2A3A4A")).pack(side="left", padx=(0, 20))

        def _btn(parent, text, color, hover, cmd):
            ctk.CTkButton(parent, text=text, width=60, height=36, corner_radius=8, font=("Arial", 12, "bold"),
                         fg_color=color, hover_color=hover, command=cmd).pack(side="left", padx=4)

        _btn(inner, "🔍 Filter", "#2471A3", "#1A5276", self.refresh_view)
        _btn(inner, "✖ Clear", "#566573", "#424949", self.clear_filters)
        _btn(inner, "📄 PDF", "#C0392B", "#922B21", self.export_pdf_reports)
        _btn(inner, "📥 Excel", "#16A085", "#117A65", self.export_to_csv)

        list_header = ctk.CTkFrame(self, fg_color="transparent")
        list_header.pack(fill="x", padx=80, pady=(4, 2))
        ctk.CTkLabel(list_header, text="Report List", font=("Arial", 14, "bold"), text_color=("black", "white")).pack(side="left", padx=4)
        self.count_lbl = ctk.CTkLabel(list_header, text="", font=("Arial", 12), text_color="#566573")
        self.count_lbl.pack(side="left", padx=8)

        self.scroll = ctk.CTkScrollableFrame(self, fg_color=self.COLOR_SCROLL_BG, corner_radius=14, border_width=1, border_color=self.COLOR_BORDER)
        self.scroll.pack(fill="both", expand=True, padx=80, pady=(0, 10))
        self.refresh_view()

    def show_form_view(self):
        self.clear_view()
        self.report_rows = []
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=80, pady=(10, 10))

        # --- Row 1: Back button ---
        top_bar = ctk.CTkFrame(header, fg_color="transparent")
        top_bar.pack(fill="x", padx=10)

        ctk.CTkButton(
            top_bar,
            text="← Back",
            width=80,
            height=36,
            fg_color=("#DBDBDB", "#333333"),
            text_color=("black", "white"),
            command=self.show_history_view
        ).pack(side="left")


        # --- Row 2: Date button (below back) ---
        date_row = ctk.CTkFrame(header, fg_color="transparent")
        date_row.pack(fill="x", padx=10, pady=(6, 4))

        self.date_picker = DatePickerButton(
            date_row,
            initial_date=datetime.today().date()
        )
        self.date_picker.pack(side="left")

        self.form_scroll = ctk.CTkScrollableFrame(self, fg_color="transparent", height=380)
        self.form_scroll.pack(fill="both", expand=True, padx=80, pady=10)

        self.report_fields = {}
        self._render_report_form()

        footer = ctk.CTkFrame(self, fg_color="transparent")
        footer.pack(fill="x", padx=80, pady=(0, 20))
        self.save_btn = ctk.CTkButton(
            footer,
            text="Save",
            width=60,
            height=30,
            corner_radius=8,
            font=("Arial", 12, "bold"),
            fg_color="#1f538d",
            hover_color="#174a7a",
            command=self.save_all_reports
        )
        self.save_btn.pack(side="right", padx=10)

    def _render_report_form(self, data=None):
        self.report_fields.clear()

        def add_section(title, key, placeholder=""):
            section = ctk.CTkFrame(self.form_scroll, fg_color=self.COLOR_CONTAINER_BG, corner_radius=10)
            section.pack(fill="x", pady=8, padx=5)
            ctk.CTkLabel(section, text=title, font=("Arial", 14, "bold"), text_color=self.COLOR_TEXT_MAIN).pack(anchor="w", padx=15, pady=(12, 6))
            text_box = ctk.CTkTextbox(section, height=120, fg_color=("#FFFFFF", "#1e1e26"), border_width=1, border_color=("#DBDBDB", "#333333"), text_color=("black", "white"))
            text_box.pack(fill="both", expand=True, padx=15, pady=(0, 15))
            if data and data.get(key):
                text_box.insert("1.0", data.get(key, ""))
            self.report_fields[key] = text_box

        add_section("Today's Work Detail", "today_work")
        add_section("Tomorrow's Work Schedule", "tomorrow_work")
        add_section("Problems / Issues", "problems_issues")
        add_section("Shared Matters", "shared_matters")

    def save_all_reports(self):
        try:
            selected_date = str(self.date_picker.get_date())

            self.db.cursor.execute(
                "SELECT COUNT(*) as count FROM daily_reports WHERE user_id=%s AND report_date=%s",
                (self.user['id'], selected_date)
            )

            if self.db.cursor.fetchone()['count'] > 0:
                self._show_message(f"Report for {selected_date} already exists!", "error")
                return

            today_work = self.report_fields['today_work'].get("1.0", "end-1c").strip()
            tomorrow_work = self.report_fields['tomorrow_work'].get("1.0", "end-1c").strip()
            problems_issues = self.report_fields['problems_issues'].get("1.0", "end-1c").strip()
            shared_matters = self.report_fields['shared_matters'].get("1.0", "end-1c").strip()

            if not today_work:
                self._show_message("Please enter today's work detail.", "warning")
                return

            self.db.cursor.execute(
                """
                INSERT INTO daily_reports
                (user_id, report_date, today_work, tomorrow_work, problems_issues, shared_matters)
                VALUES (%s, %s, %s, %s, %s, %s)
                """,
                (
                    self.user['id'],
                    selected_date,
                    today_work,
                    tomorrow_work,
                    problems_issues,
                    shared_matters
                )
            )

            self.db.conn.commit()
            self._show_message("Report added successfully.", "success")
            self.show_history_view()

        except Exception as e:
            self.db.conn.rollback()
            self._show_message(f"Error: {e}", "error")

    def show_detail_view(self, date):
        self.clear_view()

        top = ctk.CTkFrame(self, fg_color="transparent")
        top.pack(fill="x", padx=80, pady=20)

        ctk.CTkButton(
            top,
            text="← Back",
            width=80,
            height=36,
            fg_color=("#DBDBDB", "#333333"),
            text_color=("black", "white"),
            hover_color=("#CFCFCF", "#444444"),
            corner_radius=8,
            command=self.show_history_view
        ).pack(side="left")

        ctk.CTkLabel(
            top,
            text=f"Details - {date}",
            font=("Arial", 20, "bold"),
            text_color=("black", "white")
        ).pack(side="left", padx=20)

        self.form_scroll = ctk.CTkScrollableFrame(
            self,
            fg_color="transparent"
        )
        self.form_scroll.pack(
            fill="both",
            expand=True,
            padx=80,
            pady=10
        )

        try:
            self.db.cursor.execute(
                """
                SELECT today_work, tomorrow_work,
                    problems_issues, shared_matters
                FROM daily_reports
                WHERE user_id=%s AND report_date=%s
                LIMIT 1
                """,
                (self.user['id'], date)
            )

            row = self.db.cursor.fetchone()

            if not row:
                ctk.CTkLabel(
                    self.form_scroll,
                    text="No report found.",
                    font=("Arial", 14),
                    text_color=self.COLOR_TEXT_TER
                ).pack(pady=20)
                return

            def add_readonly_section(title, content):

                content = content or "-"

                # Auto height based on actual content
                lines = content.split("\n")

                total_lines = 0

                for line in lines:
                    # estimate wrapped lines
                    wrapped = max(1, (len(line) // 85) + 1)
                    total_lines += wrapped

                # smaller minimum height
                textbox_height = max(45, total_lines * 24)

                section = ctk.CTkFrame(
                    self.form_scroll,
                    fg_color=self.COLOR_CONTAINER_BG,
                    corner_radius=10
                )
                section.pack(fill="x", pady=8, padx=5)

                ctk.CTkLabel(
                    section,
                    text=title,
                    font=("Arial", 14, "bold"),
                    text_color=self.COLOR_TEXT_MAIN
                ).pack(anchor="w", padx=15, pady=(12, 6))

                text_box = ctk.CTkTextbox(
                    section,
                    height=int(textbox_height),
                    fg_color=("#FFFFFF", "#1e1e26"),
                    border_width=1,
                    border_color=("#DBDBDB", "#333333"),
                    text_color=("black", "white"),
                    wrap="word"
                )

                text_box.pack(
                    fill="both",
                    expand=True,
                    padx=15,
                    pady=(0, 15)
                )

                text_box.insert("1.0", content)
                text_box.configure(state="disabled")

            add_readonly_section(
                "Today's Work Detail",
                row.get('today_work', "")
            )

            add_readonly_section(
                "Tomorrow's Work Schedule",
                row.get('tomorrow_work', "")
            )

            add_readonly_section(
                "Problems / Issues",
                row.get('problems_issues', "")
            )

            add_readonly_section(
                "Shared Matters",
                row.get('shared_matters', "")
            )

        except Exception as e:
            print("Detail error:", e)

    def edit_report(self, date):
        self.clear_view()
        top = ctk.CTkFrame(self, fg_color="transparent")
        top.pack(fill="x", padx=80, pady=20)
        ctk.CTkButton(
            top,
            text="← Back",
            width=80,
            height = 36,
            fg_color=("#DBDBDB", "#333333"),
            text_color=("black", "white"),
            hover_color=("#CFCFCF", "#444444"),
            corner_radius=8,
            command=self.show_history_view
        ).pack(side="left")
        ctk.CTkLabel(top, text=f"Edit Report - {date}", font=("Arial", 20, "bold"), text_color=("black", "white")).pack(side="left", padx=20)

        self.form_scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.form_scroll.pack(fill="both", expand=True, padx=80, pady=10)

        try:
            self.db.cursor.execute(
                "SELECT today_work, tomorrow_work, problems_issues, shared_matters "
                "FROM daily_reports WHERE user_id=%s AND report_date=%s LIMIT 1",
                (self.user['id'], date)
            )
            data = self.db.cursor.fetchone() or {}
            self.report_fields = {}
            self._render_report_form(data)
        except Exception as e:
            print("Edit load error:", e)

        footer = ctk.CTkFrame(self, fg_color="transparent")
        footer.pack(fill="x", padx=80, pady=(0, 20))
        ctk.CTkButton(
            footer,
            text="Update",
            width=60,
            height=30,
            corner_radius=8,
            font=("Arial", 12, "bold"),
            fg_color="#F39C12",
            hover_color="#D68910",
            command=lambda: self.update_report(date)
        ).pack(side="right", padx=10)

    def update_report(self, date):
        try:
            today_work = self.report_fields['today_work'].get("1.0", "end-1c").strip()
            tomorrow_work = self.report_fields['tomorrow_work'].get("1.0", "end-1c").strip()
            problems_issues = self.report_fields['problems_issues'].get("1.0", "end-1c").strip()
            shared_matters = self.report_fields['shared_matters'].get("1.0", "end-1c").strip()

            if not today_work:
                self._show_message("Please enter today's work detail.", "warning")
                return

            self.db.cursor.execute("DELETE FROM daily_reports WHERE user_id=%s AND report_date=%s", (self.user['id'], date))
            self.db.cursor.execute(
                """
                INSERT INTO daily_reports
                (user_id, report_date, today_work, tomorrow_work, problems_issues, shared_matters)
                VALUES (%s, %s, %s, %s, %s, %s)
                """,
                (
                    self.user['id'],
                    date,
                    today_work,
                    tomorrow_work,
                    problems_issues,
                    shared_matters
                )
            )
            self.db.conn.commit()
            self._show_message("Report updated successfully", "success")
            self.show_history_view()
        except Exception as e:
            self.db.conn.rollback()
            self._show_message(str(e, "error"))

    def delete_report(self, date):
        if messagebox.askyesno(
            "Confirm Delete",
            f"Are you sure you want to delete all reports for {date}?"
        ):
            try:
                self.db.cursor.execute(
                    "DELETE FROM daily_reports WHERE user_id=%s AND report_date=%s",
                    (self.user['id'], date)
                )

                self.db.conn.commit()

                # Show green success toast
                self._show_message(
                    "Report deleted successfully.",
                    "success"
                )

                # Refresh list after a short delay
                self.after(
                    500,
                    self.show_history_view
                )

            except Exception as e:
                self._show_message(
                    f"Delete failed: {e}",
                    "error"
                )

    def clear_filters(self):
        today = datetime.today().date()
        self.start_date = datetime(today.year, today.month, 1).date()
        self.end_date = today
        self.start_cal.set_date(self.start_date)
        self.end_cal.set_date(self.end_date)
        self.refresh_view()

    def export_pdf_reports(self):
        try:
            from fpdf import FPDF
            from tkinter import filedialog

            start = self.start_cal.get_date()
            end = self.end_cal.get_date()

            self.db.cursor.execute(
                """
                SELECT report_date, today_work, tomorrow_work,
                    problems_issues, shared_matters
                FROM daily_reports
                WHERE user_id = %s
                AND report_date BETWEEN %s AND %s
                ORDER BY report_date ASC, created_at ASC
                """,
                (self.user['id'], str(start), str(end))
            )

            rows = self.db.cursor.fetchall()

            if not rows:
                return self._show_message("No reports found.", "info")

            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[
                    ("PDF files", "*.pdf")
                ],
                initialfile=f"{self.user.get('full_name', 'User')}_Reports.pdf"
            )

            if not file_path:
                return

            if not os.path.splitext(file_path)[1]:
                file_path += ".pdf"

            if file_path.lower().endswith(".xlsx"):
                try:
                    from openpyxl import Workbook
                    from openpyxl.utils import get_column_letter
                    from openpyxl.styles import Alignment
                except ImportError:
                    self._show_message("openpyxl is required to export Excel files. Install it with:\n\npip install openpyxl", "error")
                    return

                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Reports"

                headers = [
                    "Date",
                    "Today's Work Detail",
                    "Tomorrow's Work Schedule",
                    "Problems/Issues",
                    "Shared Matters"
                ]
                sheet.append(headers)

                for r in rows:
                    values = [
                        r['report_date'],
                        r.get('today_work') or "-",
                        r.get('tomorrow_work') or "-",
                        r.get('problems_issues') or "-",
                        r.get('shared_matters') or "-"
                    ]
                    row_idx = sheet.max_row + 1
                    for col_idx, value in enumerate(values, start=1):
                        cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                        cell.alignment = Alignment(wrap_text=True, vertical='top')

                for col_idx in range(1, len(headers) + 1):
                    max_width = 0
                    for row_idx in range(1, sheet.max_row + 1):
                        cell_value = sheet.cell(row=row_idx, column=col_idx).value or ""
                        for line in str(cell_value).splitlines():
                            max_width = max(max_width, len(line))
                    sheet.column_dimensions[get_column_letter(col_idx)].width = max(max_width + 2, 15)

                workbook.save(file_path)
                self._show_message("Excel exported successfully!", "success")
                return

            class PDF(FPDF):

                def header(self):
                    # Main title
                    self.set_text_color(0, 0, 0)
                    self.set_font("Arial", "B", 20)
                    self.cell(
                        0,
                        12,
                        "Daily Report History",
                        ln=True,
                        align="C"
                    )

                    self.ln(4)

                    # Name + Date range
                    self.set_font("Arial", "B", 11)

                    # Name row
                    self.set_font("Arial", "B", 11)
                    self.cell(20, 8, "Name :", ln=0)

                    self.set_font("Arial", "", 11)
                    self.cell(
                        0,
                        8,
                        f"{getattr(self, 'user_name', 'User')}",
                        ln=True
                    )

                    # Date row
                    self.set_font("Arial", "B", 11)
                    self.cell(20, 8, "Date   :", ln=0)

                    self.set_font("Arial", "", 11)
                    self.cell(
                        0,
                        8,
                        f"{getattr(self, 'start_date', '')} ~ {getattr(self, 'end_date', '')}",
                        ln=True
                    )

                    self.ln(4)

                    # Smaller header font so titles fit
                    self.set_font("Arial", "B", 10)

                    headers = [
                        "Date",
                        "Today's Work Detail",
                        "Tomorrow's Work Schedule",
                        "Problems/Issues",
                        "Shared Matters"
                    ]

                    widths = self.col_widths
                    self.set_draw_color(169, 169, 169)
                    self.set_fill_color(0, 0, 139)
                    self.set_text_color(245, 245, 245)

                    for i, header in enumerate(headers):
                        self.cell(
                            widths[i],
                            12,
                            header,
                            border=1,
                            align="C",
                            fill=True
                        )

                    self.ln()
                    self.set_text_color(0, 0, 0)
                    self.set_fill_color(255, 255, 255)

            pdf = PDF("L", "mm", "A4")
            
            pdf.user_name = self.user.get('full_name', 'User')
            pdf.start_date = start
            pdf.end_date = end

            # Landscape widths
            pdf.col_widths = [32, 68, 68, 52, 52]

            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()

            # Body font
            body_font_size = 8
            pdf.set_font("Arial", "", body_font_size)

            line_height = 5

            def split_text(text, width):
                """
                Split text into lines based on column width.
                """
                if not text:
                    return ["-"]

                pdf.set_font("Arial", "", body_font_size)

                lines = []

                for paragraph in str(text).split("\n"):

                    words = paragraph.split(" ")
                    current_line = ""

                    for word in words:

                        test_line = (
                            current_line + " " + word
                        ).strip()

                        if pdf.get_string_width(test_line) <= width - 2:
                            current_line = test_line
                        else:
                            if current_line:
                                lines.append(current_line)
                            current_line = word

                    if current_line:
                        lines.append(current_line)

                return lines if lines else ["-"]

            for r in rows:

                cells = [
                    str(r['report_date']),
                    str(r.get('today_work') or "-"),
                    str(r.get('tomorrow_work') or "-"),
                    str(r.get('problems_issues') or "-"),
                    str(r.get('shared_matters') or "-")
                ]

                split_cells = []

                max_lines = 1

                for i, text in enumerate(cells):
                    lines = split_text(
                        text,
                        pdf.col_widths[i]
                    )

                    split_cells.append(lines)

                    if len(lines) > max_lines:
                        max_lines = len(lines)

                bottom_padding = 3
                row_height = (max_lines * line_height) + bottom_padding

                # Add new page if needed
                if pdf.get_y() + row_height > 190:
                    pdf.add_page()
                    pdf.set_font("Arial", "", body_font_size)

                x = pdf.get_x()
                y = pdf.get_y()
                pdf.set_draw_color(169, 169, 169)

                # Draw all borders first
                for i, width in enumerate(pdf.col_widths):
                    pdf.rect(
                        x,
                        y,
                        width,
                        row_height
                    )
                    x += width

                # Write text
                x = pdf.l_margin

                for i, lines in enumerate(split_cells):

                    pdf.set_xy(x + 2, y)

                    current_y = y + 2

                    for line in lines:
                        pdf.set_xy(x + 2, current_y)

                        pdf.cell(
                            pdf.col_widths[i] - 4,
                            line_height,
                            line,
                            border=0
                        )

                        current_y += line_height

                    x += pdf.col_widths[i]

                pdf.set_y(y + row_height)

            pdf.output(file_path)

            self._show_message("PDF exported successfully!", "success")

        except Exception as e:
            self._show_message(str(e, "error")
            )

    def export_to_csv(self):
        try:
            start = self.start_cal.get_date()
            end = self.end_cal.get_date()

            self.db.cursor.execute(
                """
                SELECT report_date, today_work, tomorrow_work,
                    problems_issues, shared_matters
                FROM daily_reports
                WHERE user_id = %s
                AND report_date BETWEEN %s AND %s
                ORDER BY report_date ASC, created_at ASC
                """,
                (self.user['id'], str(start), str(end))
            )

            rows = self.db.cursor.fetchall()
            if not rows:
                return self._show_message("No reports found.", "info")

            try:
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
                from openpyxl.styles import Alignment
            except ImportError:
                self._show_message("openpyxl is required to export Excel files. Install it with:\n\npip install openpyxl", "error")
                return

            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"{self.user.get('full_name', 'User')}_Reports.xlsx"
            )
            if not file_path:
                return
            base, ext = os.path.splitext(file_path)
            if ext.lower() != ".xlsx":
                file_path = base + ".xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Reports"

            # Title + user name + date (mimic PDF)
            ncols = 5
            title = "Daily Report History"
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            tcell = sheet.cell(row=1, column=1, value=title)
            tcell.font = tcell.font.copy(bold=True, size=18, color="000000")
            tcell.alignment = tcell.alignment.copy(horizontal='center', vertical='center')
            sheet.row_dimensions[1].height = 26

            # Name row
            name_text = f"Name : {getattr(self, 'user', {}).get('full_name', 'User')}"
            sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
            ncell = sheet.cell(row=2, column=1, value=name_text)
            ncell.font = ncell.font.copy(bold=False, size=11, color="000000")
            ncell.alignment = ncell.alignment.copy(horizontal='left', vertical='center')
            sheet.row_dimensions[2].height = 18

            # Date row
            date_text = f"Date   : {start} ~ {end}"
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
            dcell = sheet.cell(row=3, column=1, value=date_text)
            dcell.font = dcell.font.copy(bold=False, size=11, color="000000")
            dcell.alignment = dcell.alignment.copy(horizontal='left', vertical='center')
            sheet.row_dimensions[3].height = 18

            headers = [
                "Date",
                "Today's Work Detail",
                "Tomorrow's Work Schedule",
                "Problems/Issues",
                "Shared Matters"
            ]

            hdr_row = 5

            from openpyxl.styles import (
                PatternFill,
                Border,
                Side,
                Font,
                Alignment
            )

            header_fill = PatternFill(
                fill_type="solid",
                start_color="D9D9D9",
                end_color="D9D9D9"
            )

            side = Side(
                border_style="thin",
                color="000000"
            )

            for c, h in enumerate(headers, start=1):

                cell = sheet.cell(
                    row=hdr_row,
                    column=c,
                    value=h
                )

                cell.font = Font(
                    bold=True,
                    color="000000"
                )

                cell.fill = header_fill

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

                cell.border = Border(
                    left=side,
                    right=side,
                    top=side,
                    bottom=side
                )
                side = Side(border_style="thin", color="A9A9A9")
                cell.border = Border(left=side, right=side, top=side, bottom=side)

            for r in rows:
                values = [
                    r['report_date'],
                    r.get('today_work') or "-",
                    r.get('tomorrow_work') or "-",
                    r.get('problems_issues') or "-",
                    r.get('shared_matters') or "-"
                ]
                row_idx = sheet.max_row + 1
                for col_idx, value in enumerate(values, start=1):
                    cell = sheet.cell(
                        row=row_idx,
                        column=col_idx,
                        value=value
                    )

                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical='top'
                    )

                    side = Side(
                        border_style="thin",
                        color="A9A9A9"
                    )

                    cell.border = Border(
                        left=side,
                        right=side,
                        top=side,
                        bottom=side
                    )

            # Auto-size columns and dynamic row height
            widths = []
            for col_idx in range(1, ncols + 1):
                max_width = 0
                for row_idx in range(1, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value or ""
                    for line in str(cell_value).splitlines():
                        max_width = max(max_width, len(line))
                widths.append(max(max_width + 2, 15))
                sheet.column_dimensions[get_column_letter(col_idx)].width = widths[-1]

            def estimate_row_height(row_idx):
                max_lines = 1
                for col_idx, width in enumerate(widths, start=1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value or ""
                    lines = []
                    for paragraph in str(cell_value).splitlines():
                        current = ""
                        for word in paragraph.split(" "):
                            test_line = (current + " " + word).strip()
                            if len(test_line) <= width:
                                current = test_line
                            else:
                                if current:
                                    lines.append(current)
                                current = word
                        if current:
                            lines.append(current)
                    max_lines = max(max_lines, len(lines))
                return max(18, max_lines * 15)

            for row_idx in range(4, sheet.max_row + 1):
                sheet.row_dimensions[row_idx].height = estimate_row_height(row_idx)
            from openpyxl.styles import Border, Side

            side = Side(
                border_style="thin",
                color="A9A9A9"
            )

            for row in sheet.iter_rows(
                min_row=5,      # header row
                max_row=sheet.max_row,
                min_col=1,
                max_col=5       # Date -> Shared Matters
            ):
                for cell in row:
                    cell.border = Border(
                        left=side,
                        right=side,
                        top=side,
                        bottom=side
                    )

            workbook.save(file_path)
            self._show_message("Excel exported successfully!", "success")
        except Exception as e:
            self._show_message(str(e, "error"))

    def save_filter_state(self):
        if hasattr(self, "start_cal") and hasattr(self, "end_cal"):
            self.start_date = self.start_cal.get_date()
            self.end_date = self.end_cal.get_date()