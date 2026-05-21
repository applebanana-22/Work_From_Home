import customtkinter as ctk
from database import Database
from tkinter import messagebox, filedialog, ttk
from datetime import datetime
import tkinter as tk
import calendar
import os
from fpdf import FPDF
from tkcalendar import DateEntry, Calendar
from collections import defaultdict

# ── Reusable modern date-picker ──────────────────────────────────────────────
class DatePickerButton(ctk.CTkFrame):
    def __init__(self, master, initial_date=None):
        super().__init__(master, fg_color="transparent")

        self._date = initial_date or datetime.today().date()
        self._open = False

        # STYLE
        style = ttk.Style()
        style.theme_use("clam")

        # The Calendar widget is a standard TK widget, so we use a neutral dark-friendly theme 
        # that remains legible, but the wrapper around it is theme-aware.
        style.configure("Custom.Calendar",
            background="#1A1A2E",
            foreground="white",
            headersbackground="#16213E",
            headersforeground="#4FC3F7",
            selectbackground="#3498DB",
            selectforeground="white",
            normalbackground="#1A1A2E",
            normalforeground="#CCCCCC",
            weekendbackground="#1A1A2E",
            weekendforeground="#F39C12",
            othermonthforeground="#555555",
            bordercolor="#2A2A4A",
            relief="flat"
        )

        # BUTTON - Theme Aware (Light/Dark)
        self.btn = ctk.CTkButton(
            self,
            text=self._fmt(),
            width=140,
            height=36,
            corner_radius=8,
            hover_color=("#EBEBEB", "#2A2A2A"),
            border_width=1,
            fg_color=("#F9F9F9", "#1E1E1E"),
            border_color=("#DBDBDB", "#2C2C2C"),
            text_color=("black", "white"),
            anchor="w",
            command=self.toggle
        )
        self.btn.pack()

        # FLOATING PANEL - Theme Aware
        self.panel = ctk.CTkFrame(
            self.winfo_toplevel(),
            fg_color=("#FFFFFF", "#141E2B"),
            corner_radius=12,
            border_width=1,
            border_color=("#CCCCCC", "#2A3A4A")
        )

        # CALENDAR
        self.cal = Calendar(
            self.panel,
            style="Custom.Calendar",
            selectmode="day",
            date_pattern="yyyy-mm-dd",
            year=self._date.year,
            month=self._date.month,
            day=self._date.day,
            maxdate=datetime.today().date(),
            showweeknumbers=False,
            firstweekday="monday",
            font=("Arial", 11),
            cursor="hand2"
        )
        self.cal.pack(padx=8, pady=8)

        # AUTO SELECT
        self.cal.bind("<<CalendarSelected>>", self._select)

    def toggle(self):
        if self._open:
            self.panel.place_forget()
        else:
            self.panel.lift()
            self.panel.place(in_=self, x=0, y=self.btn.winfo_height() + 2)
        self._open = not self._open

    def _select(self, event):
        selected = self.cal.get_date()
        self._date = datetime.strptime(selected, "%Y-%m-%d").date()
        self.btn.configure(text=self._fmt())
        self.toggle()

    def get_date(self):
        return self._date

    def set_date(self, d):
        self._date = d
        self.cal.selection_set(d)
        self.btn.configure(text=self._fmt())

    def _fmt(self):
        return f"  📅  {self._date.strftime('%Y-%m-%d')}"

class LeaderReportView(ctk.CTkFrame):
    def __init__(self, master, user, **kwargs):
        super().__init__(master, **kwargs)
        self.db = Database()
        self.user = user
        self.configure(fg_color="transparent")

        # --- Filter State ---
        today = datetime.today().date()
        self.start_date = today
        self.end_date = today
        self.member_search = ""

        # Get team name for this user
        self.team_id = self.user.get('team_id')
        self.team_name = self.db.get_team_name(self.team_id) or "Unknown Team"

        # Main Container
        self.container = ctk.CTkFrame(self, fg_color="transparent")
        self.container.pack(fill="both", expand=True)

        self.show_reports_list()

    def _show_message(self, message, message_type="info", duration=3000):
        if message_type == "error":
            bg_color = "#E74C3C"
        elif message_type == "warning":
            bg_color = "#F39C12"
        elif message_type == "success":
            bg_color = "#27AE60"
        else:
            bg_color = "#3498DB"

        message_frame = ctk.CTkFrame(
            self.winfo_toplevel(),
            fg_color=bg_color,
            corner_radius=8
        )
        message_frame.place(relx=1.0, rely=0, x=-20, y=20, anchor="ne")

        ctk.CTkLabel(
            message_frame,
            text=message,
            text_color="white",
            font=("Arial", 12, "bold"),
            wraplength=250
        ).pack(padx=15, pady=10)

        self.after(duration, message_frame.destroy)

    def get_member_names(self):
        try:
            query = "SELECT full_name FROM users WHERE role = 'member' AND team_id = %s ORDER BY full_name ASC"
            self.db.cursor.execute(query, (self.team_id,))
            return [row['full_name'] for row in self.db.cursor.fetchall()]
        except Exception:
            return []

    def clear_container(self):
        for widget in self.container.winfo_children():
            widget.destroy()

    def show_reports_list(self):
        content_padx = 80
        self.clear_container()

        # Header Section
        header = ctk.CTkFrame(self.container, fg_color="transparent")
        header.pack(fill="x", padx=80, pady=14)

        # Left title
        ctk.CTkLabel(
            header,
            text=f"📊 {self.team_name} · Report View",
            font=("Arial", 20, "bold"),
            text_color=("#333333", "#FFFFFF")
        ).pack(side="left")

        # (counts UI moved below title, before filters)


        # Filter Card - Theme Aware Colors
        # ── Counts Bar ─────────────────────────────────────
        counts_bar = ctk.CTkFrame(
            self.container,
            corner_radius=14,
            fg_color=("#F9F9F9", "#1E1E1E"),
            border_width=1,
            border_color=("#E0E0E0", "#2C2C2C")
        )
        counts_bar.pack(fill="x", padx=content_padx, pady=(0, 10))

        # LEFT SIDE CONTAINER
        left_wrap = ctk.CTkFrame(counts_bar, fg_color="transparent")
        left_wrap.pack(anchor="w", padx=18, pady=14)

        # TOP ROW (DATE + FILTER BUTTON)
        top_row = ctk.CTkFrame(left_wrap, fg_color="transparent")
        top_row.pack(anchor="w", pady=(0, 6))

        # Date picker
        self.count_date = DatePickerButton(
            top_row,
            initial_date=datetime.today().date()
        )
        self.count_date.pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            top_row,
            text="🔍 Filter",
            width=60,
            height=36,
            corner_radius=8,
            fg_color="#2471A3",
            hover_color="#1A5276",
            font=("Arial", 12, "bold"),
            command=self.refresh_counts
        ).pack(side="left", padx=(0, 6))

        ctk.CTkButton(
            top_row,
            text="✖ Clear",
            width=60,
            height=36,
            corner_radius=8,
            fg_color="#566573",
            hover_color="#424949",
            font=("Arial", 12, "bold"),
            command=self.clear_counts
        ).pack(side="left")

        # SECOND ROW (SMALL CARDS)
        stats_wrap = ctk.CTkFrame(left_wrap, fg_color="transparent")
        stats_wrap.pack(anchor="w")

        def create_small_card(parent, title, color, command=None):

            card = ctk.CTkFrame(
                parent,
                width=130,
                height=72,
                corner_radius=15,
                fg_color=("#FFFFFF", "#2B2B2B"),
                border_width=1,
                border_color=("#DDDDDD", "#3A3A3A")
            )

            card.pack(side="left", padx=(0, 16))
            card.pack_propagate(False)
            card.grid_propagate(False)

            title_lbl = ctk.CTkLabel(
                card,
                text=title,
                font=("Arial", 13),
                text_color=("#666666", "#888888")
            )

            title_lbl.pack(
                pady=(10, 0)
            )

            value_lbl = ctk.CTkLabel(
                card,
                text="0",
                font=("Arial", 20, "bold"),
                text_color=color,
                cursor="hand2" if command else ""
            )

            value_lbl.pack(
                pady=(2, 0)
            )

            if command:
                value_lbl.bind("<Button-1>", lambda e: command())
                title_lbl.bind("<Button-1>", lambda e: command())
                card.bind("<Button-1>", lambda e: command())

            return value_lbl

        # Cards
        self.total_lbl = create_small_card(
            stats_wrap,
            "Total Members",
            "#3498DB"
        )

        self.submitted_btn = create_small_card(
            stats_wrap,
            "Submitted",
            "#10B981",
            lambda: self.show_member_list(True)
        )

        self.not_submitted_btn = create_small_card(
            stats_wrap,
            "Not Submitted",
            "#E74C3C",
            lambda: self.show_member_list(False)
        )

        # INITIAL LOAD ONLY
        self.refresh_counts()

        # Filter Card - Theme Aware Colors
        filter_card = ctk.CTkFrame(
            self.container,
            corner_radius=14,
            fg_color=("#F9F9F9", "#1E1E1E"),
            border_width=1,
            border_color=("#E0E0E0", "#2C2C2C")
        )
        filter_card.pack(fill="x", padx=content_padx, pady=(4, 10))
        inner = ctk.CTkFrame(filter_card, fg_color="transparent")
        inner.pack(fill="x", padx=18, pady=14)

        def _lbl(parent, text):
            ctk.CTkLabel(parent, text=text,
                         font=("Arial", 11, "bold"),
                         text_color=("#555555", "#FFFFFF")).pack(side="left", padx=(0, 4))

        # Date pickers
        _lbl(inner, "From")
        self.start_cal = DatePickerButton(inner, initial_date=self.start_date)
        self.start_cal.pack(side="left", padx=(0, 16))
        _lbl(inner, "To")
        self.end_cal = DatePickerButton(inner, initial_date=self.end_date)
        self.end_cal.pack(side="left", padx=(0, 24))

        # Separator
        sep = ctk.CTkFrame(inner, width=1, height=32, fg_color=("#CCCCCC", "#2A3A4A"))
        sep.pack(side="left", padx=(0, 20))

        # Member search
        _lbl(inner, "Member")
        self.member_entry = ctk.CTkEntry(
            inner, placeholder_text="Search name…",
            width=100, height=36, corner_radius=8,
            border_color=("#CCCCCC", "#2C2C2C"), border_width=1,
            fg_color=("#FFFFFF", "#1E1E1E"),
            text_color=("#000000", "#FFFFFF"),
            font=("Arial", 12)
        )
        self.member_entry.pack(side="left", padx=(0, 8))

        # Action buttons
        def _btn(parent, text, color, hover, cmd, width=75):
            b = ctk.CTkButton(
                parent, text=text, width=width, height=36,
                corner_radius=8, font=("Arial", 12, "bold"),
                fg_color=color, hover_color=hover, command=cmd
            )
            b.pack(side="left", padx=4)
            return b

        _btn(inner, "🔍 Filter", "#2471A3", "#1A5276", self.refresh_view, width=60)
        _btn(inner, "✖ Clear", "#566573", "#424949", self.clear_filters, width=60)
        _btn(inner, "📄 PDF", "#C0392B", "#922B21", self.export_to_pdf, width=60)
        _btn(inner, "📥 Excel", "#16A085", "#117A65", self.export_to_csv, width=60)
        
        # List header
        list_header = ctk.CTkFrame(self.container, fg_color="transparent")
        list_header.pack(fill="x", padx=content_padx, pady=(4, 2))
        ctk.CTkLabel(list_header, text="Report List", font=("Arial", 14, "bold"), text_color=("#333333", "#FFFFFF")).pack(side="left", padx=4)
        self.count_lbl = ctk.CTkLabel(list_header, text="", font=("Arial", 12), text_color="#566573")
        self.count_lbl.pack(side="left", padx=8)

        self.scroll = ctk.CTkScrollableFrame(
            self.container,
            fg_color=("#EEEEEE", "#1A1A1A"),
            corner_radius=14,
            border_width=1,
            border_color=("#DDDDDD", "#2C2C2C")
        )
        self.scroll.pack(fill="both", expand=True, padx=content_padx, pady=(0, 10))

        self.refresh_view()
        # initial counts for today
        self.refresh_counts()

    def refresh_view(self):
        for w in self.scroll.winfo_children():
            w.destroy()

        team_id = self.team_id
        start = self.start_cal.get_date()
        end = self.end_cal.get_date()
        if start > end:
            self._show_message("Start date cannot be later than End date.", "error")
            return
        member_search = self.member_entry.get().strip()
        today_date = datetime.today().date()

        query = '''SELECT u.full_name, dr.report_date, dr.today_work, dr.tomorrow_work, dr.problems_issues, dr.shared_matters
                FROM users u
                JOIN daily_reports dr ON u.id = dr.user_id
                WHERE u.team_id=%s AND dr.report_date BETWEEN %s AND %s AND u.role='member' '''
        params = [team_id, start, end]
        if member_search:
            query += " AND u.full_name LIKE %s"
            params.append(f"%{member_search}%")

        query += " ORDER BY dr.report_date DESC, u.full_name ASC, dr.created_at DESC"

        try:
            self.db.cursor.execute(query, tuple(params))
            rows = self.db.cursor.fetchall()
            
            grouped = defaultdict(list)
            for r in rows:
                key = (r['report_date'], r['full_name'])
                grouped[key].append(r)
                
            self.count_lbl.configure(text=f"({len(grouped)} records)")
            
            if not grouped:
                ctk.CTkLabel(
                    self.scroll,
                    text="No records found.",
                    font=("Arial", 15, "bold"),
                    text_color=("#666666", "#AAAAAA")
                ).pack(
                    pady=20
                )
                return

            last_date = None
            for (date_val, member), tasks in grouped.items():
                is_today = date_val == today_date
                if last_date != date_val:
                    date_sep = ctk.CTkFrame(self.scroll, fg_color="transparent")
                    date_sep.pack(fill="x", padx=8, pady=(8, 1))
                    day_str = date_val.strftime("%A")
                    date_str = date_val.strftime("%d %b %Y")
                    badge_color = "#E67E22" if is_today else ("#BDC3C7", "#2C3E50")
                    badge_text = f"  {date_str}  ·  {day_str}{'  ← TODAY' if is_today else ''}  "
                    ctk.CTkLabel(
                        date_sep, text=badge_text,
                        font=("Arial", 11, "bold"),
                        text_color=("black", "white"),
                        fg_color=badge_color, corner_radius=6,
                        padx=6, pady=2
                    ).pack(side="left")
                    last_date = date_val

                card = ctk.CTkFrame(self.scroll,
                    corner_radius=12,
                    fg_color=("#FFFFFF", "#1E1E1E"),
                    border_width=1,
                    border_color=("#E0E0E0", "#2C2C2C")
                )
                card.pack(fill="x", pady=1, padx=10)
                info = ctk.CTkFrame(card, fg_color="transparent")
                info.pack(side="left", fill="both", expand=True, padx=10, pady=6)

                ctk.CTkLabel(info, text=member, font=("Arial", 12, "bold"), text_color=("#333333", "#E8EDF2")).pack(anchor="w")
                # Show preview like member report
                preview = tasks[0].get('today_work', "") if tasks else ""
                preview = preview or "No details yet."

                lines = preview.splitlines()

                # More than 2 lines -> show ...
                if len(lines) > 2:
                    preview = "\n".join(lines[:2]) + " ..."
                elif len(preview) > 120:
                    preview = preview[:120] + "..."

                ctk.CTkLabel(
                    info,
                    text=preview,
                    wraplength=650,
                    justify="left",
                    anchor="w",
                    text_color=("#555555", "#AAB7C4")
                ).pack(anchor="w", fill="x", pady=(4, 0))

                btn_frame = ctk.CTkFrame(card, fg_color="transparent")
                btn_frame.pack(side="right", padx=12, pady=6)

                ctk.CTkButton(
                    btn_frame,
                    text="View",
                    width=60,
                    height=30,
                    corner_radius=8,
                    font=("Arial", 12, "bold"),
                    fg_color="#1f538d",
                    hover_color="#174a7a",
                    command=lambda tlist=tasks, m=member, d=date_val:
                    self.open_detail_with_state(m, d, tlist)
                ).pack()
        except Exception as e:
            self._show_message(str(e, "error"))
    
    def refresh_counts(self):
        try:
            sel_date = self.count_date.get_date()

            # TOTAL MEMBERS
            q_total = """
                SELECT COUNT(*) AS cnt
                FROM users
                WHERE role='member'
                AND team_id=%s
            """

            self.db.cursor.execute(
                q_total,
                (self.team_id,)
            )

            total = self.db.cursor.fetchone()['cnt'] or 0

            # SUBMITTED MEMBERS
            q_submitted = """
                SELECT COUNT(DISTINCT dr.user_id) AS cnt
                FROM daily_reports dr
                JOIN users u ON u.id = dr.user_id
                WHERE u.team_id=%s
                AND u.role='member'
                AND dr.report_date=%s
            """

            self.db.cursor.execute(
                q_submitted,
                (self.team_id, sel_date)
            )

            submitted = self.db.cursor.fetchone()['cnt'] or 0

            # NOT SUBMITTED
            not_submitted = total - submitted

            # UPDATE UI
            self.total_lbl.configure(
                text=str(total)
            )

            self.submitted_btn.configure(
                text=str(submitted)
            )

            self.not_submitted_btn.configure(
                text=str(not_submitted)
            )

        except Exception as e:
            print("Count Error:", e)
            
    def clear_counts(self):
        today = datetime.today().date()

        try:
            if hasattr(self, 'count_date'):
                self.count_date.set_date(today)
        except Exception:
            pass

        self.refresh_counts()
        
    def show_member_list(self, submitted=True):
        """Render member list for the selected date in the current view (no popup)."""
        try:
            sel_date = self.count_date.get_date()
        except Exception:
            sel_date = datetime.today().date()

        if submitted:
            q = '''SELECT DISTINCT u.full_name FROM users u
                   JOIN daily_reports dr ON u.id=dr.user_id
                   WHERE u.team_id=%s AND dr.report_date=%s AND u.role='member'
                   ORDER BY u.full_name ASC'''
            params = (self.team_id, sel_date)
            title = f"Submitted on {sel_date}"
        else:
            q = '''SELECT u.full_name FROM users u
                   WHERE u.team_id=%s AND u.role='member' AND u.id NOT IN
                   (SELECT user_id FROM daily_reports WHERE report_date=%s)
                   ORDER BY u.full_name ASC'''
            params = (self.team_id, sel_date)
            title = f"Not Submitted on {sel_date}"

        try:
            self.db.cursor.execute(q, params)
            rows = self.db.cursor.fetchall()
            names = [r['full_name'] for r in rows] if rows else []

            # build view inside main container
            self.clear_container()

            content_padx = 80

            top = ctk.CTkFrame(self.container, fg_color="transparent")
            top.pack(fill="x", padx=content_padx, pady=20)

            ctk.CTkButton(
                top,
                text="← Back",
                width=80,
                height = 36,
                fg_color=("#DBDBDB", "#333333"),
                text_color=("black", "white"),
                hover_color=("#CFCFCF", "#444444"),
                corner_radius=8,
                command=self.back_to_list
            ).pack(side="left")

            ctk.CTkLabel(
                top,
                text=title,
                font=("Arial", 20, "bold"),
                text_color=("black", "white")
            ).pack(side="left", padx=20)

            list_frame = ctk.CTkScrollableFrame(self.container, fg_color="transparent")
            list_frame.pack(fill="both", expand=True, padx=content_padx, pady=10)

            if not names:
                ctk.CTkLabel(list_frame, text="(no members)", font=("Arial", 12)).pack(anchor="w", padx=8, pady=6)
            else:
                for i, n in enumerate(names, start=1):
                    ctk.CTkLabel(list_frame, text=f"{i}. {n}", font=("Arial", 12)).pack(anchor="w", padx=8, pady=4)

            # ensure counts reflect current selection
            self.refresh_counts()
        except Exception as e:
            self._show_message(str(e, "error"))
            
    def open_detail_with_state(self, member, date_val, tasks):
        self.start_date = self.start_cal.get_date()
        self.end_date = self.end_cal.get_date()
        self.member_search = self.member_entry.get()
        self.show_member_detail(member, date_val, tasks)

    def show_member_list(self, submitted=True):

        """Render submitted / not submitted member list."""

        try:
            sel_date = self.count_date.get_date()
        except Exception:
            sel_date = datetime.today().date()

        if submitted:

            q = '''
                SELECT DISTINCT u.full_name
                FROM users u
                JOIN daily_reports dr ON u.id = dr.user_id
                WHERE u.team_id=%s
                AND dr.report_date=%s
                AND u.role='member'
                ORDER BY u.full_name ASC
            '''

            params = (self.team_id, sel_date)

            title = f"Members who submitted on {sel_date}"

        else:

            q = '''
                SELECT u.full_name
                FROM users u
                WHERE u.team_id=%s
                AND u.role='member'
                AND u.id NOT IN
                (
                    SELECT user_id
                    FROM daily_reports
                    WHERE report_date=%s
                )
                ORDER BY u.full_name ASC
            '''

            params = (self.team_id, sel_date)

            title = f"Members who not submitted on {sel_date}"

        try:

            self.db.cursor.execute(q, params)

            rows = self.db.cursor.fetchall()

            names = [r['full_name'] for r in rows] if rows else []

            self.clear_container()

            content_padx = 80

            # TOP BAR
            top = ctk.CTkFrame(
                self.container,
                fg_color="transparent"
            )

            top.pack(
                fill="x",
                padx=content_padx,
                pady=20
            )

            # BACK BUTTON (same as admin)
            ctk.CTkButton(
                top,
                text="← Back",
                width=80,
                height = 36,
                fg_color=("#DBDBDB", "#333333"),
                text_color=("black", "white"),
                hover_color=("#CFCFCF", "#444444"),
                corner_radius=8,
                command=self.back_to_list
            ).pack(side="left")

            # TITLE (same style as admin)
            ctk.CTkLabel(
                top,
                text=title,
                font=("Arial", 20, "bold"),
                text_color=("black", "white")
            ).pack(side="left", padx=20)

            # MAIN CARD
            main_card = ctk.CTkFrame(
                self.container,
                corner_radius=14,
                fg_color=("#F0F0F0", "#252525"),
                border_width=1,
                border_color=("#DDDDDD", "#333333")
            )

            main_card.pack(
                fill="both",
                expand=True,
                padx=content_padx,
                pady=(0, 20)
            )

            # INNER LIST BOX
            list_frame = ctk.CTkScrollableFrame(
                main_card,
                fg_color=("#FFFFFF", "#1E1E26"),
                corner_radius=12,
                border_width=1,
                border_color=("#DDDDDD", "#333333")
            )

            list_frame.pack(
                fill="both",
                expand=True,
                padx=20,
                pady=20
            )

            if not names:

                ctk.CTkLabel(
                    list_frame,
                    text="No member found.",
                    font=("Arial", 12)
                ).pack(
                    anchor="w",
                    padx=8,
                    pady=6
                )

            else:

                for i, n in enumerate(names, start=1):

                    ctk.CTkLabel(
                        list_frame,
                        text=f"{i}. {n}",
                        font=("Arial", 12),
                        text_color=("black", "white")
                    ).pack(
                        anchor="w",
                        padx=8,
                        pady=4
                    )

        except Exception as e:

            self._show_message(str(e, "error")
            )
            
    def back_to_list(self):
        self.show_reports_list()
        self.start_cal.set_date(self.start_date)
        self.end_cal.set_date(self.end_date)
        self.member_entry.delete(0, "end")
        self.member_entry.insert(0, self.member_search)
        self.refresh_view()

    def clear_filters(self):
        today = datetime.today().date()

        # Reset ONLY report-list filters
        self.start_cal.set_date(today)
        self.end_cal.set_date(today)

        self.member_entry.delete(0, "end")

        # Refresh only the report list
        self.refresh_view()
        
    def open_detail_with_state(self, member, date_val, tasks):
        self.start_date = self.start_cal.get_date()
        self.end_date = self.end_cal.get_date()
        self.member_search = self.member_entry.get()
        self.show_member_detail(member, date_val, tasks)
        
    def show_member_detail(self, member_name, date_val, tasks):

        self.clear_container()

        content_padx = 80

        top = ctk.CTkFrame(
            self.container,
            fg_color="transparent"
        )

        top.pack(
            fill="x",
            padx=content_padx,
            pady=20
        )

        ctk.CTkButton(
            top,
            text="← Back",
            width=80,
            fg_color=("#DBDBDB", "#333333"),
            text_color=("black", "white"),
            hover_color=("#CFCFCF", "#444444"),
            corner_radius=8,
            command=self.back_to_list
        ).pack(side="left")

        ctk.CTkLabel(
            top,
            text=f"{member_name} - {date_val}",
            font=("Arial", 20, "bold"),
            text_color=("black", "white")
        ).pack(side="left", padx=20)

        self.form_scroll = ctk.CTkScrollableFrame(
            self.container,
            fg_color="transparent"
        )

        self.form_scroll.pack(
            fill="both",
            expand=True,
            padx=content_padx,
            pady=10
        )

        COLOR_CONTAINER_BG = ("#F0F0F0", "#252525")
        COLOR_TEXT_MAIN = ("#1A1A1A", "#E8EDF2")

        for r in tasks:

            def add_readonly_section(title, content):

                content = content or "-"

                lines = content.split("\n")

                total_lines = 0

                for line in lines:
                    wrapped = max(1, (len(line) // 85) + 1)
                    total_lines += wrapped

                textbox_height = max(45, total_lines * 24)

                section = ctk.CTkFrame(
                    self.form_scroll,
                    fg_color=COLOR_CONTAINER_BG,
                    corner_radius=10
                )

                section.pack(fill="x", pady=8, padx=5)

                ctk.CTkLabel(
                    section,
                    text=title,
                    font=("Arial", 14, "bold"),
                    text_color=COLOR_TEXT_MAIN
                ).pack(anchor="w", padx=15, pady=(12, 6))

                text_box = ctk.CTkTextbox(
                    section,
                    height=int(textbox_height),
                    fg_color=("#FFFFFF", "#1e1e26"),
                    border_width=1,
                    border_color=("#DBDBDB", "#333333"),
                    text_color=("black", "white"),
                    wrap="word"
                )

                text_box.pack(
                    fill="both",
                    expand=True,
                    padx=15,
                    pady=(0, 15)
                )

                text_box.insert("1.0", content)

                text_box.configure(state="disabled")

            add_readonly_section(
                "Today's Work Detail",
                r.get('today_work', "")
            )

            add_readonly_section(
                "Tomorrow's Work Schedule",
                r.get('tomorrow_work', "")
            )

            add_readonly_section(
                "Problems / Issues",
                r.get('problems_issues', "")
            )

            add_readonly_section(
                "Shared Matters",
                r.get('shared_matters', "")
            )

    def export_to_pdf(self):

        start = self.start_cal.get_date()
        end = self.end_cal.get_date()
        member_search = self.member_entry.get().strip()

        query = '''
            SELECT
                u.full_name,
                dr.report_date,
                dr.today_work,
                dr.tomorrow_work,
                dr.problems_issues,
                dr.shared_matters
            FROM users u
            JOIN daily_reports dr ON u.id = dr.user_id
            WHERE u.team_id=%s
            AND dr.report_date BETWEEN %s AND %s
            AND u.role='member'
        '''

        params = [self.team_id, start, end]

        if member_search:
            query += " AND u.full_name LIKE %s"
            params.append(f"%{member_search}%")

        query += """
            ORDER BY
            dr.report_date DESC,
            u.full_name ASC,
            dr.created_at ASC
        """

        try:

            self.db.cursor.execute(query, tuple(params))
            rows = self.db.cursor.fetchall()

            if not rows:
                self._show_message("No reports found.", "info")
                return

            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[
                    ("PDF files", "*.pdf")
                ],
                initialfile=f"{self.team_name}_Reports.pdf"
            )

            if not file_path:
                return

            if not os.path.splitext(file_path)[1]:
                file_path += ".pdf"

            if file_path.lower().endswith(".xlsx"):
                try:
                    from openpyxl import Workbook
                    from openpyxl.utils import get_column_letter
                    from openpyxl.styles import Alignment
                except ImportError:
                    self._show_message("openpyxl is required to export Excel files. Install it with:\n\npip install openpyxl", "error")
                    return

                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Reports"

                headers = [
                    "Date",
                    "Name",
                    "Today's Work Detail",
                    "Tomorrow's Work Schedule",
                    "Problems/Issues",
                    "Shared Matters"
                ]
                sheet.append(headers)

                for r in rows:
                    values = [
                        r['report_date'],
                        r.get('full_name') or "-",
                        r.get('today_work') or "-",
                        r.get('tomorrow_work') or "-",
                        r.get('problems_issues') or "-",
                        r.get('shared_matters') or "-"
                    ]
                    row_idx = sheet.max_row + 1
                    for col_idx, value in enumerate(values, start=1):
                        cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                        cell.alignment = Alignment(wrap_text=True, vertical='top')

                for col_idx in range(1, len(headers) + 1):
                    max_width = 0
                    for row_idx in range(1, sheet.max_row + 1):
                        cell_value = sheet.cell(row=row_idx, column=col_idx).value or ""
                        for line in str(cell_value).splitlines():
                            max_width = max(max_width, len(line))
                    sheet.column_dimensions[get_column_letter(col_idx)].width = max(max_width + 2, 15)

                workbook.save(file_path)
                self._show_message("Excel exported successfully!", "success")
                return

            class PDF(FPDF):

                def header(self):

                    # Title
                    self.set_text_color(0, 0, 0)
                    self.set_font("Arial", "B", 20)

                    self.cell(
                        0,
                        12,
                        f"{self.team_name} Reports",
                        ln=True,
                        align="C"
                    )

                    self.ln(4)

                    # Date row
                    self.set_font("Arial", "B", 11)

                    self.cell(
                        14,
                        8,
                        "Date :",
                        ln=0
                    )

                    self.set_font("Arial", "", 11)

                    self.cell(
                        0,
                        8,
                        f"{self.start_date} ~ {self.end_date}",
                        ln=True
                    )

                    self.ln(4)

                    # Table header
                    self.set_font("Arial", "B", 10)

                    headers = [
                        "Date",
                        "Name",
                        "Today's Work Detail",
                        "Tomorrow's Work Schedule",
                        "Problems/Issues",
                        "Shared Matters"
                    ]

                    widths = self.col_widths
                    self.set_draw_color(169, 169, 169)
                    self.set_fill_color(0, 0, 139)
                    self.set_text_color(245, 245, 245)

                    for i, header in enumerate(headers):

                        self.cell(
                            widths[i],
                            12,
                            header,
                            border=1,
                            align="C",
                            fill=True
                        )

                    self.ln()
                    self.set_text_color(0, 0, 0)
                    self.set_fill_color(255, 255, 255)

            pdf = PDF("L", "mm", "A4")

            pdf.team_name = self.team_name
            pdf.start_date = start
            pdf.end_date = end

            # widths
            pdf.col_widths = [28, 38, 60, 60, 42, 42]

            pdf.set_auto_page_break(auto=True, margin=15)

            pdf.add_page()

            body_font_size = 8

            pdf.set_font("Arial", "", body_font_size)

            line_height = 5

            def split_text(text, width):

                if not text:
                    return ["-"]

                pdf.set_font("Arial", "", body_font_size)

                lines = []

                for paragraph in str(text).split("\n"):

                    words = paragraph.split(" ")

                    current_line = ""

                    for word in words:

                        test_line = (
                            current_line + " " + word
                        ).strip()

                        if pdf.get_string_width(test_line) <= width - 2:
                            current_line = test_line
                        else:
                            if current_line:
                                lines.append(current_line)

                            current_line = word

                    if current_line:
                        lines.append(current_line)

                return lines if lines else ["-"]

            for r in rows:

                cells = [
                    str(r['report_date']),
                    str(r.get('full_name') or "-"),
                    str(r.get('today_work') or "-"),
                    str(r.get('tomorrow_work') or "-"),
                    str(r.get('problems_issues') or "-"),
                    str(r.get('shared_matters') or "-")
                ]

                split_cells = []

                max_lines = 1

                for i, text in enumerate(cells):

                    lines = split_text(
                        text,
                        pdf.col_widths[i]
                    )

                    split_cells.append(lines)

                    if len(lines) > max_lines:
                        max_lines = len(lines)

                bottom_padding = 3

                row_height = (
                    max_lines * line_height
                ) + bottom_padding

                # page break
                if pdf.get_y() + row_height > 190:

                    pdf.add_page()

                    pdf.set_font(
                        "Arial",
                        "",
                        body_font_size
                    )

                x = pdf.get_x()
                y = pdf.get_y()
                pdf.set_draw_color(169, 169, 169)

                # borders
                for i, width in enumerate(pdf.col_widths):

                    pdf.rect(
                        x,
                        y,
                        width,
                        row_height
                    )

                    x += width

                # text
                x = pdf.l_margin

                for i, lines in enumerate(split_cells):

                    current_y = y + 2

                    for line in lines:

                        pdf.set_xy(
                            x + 2,
                            current_y
                        )

                        pdf.cell(
                            pdf.col_widths[i] - 4,
                            line_height,
                            line,
                            border=0
                        )

                        current_y += line_height

                    x += pdf.col_widths[i]

                pdf.set_y(y + row_height)

            pdf.output(file_path)

            self._show_message("PDF exported successfully!", "success")

        except Exception as e:

            self._show_message(str(e, "error")
            )

    def export_to_csv(self):

        start = self.start_cal.get_date()
        end = self.end_cal.get_date()
        member_search = self.member_entry.get().strip()

        query = '''
            SELECT
                u.full_name,
                dr.report_date,
                dr.today_work,
                dr.tomorrow_work,
                dr.problems_issues,
                dr.shared_matters
            FROM users u
            JOIN daily_reports dr ON u.id = dr.user_id
            WHERE u.team_id=%s
            AND dr.report_date BETWEEN %s AND %s
            AND u.role='member'
        '''
        params = [self.team_id, start, end]

        if member_search:
            query += " AND u.full_name LIKE %s"
            params.append(f"%{member_search}%")

        query += " ORDER BY dr.report_date DESC, u.full_name ASC, dr.created_at ASC"

        try:
            self.db.cursor.execute(query, tuple(params))
            rows = self.db.cursor.fetchall()

            if not rows:
                self._show_message("No reports found.", "info")
                return

            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"{self.team_name}_Reports.xlsx"
            )
            if not file_path:
                return
            base, ext = os.path.splitext(file_path)
            if ext.lower() != ".xlsx":
                file_path = base + ".xlsx"

            try:
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
                from openpyxl.styles import Alignment
            except ImportError:
                self._show_message("openpyxl is required to export Excel files. Install it with:\n\npip install openpyxl", "error")
                return

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Reports"

            # Title and date rows
            ncols = 6
            title = f"{self.team_name} Reports"
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            tcell = sheet.cell(row=1, column=1, value=title)
            tcell.font = tcell.font.copy(bold=True, size=18, color="000000")
            tcell.alignment = tcell.alignment.copy(horizontal='center', vertical='center')
            sheet.row_dimensions[1].height = 26

            date_text = f"Date : {self.start_date} ~ {self.end_date}"
            sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
            dcell = sheet.cell(row=2, column=1, value=date_text)
            dcell.font = dcell.font.copy(bold=False, size=11, color="000000")
            dcell.alignment = dcell.alignment.copy(horizontal='left', vertical='center')
            sheet.row_dimensions[2].height = 18

            headers = [
                "Date",
                "Name",
                "Today's Work Detail",
                "Tomorrow's Work Schedule",
                "Problems/Issues",
                "Shared Matters"
            ]

            hdr_row = 4

            from openpyxl.styles import (
                PatternFill,
                Border,
                Side,
                Font,
                Alignment
            )

            header_fill = PatternFill(
                fill_type="solid",
                start_color="D9D9D9",
                end_color="D9D9D9"
            )

            side = Side(
                border_style="thin",
                color="000000"
            )

            for c, h in enumerate(headers, start=1):

                cell = sheet.cell(
                    row=hdr_row,
                    column=c,
                    value=h
                )

                cell.font = Font(
                    bold=True,
                    color="000000"
                )

                cell.fill = header_fill

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

                cell.border = Border(
                    left=side,
                    right=side,
                    top=side,
                    bottom=side
                )

            sheet.row_dimensions[hdr_row].height = 25

            for r in rows:
                values = [
                    r['report_date'],
                    r.get('full_name') or "-",
                    r.get('today_work') or "-",
                    r.get('tomorrow_work') or "-",
                    r.get('problems_issues') or "-",
                    r.get('shared_matters') or "-"
                ]
                row_idx = sheet.max_row + 1
                for col_idx, value in enumerate(values, start=1):
                    cell = sheet.cell(
                        row=row_idx,
                        column=col_idx,
                        value=value
                    )

                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical='top'
                    )

                    side = Side(
                        border_style="thin",
                        color="A9A9A9"
                    )

                    cell.border = Border(
                        left=side,
                        right=side,
                        top=side,
                        bottom=side
                    )

            # Auto-size columns
            widths = []
            for col_idx in range(1, ncols + 1):
                max_width = 0
                for row_idx in range(1, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value or ""
                    for line in str(cell_value).splitlines():
                        max_width = max(max_width, len(line))
                widths.append(max(max_width + 2, 15))
                sheet.column_dimensions[get_column_letter(col_idx)].width = widths[-1]

            def estimate_row_height(row_idx):
                max_lines = 1
                for col_idx, width in enumerate(widths, start=1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value or ""
                    lines = []
                    for paragraph in str(cell_value).splitlines():
                        current = ""
                        for word in paragraph.split(" "):
                            test_line = (current + " " + word).strip()
                            if len(test_line) <= width:
                                current = test_line
                            else:
                                if current:
                                    lines.append(current)
                                current = word
                        if current:
                            lines.append(current)
                    max_lines = max(max_lines, len(lines))
                return max(18, max_lines * 15)

            for row_idx in range(3, sheet.max_row + 1):
                sheet.row_dimensions[row_idx].height = estimate_row_height(row_idx)
                
            from openpyxl.styles import Border, Side

            side = Side(
                border_style="thin",
                color="A9A9A9"
            )

            for row in sheet.iter_rows(
                min_row=4,
                max_row=sheet.max_row,
                min_col=1,
                max_col=6
            ):
                for cell in row:
                    cell.border = Border(
                        left=side,
                        right=side,
                        top=side,
                        bottom=side
                    )

            workbook.save(file_path)
            self._show_message("Excel exported successfully!", "success")
        except Exception as e:
            self._show_message(str(e, "error"))