import customtkinter as ctk
from database import Database
from tkinter import messagebox, filedialog, ttk
from tkcalendar import Calendar
import tkinter as tk
import random
from datetime import datetime, timedelta
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import csv
import calendar


# ── Reusable modern date-picker ──────────────────────────────────────────────
class DatePickerButton(ctk.CTkFrame):
    """A styled button that opens a compact calendar dropdown."""
    def __init__(self, master, label="Date", initial_date=None, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self._date = initial_date or datetime.today().date()
        self._open = False
        self._callback = None

        self.apply_calendar_style()

        self.btn = ctk.CTkButton(
            self,
            text=self._fmt(),
            width=170,
            height=36,
            corner_radius=10,
            fg_color=("#EAECEE", "#1E2A3A"),
            hover_color=("#D5D8DC", "#2C3E50"),
            border_width=1,
            border_color=("#ABB2B9", "#3D5166"),
            text_color=("#1A1A1A", "#FFFFFF"),
            anchor="w",
            command=self.toggle
        )
        self.btn.pack()

        self.panel = ctk.CTkFrame(
            self.winfo_toplevel(),
            fg_color=("#FFFFFF", "#141E2B"),
            corner_radius=12,
            border_width=1,
            border_color=("#ABB2B9", "#2A3A4A")
        )
                
        self.cal = Calendar(
            self.panel,
            style="Custom.Calendar",
            selectmode="day",
            date_pattern="yyyy-mm-dd",
            year=self._date.year,
            month=self._date.month,
            day=self._date.day,
            showweeknumbers=False,
            firstweekday="monday",
            font=("Arial", 11),
            cursor="hand2"
        )
        self.cal.pack(padx=8, pady=8)
        self.cal.bind("<<CalendarSelected>>", self._select)

        # Re-apply style whenever the widget is rendered to handle theme changes
        self.bind("<Expose>", lambda e: self.apply_calendar_style())

    def apply_calendar_style(self):
        style = ttk.Style()
        style.theme_use("clam")
        is_dark = ctk.get_appearance_mode() == "Dark"
        
        bg = "#1A1A2E" if is_dark else "#FFFFFF"
        fg = "white" if is_dark else "black"
        h_bg = "#16213E" if is_dark else "#EAECEE"
        h_fg = "#4FC3F7" if is_dark else "#2471A3"

        style.configure("Custom.Calendar", 
                        background=bg, foreground=fg,
                        headersbackground=h_bg, headersforeground=h_fg,
                        selectbackground="#3498DB", selectforeground="white",
                        normalbackground=bg, normalforeground=fg,
                        weekendbackground=bg, weekendforeground="#F39C12",
                        bordercolor=h_bg, relief="flat")
        
        # Safely configure the calendar if it exists
        if hasattr(self, 'cal'):
            self.cal.configure(style="Custom.Calendar")

    def _fmt(self):
        return f"  📅  {self._date.strftime('%Y-%m-%d')}"

    def toggle(self):
        if self._open:
            self.panel.place_forget()
        else:
            self.panel.lift()
            self.panel.place(in_=self, x=0, y=self.btn.winfo_height() + 2)
        self._open = not self._open

    def _select(self, event):
        selected = self.cal.get_date()
        self._date = datetime.strptime(selected, "%Y-%m-%d").date()
        self.btn.configure(text=self._fmt())
        self.toggle()
        if self._callback:
            self._callback(self._date)

    def get_date(self):
        return self._date

    def set_date(self, d):
        self._date = d
        self.cal.selection_set(d)
        self.btn.configure(text=self._fmt())

    def on_change(self, callback):
        self._callback = callback


# ── Main Schedule Frame ───────────────────────────────────────────────────────
class LeaderSchedule(ctk.CTkFrame):
    def __init__(self, master, user):
        super().__init__(master, fg_color=("#FFFFFF", "#1A1A1A"))
        self.db = Database()
        self.user = user

        today = datetime.today()
        self.current_date = today.date()
        self.current_month_start = today.replace(day=1).date()
        next_month = today.replace(day=28) + timedelta(days=4)
        self.current_month_end = (next_month - timedelta(days=next_month.day)).date()

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.setup_ui()
        self.auto_generate_monthly_schedule()
        self.refresh_view()

    # ── UI Setup ──────────────────────────────────────────────────────────────
    def setup_ui(self):
        self.wrapper = ctk.CTkFrame(self, fg_color="transparent")
        self.wrapper.grid(row=0, column=0, sticky="nsew", padx=80, pady=5)

        # ── Page Title ────────────────────────────────────────────────────────
        title_row = ctk.CTkFrame(self.wrapper, fg_color="transparent")
        title_row.pack(fill="x", padx=10, pady=(14, 4))

        team_id = self.user.get('team_id', '—')
        ctk.CTkLabel(
            title_row,
            text=f"🗓  Team {team_id}  ·  WFH Schedule Manager",
            font=("Arial", 20, "bold"),
            text_color=("#1A1A1A", "#FFFFFF")
        ).pack(side="left")

        # ── Filter Card ───────────────────────────────────────────────────────
        filter_card = ctk.CTkFrame(
            self.wrapper, corner_radius=14,
            fg_color=("#F2F4F4", "#141E2B"),
            border_width=1, border_color=("#D5D8DC", "#253545")
        )
        filter_card.pack(fill="x", padx=10, pady=(4, 10))

        inner = ctk.CTkFrame(filter_card, fg_color="transparent")
        inner.pack(fill="x", padx=18, pady=14)

        # ── Date pickers ──────────────────────────────────────────────────────
        def _lbl(parent, text):
            ctk.CTkLabel(parent, text=text,
                         font=("Arial", 11),
                         text_color=("#475569", "#8899AA")).pack(side="left", padx=(0, 4))

        _lbl(inner, "From")
        self.start_cal = DatePickerButton(inner, initial_date=self.current_date)
        self.start_cal.pack(side="left", padx=(0, 16))

        _lbl(inner, "To")
        self.end_cal = DatePickerButton(inner, initial_date=self.current_date)
        self.end_cal.pack(side="left", padx=(0, 24))

        # ── Separator ─────────────────────────────────────────────────────────
        sep = ctk.CTkFrame(inner, width=1, height=32, fg_color="#2A3A4A")
        sep.pack(side="left", padx=(0, 20))

        # ── Name filter ───────────────────────────────────────────────────────
        _lbl(inner, "Member")
        self.name_filter = ctk.CTkEntry(
            inner, placeholder_text="Search name…",
            width=100, height=36,
            corner_radius=8,
            border_color=("#ABB2B9", "#3D5166"), border_width=1,
            fg_color=("#FFFFFF", "#1E2A3A"),
            text_color=("#1A1A1A", "#FFFFFF"),
            placeholder_text_color=("#7B7D7D", "#8CA0B3"),
            font=("Arial", 12)
        )
        self.name_filter.pack(side="left", padx=(0, 16))

        # ── Status dropdown ───────────────────────────────────────────────────
        _lbl(inner, "Status")
        self.status_filter = ctk.CTkComboBox(
            inner, values=["All", "Office", "WFH"],
            width=110, height=36,
            corner_radius=8,
            border_color=("#ABB2B9", "#3D5166"), border_width=1,
            fg_color=("#FFFFFF", "#1E2A3A"),
            button_color=("#D5D8DC", "#2C3E50"),
            text_color=("#1A1A1A", "#FFFFFF"),
            font=("Arial", 12)
        )
        self.status_filter.set("All")
        self.status_filter.pack(side="left", padx=(0, 24))

        # ── Action buttons ────────────────────────────────────────────────────
        def _btn(parent, text, color, hover, cmd, width=60):
            b = ctk.CTkButton(
                parent, text=text, width=width, height=36,
                corner_radius=8,
                font=("Arial", 12, "bold"),
                fg_color=(color, color), hover_color=(hover, hover),
                text_color="#FFFFFF",
                command=cmd
            )
            b.pack(side="left", padx=4)
            return b

        _btn(inner, "🔍  Filter",  "#2471A3", "#1A5276", self.refresh_view)
        _btn(inner, "✖  Clear",   "#566573", "#424949", self.clear_filters)
        _btn(inner, "📥  CSV", "#16A085", "#117A65", self.export_to_xlsx)
        _btn(inner, "📄  PDF", "#C0392B", "#922B21", self.export_to_pdf)

        # ── List header ───────────────────────────────────────────────────────
        list_header = ctk.CTkFrame(self.wrapper, fg_color="transparent")
        list_header.pack(fill="x", padx=10, pady=(4, 2))

        ctk.CTkLabel(list_header,
                     text="Team Schedule Log",
                     font=("Arial", 14, "bold"),
                     text_color=("#4E5D6C", "#AABBCC")).pack(side="left", padx=4)

        self.count_lbl = ctk.CTkLabel(
            list_header,
            text="",
            font=("Arial", 10),
            text_color=("#93A0AA", "#6F8292")
        )
        self.count_lbl.pack(side="left", padx=8)

        # ── View toggle (List / Overview) ─────────────────────────────────────
        view_toggle = ctk.CTkFrame(self.wrapper, fg_color="transparent")
        view_toggle.pack(fill="x", padx=10, pady=(0, 4))
        self.view_mode = tk.StringVar(value="List")
        seg_view = ctk.CTkSegmentedButton(
            view_toggle, values=["List", "Overview"],
            variable=self.view_mode,
            font=("Arial", 11, "bold"),
            selected_color=("#2471A3", "#2471A3"),
            unselected_color=("#F0F3F5", "#2C3E50"),
            text_color=("#2C3E50", "#FFFFFF"),
            command=lambda v=None: self.switch_view(v)
        )
        seg_view.pack(side="right")
        # React to changes in the segmented control as a fallback
        try:
            self.view_mode.trace_add("write", lambda *a: self.switch_view())
        except Exception:
            # fallback for older tkinter versions
            self.view_mode.trace("w", lambda *a: self.switch_view())

        # ── Calendar overview frame (hidden by default)
        # Use a scrollable frame so full-month view can scroll when tall
        self.overview_frame = ctk.CTkScrollableFrame(
            self.wrapper,
            fg_color=("#FFFFFF", "#0D1117"),
            corner_radius=14,
            border_width=1,
            border_color=("#D5D8DC", "#1E2A3A")
        )
        # Do not pack: packed when toggled to Overview

        # ── Scroll list ───────────────────────────────────────────────────────
        self.scroll = ctk.CTkScrollableFrame(
            self.wrapper,
            fg_color=("#FFFFFF", "#0D1117"),
            corner_radius=14,
            border_width=1,
            border_color=("#D5D8DC", "#1E2A3A")
        )
        self.scroll.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    # ── Auto-generate ─────────────────────────────────────────────────────────
    def auto_generate_monthly_schedule(self):
        leader_team_id = self.user.get('team_id')
        if not leader_team_id:
            return
        self.db.cursor.execute(
            "SELECT id FROM users WHERE role='member' AND team_id=%s", (leader_team_id,))
        members = self.db.cursor.fetchall()
        if not members:
            return
        try:
            cur = self.current_month_start
            while cur <= self.current_month_end:
                if cur.weekday() < 5:
                    for m in members:
                        self.db.cursor.execute(
                            "SELECT id FROM wfh_schedules WHERE user_id=%s AND schedule_date=%s",
                            (m['id'], cur))
                        if not self.db.cursor.fetchone():
                            status = random.choice(['Office', 'WFH'])
                            self.db.cursor.execute(
                                "INSERT INTO wfh_schedules (user_id,leader_id,schedule_date,status) VALUES(%s,%s,%s,%s)",
                                (m['id'], self.user['id'], cur, status))
                cur += timedelta(days=1)
            self.db.conn.commit()
        except Exception as e:
            print(f"Sync Error: {e}")

    # ── Refresh list ──────────────────────────────────────────────────────────
    def refresh_view(self):
        for w in self.scroll.winfo_children():
            w.destroy()

        leader_team_id = self.user.get('team_id')
        start        = self.start_cal.get_date()
        end          = self.end_cal.get_date()
        name_search  = f"%{self.name_filter.get()}%"
        status_search = self.status_filter.get()
        today_date   = datetime.today().date()

        # Validate dates: allow past ranges; only ensure end is not before start
        if end < start:
            try:
                self._show_message("End date cannot be earlier than Start date", "error")
            except Exception:
                messagebox.showerror("Invalid date", "End date cannot be earlier than Start date")
            return

        query = """SELECT s.*, u.full_name
                   FROM wfh_schedules s
                   JOIN users u ON s.user_id = u.id
                   WHERE u.team_id=%s
                     AND s.schedule_date BETWEEN %s AND %s
                     AND u.full_name LIKE %s"""
        params = [leader_team_id, start, end, name_search]
        if status_search != "All":
            query += " AND s.status=%s"
            params.append(status_search)
        query += " ORDER BY s.schedule_date ASC, u.full_name ASC"

        try:
            self.db.cursor.execute(query, tuple(params))
            rows = self.db.cursor.fetchall()

            self.count_lbl.configure(text=f"({len(rows)} records)")

            if not rows:
                # If member filter is present, check whether any member matches the filter
                member_text = self.name_filter.get().strip()
                if member_text:
                    try:
                        self.db.cursor.execute(
                            "SELECT COUNT(*) AS cnt FROM users WHERE team_id=%s AND full_name LIKE %s",
                            (leader_team_id, f"%{member_text}%")
                        )
                        found = self.db.cursor.fetchone()
                        if found and found.get('cnt', 0) == 0:
                            # No member matches the entered name
                            ctk.CTkLabel(self.scroll, text=f"No members found matching '{member_text}'.", text_color=("#7B7D7D", "#AABBCD")).pack(pady=20)
                            try:
                                self._show_message(f"No members match '{member_text}'", "info")
                            except Exception:
                                pass
                            return
                    except Exception:
                        # Fallback to generic message on error
                        pass

                # Generic no-results message when members exist but no schedules
                ctk.CTkLabel(self.scroll, text="No schedule assigned for this period.", text_color=("#7B7D7D", "#AABBCD")).pack(pady=20)
                try:
                    self._show_message("No record found for selected period", "info")
                except Exception:
                    pass
                return

            # Group by date
            from itertools import groupby
            for date_val, group in groupby(rows, key=lambda r: r['schedule_date']):
                group = list(group)
                is_today = date_val == today_date

                # ── Date separator row ─────────────────────────────────────
                date_sep = ctk.CTkFrame(self.scroll, fg_color="transparent")
                date_sep.pack(fill="x", padx=8, pady=(8, 1))

                day_str  = date_val.strftime("%A")
                date_str = date_val.strftime("%d %b %Y")
                badge_color = ("#FDEBD0", "#E67E22") if is_today else ("#D6DBDF", "#2C3E50")
                badge_text  = f"  {date_str}  ·  {day_str}{'  ← TODAY' if is_today else ''}  "

                ctk.CTkLabel(
                    date_sep,
                    text=badge_text,
                    font=("Arial", 11, "bold"),
                    text_color=("#7E5109", "#FFFFFF") if is_today else ("#1A1A1A", "#FFFFFF"),
                    fg_color=badge_color,
                    corner_radius=6,
                    padx=6, pady=2
                ).pack(side="left")

                # ── Member cards under that date ───────────────────────────
                for r in group:
                    is_wfh = r['status'] == 'WFH'
                    accent = "#F39C12" if is_today else ("#3498DB" if is_wfh else "#27AE60")

                    # Card — compact, with colored left border
                    card = ctk.CTkFrame(
                        self.scroll,
                        corner_radius=8,
                        fg_color=("#FFFFFF", "#151C28"),
                        border_width=1,
                        border_color=accent
                    )
                    card.pack(fill="x", pady=1, padx=10)

                    # ── Left: Avatar ──────────────────────────────────────
                    name       = r['full_name']
                    initials   = "".join([n[0] for n in name.split()[:2]]).upper()
                    avatar_clr = ("#D6EAF8", "#1A3A5C") if is_wfh else ("#D5F5E3", "#1A4032")

                    avatar = ctk.CTkFrame(card, width=34, height=34,
                                          corner_radius=17, fg_color=avatar_clr,
                                          border_width=2, border_color=accent)
                    avatar.pack(side="left", padx=(10, 0), pady=6)
                    avatar.pack_propagate(False)
                    ctk.CTkLabel(avatar, text=initials,
                                 font=("Arial", 11, "bold"),
                                 text_color=("#1B4F72", "#5DADE2") if is_wfh else ("#145A32", "#58D68D")).pack(expand=True)

                    # ── Middle: Name + sub info ───────────────────────────
                    info = ctk.CTkFrame(card, fg_color="transparent")
                    info.pack(side="left", fill="both", expand=True, padx=10, pady=6)

                    ctk.CTkLabel(info, text=name,
                                 font=("Arial", 12, "bold"),
                                 text_color=("#1A1A1A", "#E8EDF2")).pack(anchor="w")

                    emp_id = r.get('employee_id', '')
                    sub_text = f"ID: {emp_id}" if emp_id else "Team Member"
                    ctk.CTkLabel(info, text=sub_text,
                                 font=("Arial", 9),
                                 text_color=("#5F6D7A", "#4A5568")).pack(anchor="w")

                    # ── Right: Status pill + edit icon ────────────────────
                    right = ctk.CTkFrame(card, fg_color="transparent")
                    right.pack(side="right", padx=(0, 10), pady=6)

                    pill_bg   = ("#D6EAF8", "#1A3A5C") if is_wfh else ("#D5F5E3", "#1A4032")
                    pill_text = ("#1B4F72", "#5DADE2") if is_wfh else ("#145A32", "#58D68D")
                    status_icon = "🏠" if is_wfh else "🏢"

                    ctk.CTkLabel(
                        right,
                        text=f" {status_icon} {r['status']} ",
                        font=("Arial", 11, "bold"),
                        text_color=pill_text,
                        fg_color=pill_bg,
                        corner_radius=6,
                        padx=6, pady=2
                    ).pack(side="left", padx=(0, 6))

                    ctk.CTkButton(
                        right,
                        text="Edit",
                        width=60, height=30,
                        corner_radius=8,
                        font=("Arial", 12, "bold"),
                        fg_color=("#F39C12", "#F39C12"),
                        hover_color=("#E08E00", "#E08E00"),
                        text_color=("#1A1A1A", "#FFFFFF"),
                        border_width=0,
                        command=lambda row=r: self.open_edit_popup(row)
                    ).pack(side="left", padx=(8,0))

        except Exception as e:
            messagebox.showerror("Error", str(e))

    # ── Overview rendering and view switching ─────────────────────────────────
    def switch_view(self, _=None):
        mode = self.view_mode.get()
        if mode == "Overview":
            try:
                self.scroll.pack_forget()
            except:
                pass
            # initialize overview month/year if not present
            if not hasattr(self, 'overview_year') or not hasattr(self, 'overview_month'):
                d = self.start_cal.get_date()
                self.overview_year = d.year
                self.overview_month = d.month
            self.overview_frame.pack(fill="both", expand=True, padx=10, pady=(0,10))
            self.render_overview()
        else:
            try:
                self.overview_frame.pack_forget()
            except:
                pass
            self.scroll.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    def prev_month(self):
        if not hasattr(self, 'overview_month'):
            return
        m = self.overview_month - 1
        y = self.overview_year
        if m < 1:
            m = 12
            y -= 1
        self.overview_month = m
        self.overview_year = y
        self.render_overview()

    def next_month(self):
        if not hasattr(self, 'overview_month'):
            return
        m = self.overview_month + 1
        y = self.overview_year
        if m > 12:
            m = 1
            y += 1
        self.overview_month = m
        self.overview_year = y
        self.render_overview()

    def render_overview(self):
        # Clear previous
        for w in self.overview_frame.winfo_children():
            w.destroy()

        leader_team_id = self.user.get('team_id')
        year = getattr(self, 'overview_year', self.start_cal.get_date().year)
        month = getattr(self, 'overview_month', self.start_cal.get_date().month)
        self.overview_year = year
        self.overview_month = month

        # Build summary counts per date
        first_day = datetime(year, month, 1).date()
        import calendar as _cal
        last_day_num = _cal.monthrange(year, month)[1]
        last_day = datetime(year, month, last_day_num).date()

        try:
            self.db.cursor.execute(
                """SELECT s.schedule_date, s.status, COUNT(*) AS cnt
                   FROM wfh_schedules s JOIN users u ON s.user_id=u.id
                   WHERE u.team_id=%s AND s.schedule_date BETWEEN %s AND %s
                   GROUP BY s.schedule_date, s.status""",
                (leader_team_id, first_day, last_day)
            )
            rows = self.db.cursor.fetchall()
            summary = {}
            for r in rows:
                d = r['schedule_date']
                summary.setdefault(d, {'WFH':0, 'Office':0})
                summary[d][r['status']] = r['cnt']
        except Exception as e:
            messagebox.showerror("Overview Error", str(e))
            summary = {}

        # Colors adjusted for light/dark modes
        is_dark = ctk.get_appearance_mode() == "Dark"
        month_label_color = ("#2C3E50", "#E6EEF6")  # light, dark
        weekday_text_color = ("#2C3E50", "#BFC9D3")
        # Day numbers need dark text in light mode and light text in dark mode.
        date_text_color = ("#0F172A", "#FFFFFF")
        nav_btn_bg = ("#F0F3F5","#2F3B43")
        nav_btn_text = ("#2C3E50","#E6EEF6")
        overview_card_bg = ("#F2F4F4", "#0B0F12")
        day_cell_bg = ("#FFFFFF", "#0E1518")
        day_cell_border = ("#CBD5E1", "#24313A")
        day_cell_hover_bg = ("#EAF2F8", "#0F2228")
        day_cell_hover_border = ("#7FB3D5", "#57A6FF")

        # Top: Month navigation
        nav = ctk.CTkFrame(self.overview_frame, fg_color="transparent")
        nav.pack(fill="x", pady=(6,8))
        ctk.CTkButton(nav, text="◀", width=36, height=30, fg_color=nav_btn_bg, hover_color=("#E0EAF4","#3E4A52"), text_color=nav_btn_text, command=self.prev_month).pack(side="left")
        ctk.CTkLabel(nav, text=f"{calendar.month_name[month]} {year}", font=("Arial", 18, "bold"), text_color=month_label_color).pack(side="left", padx=12)
        ctk.CTkButton(nav, text="▶", width=36, height=30, fg_color=nav_btn_bg, hover_color=("#E0EAF4","#3E4A52"), text_color=nav_btn_text, command=self.next_month).pack(side="left")

        # Single master grid container keeps header and calendar cells aligned.
        calendar_wrap = ctk.CTkFrame(
            self.overview_frame,
            fg_color=overview_card_bg,
            corner_radius=14,
            border_width=1,
            border_color=("#D5D8DC", "#1E2A3A"),
        )
        calendar_wrap.pack(fill="both", expand=True, pady=12, padx=(12, 24))

        weekdays = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        # Mon-Fri stretch, Sat/Sun stay narrow.
        day_weights = [5, 5, 5, 5, 5, 0, 0]
        day_minsizes = [0, 0, 0, 0, 0, 72, 72]
        day_pads = [6, 6, 6, 6, 6, 1, 1]
        weekend_width = 72

        for idx in range(7):
            calendar_wrap.grid_columnconfigure(idx, weight=day_weights[idx], minsize=day_minsizes[idx])

        for idx, wd in enumerate(weekdays):
            label_width = weekend_width if idx >= 5 else 1
            ctk.CTkLabel(
                calendar_wrap,
                text=wd,
                anchor="center",
                font=("Arial", 13, "bold"),
                text_color=weekday_text_color,
                width=label_width,
            ).grid(row=0, column=idx, sticky="ew", padx=(day_pads[idx], day_pads[idx]), pady=(0, 6))

        month_matrix = calendar.monthcalendar(year, month)
        normal_bg = ("#0E1518", "#0B0F12")
        normal_border = ("#24313A", "#2B3B44")
        hover_border = ("#57A6FF", "#57A6FF")

        for r_idx, week in enumerate(month_matrix, start=1):
            calendar_wrap.grid_rowconfigure(r_idx, weight=1)
            for c_idx, day in enumerate(week):
                cell = ctk.CTkFrame(
                    calendar_wrap,
                    width=weekend_width if c_idx >= 5 else 1,
                    height=120,
                    fg_color=day_cell_bg,
                    corner_radius=10,
                    border_width=1,
                    border_color=day_cell_border,
                )
                cell.grid(
                    row=r_idx,
                    column=c_idx,
                    sticky="nsew",
                    padx=(day_pads[c_idx], day_pads[c_idx]),
                    pady=4,
                )
                cell.pack_propagate(False)

                if day == 0:
                    placeholder = ctk.CTkFrame(cell, fg_color=day_cell_bg)
                    placeholder.pack(fill="both", expand=True)
                    continue

                d_date = datetime(year, month, day).date()
                is_today = d_date == datetime.today().date()

                def _on_enter(ev, widget=cell):
                    try:
                        widget.configure(border_color=day_cell_hover_border)
                        widget.configure(fg_color=day_cell_hover_bg)
                    except Exception:
                        pass

                def _on_leave(ev, widget=cell):
                    try:
                        widget.configure(border_color=day_cell_border)
                        widget.configure(fg_color=day_cell_bg)
                    except Exception:
                        pass

                cell.bind("<Enter>", _on_enter)
                cell.bind("<Leave>", _on_leave)

                top_row = ctk.CTkFrame(cell, fg_color="transparent")
                top_row.pack(fill="x", padx=10, pady=(8, 2))
                date_txt = str(day) + ("  ← Today" if is_today else "")
                ctk.CTkLabel(
                    top_row,
                    text=date_txt,
                    anchor="w",
                    font=("Arial", 14, "bold"),
                    text_color=date_text_color,
                ).pack(side="left")

                counts_frame = ctk.CTkFrame(cell, fg_color="transparent")
                counts_frame.pack(fill="both", expand=True, padx=10, pady=6)
                counts = summary.get(d_date, {"WFH": 0, "Office": 0})

                if counts.get("WFH", 0):
                    ctk.CTkLabel(
                        counts_frame,
                        text=f"  🏠 {counts['WFH']}  ",
                        font=("Arial", 12),
                        fg_color=("#0F3B4A", "#0F3B4A"),
                        text_color=("#8FD8FF", "#8FD8FF"),
                        corner_radius=8,
                    ).pack(anchor="w", pady=4)
                if counts.get("Office", 0):
                    ctk.CTkLabel(
                        counts_frame,
                        text=f"  🏢 {counts['Office']}  ",
                        font=("Arial", 12),
                        fg_color=("#0F3A24", "#0F3A24"),
                        text_color=("#9EE6B6", "#9EE6B6"),
                        corner_radius=8,
                    ).pack(anchor="w", pady=4)

    # ── Edit popup ────────────────────────────────────────────────────────────
    def open_edit_popup(self, row_data):
        parent = self.winfo_toplevel()
        popup = ctk.CTkToplevel(parent)
        popup.title("Update Schedule Status")
        popup.geometry("360x240")
        popup.resizable(False, False)
        popup.transient(parent)
        popup.attributes('-topmost', True)
        popup_bg = "#F8FAFC" if ctk.get_appearance_mode() == "Light" else "#0D1117"
        popup.configure(fg_color=popup_bg)

        # Center the popup relative to the current application window.
        popup.update_idletasks()
        parent.update_idletasks()
        popup_width = 360
        popup_height = 240
        parent_x = parent.winfo_rootx()
        parent_y = parent.winfo_rooty()
        parent_width = parent.winfo_width()
        parent_height = parent.winfo_height()
        center_x = parent_x + (parent_width // 2) - (popup_width // 2)
        center_y = parent_y + (parent_height // 2) - (popup_height // 2)
        popup.geometry(f"{popup_width}x{popup_height}+{center_x}+{center_y}")

        ctk.CTkLabel(popup, text="✏  Edit Schedule",
                     font=("Arial", 16, "bold"),
                     text_color=("#0F172A", "#FFFFFF")).pack(pady=(20, 4))

        ctk.CTkLabel(popup, text=f"{row_data['full_name']}  ·  {row_data['schedule_date']}",
                     font=("Arial", 12),
                     text_color=("#475569", "#8899AA")).pack()

        ctk.CTkFrame(popup, height=1, fg_color=("#D7DEE8", "#1E2A3A")).pack(fill="x", padx=20, pady=12)

        status_var = ctk.StringVar(value=row_data['status'])
        seg = ctk.CTkSegmentedButton(
            popup, values=["Office", "WFH"],
            variable=status_var,
            font=("Arial", 13, "bold"),
            selected_color=("#2471A3", "#2471A3"),
            selected_hover_color=("#2471A3", "#2471A3"),
            unselected_color=("#CBD5E1", "#1E2A3A"),
            unselected_hover_color=("#CBD5E1", "#1E2A3A"),
            fg_color=("#E2E8F0", "#141E2B"),
            text_color=("#0F172A", "#FFFFFF")
        )
        seg.pack(pady=8)

        def save_change():
            try:
                self.db.cursor.execute(
                    "UPDATE wfh_schedules SET status=%s WHERE id=%s",
                    (status_var.get(), row_data['id']))
                self.db.conn.commit()
                popup.destroy()
                self.refresh_view()
            except Exception as e:
                messagebox.showerror("Update Fail", str(e))

        ctk.CTkButton(popup, text="✔  Save Change",
                      height=38, corner_radius=10,
                      font=("Arial", 13, "bold"),
                      fg_color=("#27AE60", "#27AE60"), hover_color=("#1E8449", "#1E8449"),
                      command=save_change).pack(pady=(12, 0), padx=30, fill="x")

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _show_message(self, message, message_type="info", duration=3000):
        """
        Displays a transient message in the top-right corner of the master window.
        message_type: "info", "warning", "error", "success"
        """
        # Determine colors based on message_type
        if message_type == "error":
            bg_color = "#E74C3C"  # Red
            text_color = "white"
        elif message_type == "warning":
            bg_color = "#F39C12"  # Orange
            text_color = "white"
        elif message_type == "success":
            bg_color = "#27AE60"  # Green
            text_color = "white"
        else:  # info
            bg_color = "#3498DB"  # Blue
            text_color = "white"

        # Create a frame for the message
        message_frame = ctk.CTkFrame(
            self.winfo_toplevel(),
            fg_color=bg_color,
            corner_radius=8
        )
        # Position in the top right corner, with some padding
        message_frame.place(relx=1.0, rely=0, x=-20, y=20, anchor="ne") 

        ctk.CTkLabel(
            message_frame,
            text=message,
            text_color=text_color,
            font=("Arial", 12, "bold"),
            wraplength=250 # Wrap text if too long
        ).pack(padx=15, pady=10)

        # Destroy the message after 'duration' milliseconds
        self.winfo_toplevel().after(duration, message_frame.destroy)

    def clear_filters(self, refresh=True):
        self.start_cal.set_date(self.current_date)
        self.end_cal.set_date(self.current_date)
        self.name_filter.delete(0, 'end')
        self.status_filter.set("All")
        if refresh:
            self.refresh_view()

    # ── PDF Export ────────────────────────────────────────────────────────────

    # ── XLSX/CSV Export helper (styled XLSX when possible) ─────────────────────
    def export_to_pdf(self):
        leader_team_id = self.user.get('team_id')
        start = self.start_cal.get_date()
        end   = self.end_cal.get_date()

        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            initialfile=f"Team_{leader_team_id}_Schedule.pdf"
        )
        if not file_path:
            return
        try:
            self.db.cursor.execute(
                """SELECT s.schedule_date, u.full_name, s.status
                   FROM wfh_schedules s JOIN users u ON s.user_id=u.id
                   WHERE u.team_id=%s AND s.schedule_date BETWEEN %s AND %s
                   ORDER BY s.schedule_date ASC""",
                (leader_team_id, start, end))
            data = self.db.cursor.fetchall()

            doc      = SimpleDocTemplate(file_path, pagesize=letter)
            styles   = getSampleStyleSheet()
            elements = [
                Paragraph(f"Team {leader_team_id} Work Schedule", styles['Title']),
                Paragraph(f"Period: {start} to {end}", styles['Normal']),
                Spacer(1, 15)
            ]
            table_data = [["Date", "Member Name", "Status"]]
            for r in data:
                day = r['schedule_date'].strftime('%a')
                table_data.append([f"{r['schedule_date']} ({day})", r['full_name'], r['status']])

            t = Table(table_data, colWidths=[130, 200, 100])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR',  (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID',       (0, 0), (-1, -1), 0.5, colors.grey)
            ]))
            elements.append(t)
            doc.build(elements)
            self._show_message("PDF exported successfully!", "success", duration=3000)
        except Exception as e:
            self._show_message(f"PDF Error: {e}", "error", duration=3000)

    def export_to_xlsx(self):
        """Export as .xlsx with borders and auto-fit widths when possible; fallback to CSV.
        This function is wired to the CSV button (shows 📥) so users get a formatted spreadsheet by default.
        """
        leader_team_id = self.user.get('team_id')
        start = self.start_cal.get_date()
        end   = self.end_cal.get_date()

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=f"Team_{leader_team_id}_Schedule.xlsx"
        )
        if not file_path:
            return

        try:
            self.db.cursor.execute(
                """SELECT s.schedule_date, u.full_name, s.status
                   FROM wfh_schedules s JOIN users u ON s.user_id=u.id
                   WHERE u.team_id=%s AND s.schedule_date BETWEEN %s AND %s
                   ORDER BY s.schedule_date ASC""",
                (leader_team_id, start, end))
            data = self.db.cursor.fetchall()

            try:
                import openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from openpyxl.utils import get_column_letter

                wb = Workbook()
                ws = wb.active
                ws.title = "Schedule"

                headers = ["Date", "Member Name", "Status"]
                ws.append(headers)
                for r in data:
                    dstr = r['schedule_date'].strftime('%Y-%m-%d') if hasattr(r['schedule_date'], 'strftime') else str(r['schedule_date'])
                    ws.append([dstr, r['full_name'], r['status']])

                thin = Side(border_style="thin", color="444444")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                col_widths = [0] * ws.max_column
                for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), start=1):
                    for col_idx, cell in enumerate(row, start=1):
                        cell.border = border
                        if row_idx == 1:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(fill_type="solid", fgColor="DDDDDD")
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        val = str(cell.value) if cell.value is not None else ""
                        col_widths[col_idx-1] = max(col_widths[col_idx-1], len(val))

                for i, wth in enumerate(col_widths, start=1):
                    col = get_column_letter(i)
                    ws.column_dimensions[col].width = max(10, wth + 2)

                save_path = file_path if file_path.lower().endswith('.xlsx') else file_path.rsplit('.',1)[0] + '.xlsx'
                wb.save(save_path)
                self._show_message("Excel exported successfully!", "success", duration=3000)
                return
            except Exception:
                # openpyxl not available or failed — fall back to CSV
                pass

            # Fallback CSV
            csv_path = file_path if file_path.lower().endswith('.csv') else file_path.rsplit('.',1)[0] + '.csv'
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["Date", "Member Name", "Status"])
                for r in data:
                    dstr = r['schedule_date'].strftime('%Y-%m-%d') if hasattr(r['schedule_date'], 'strftime') else str(r['schedule_date'])
                    writer.writerow([dstr, r['full_name'], r['status']])
            self._show_message("Excel export failed, saved as CSV instead!", "warning", duration=3000)
        except Exception as e:
            self._show_message(f"Export Error: {e}", "error", duration=3000)

    def export_to_csv(self):
        leader_team_id = self.user.get('team_id')
        start = self.start_cal.get_date()
        end   = self.end_cal.get_date()

        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            initialfile=f"Team_{leader_team_id}_Schedule.csv"
        )
        if not file_path:
            return
        try:
            self.db.cursor.execute(
                """SELECT s.schedule_date, u.full_name, s.status
                   FROM wfh_schedules s JOIN users u ON s.user_id=u.id
                   WHERE u.team_id=%s AND s.schedule_date BETWEEN %s AND %s
                   ORDER BY s.schedule_date ASC""",
                (leader_team_id, start, end))
            data = self.db.cursor.fetchall()
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["Date", "Member Name", "Status"])
                for r in data:
                    dstr = r['schedule_date'].strftime('%Y-%m-%d') if hasattr(r['schedule_date'], 'strftime') else str(r['schedule_date'])
                    writer.writerow([dstr, r['full_name'], r['status']])
            self._show_message("CSV exported successfully!", "success", duration=3000)
        except Exception as e:
            self._show_message(f"CSV Error: {e}", "error", duration=3000)
