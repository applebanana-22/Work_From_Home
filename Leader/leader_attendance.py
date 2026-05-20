import customtkinter as ctk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime, timedelta
from datetime import time as dt_time
from xml.sax.saxutils import escape
from tkcalendar import Calendar
import tkinter as tk
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database import Database


class DatePickerButton(ctk.CTkFrame):
    def __init__(self, master, initial_date=None, **kwargs):
        super().__init__(master, fg_color="transparent")

        self._date = initial_date or datetime.today().date()
        self._open = False
        self._callback = None

        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Custom.Calendar",
            background="#1A1A2E",
            foreground="white",
            headersbackground="#16213E",
            headersforeground="#4FC3F7",
            selectbackground="#3498DB",
            selectforeground="white",
            normalbackground="#1A1A2E",
            normalforeground="#CCCCCC",
            weekendbackground="#1A1A2E",
            weekendforeground="#F39C12",
            othermonthforeground="#555555",
            bordercolor="#2A2A4A",
            relief="flat",
        )

        self.btn = ctk.CTkButton(
            self,
            text=self._fmt(),
            width=140,
            height=36,
            corner_radius=8,
            fg_color=("#F2F2F2", "#1E1E1E"),
            text_color=("#000000", "#FFFFFF"),
            hover_color=("#E5E5E5", "#2A2A2A"),
            border_width=1,
            border_color=("#CCCCCC", "#2C2C2C"),
            anchor="w",
            command=self.toggle,
        )
        self.btn.pack()

        self.panel = ctk.CTkFrame(
            self.winfo_toplevel(),
            fg_color=("#FFFFFF", "#141E2B"),
            corner_radius=12,
            border_width=1,
            border_color=("#CCCCCC", "#2A3A4A"),
        )

        self.cal = Calendar(
            self.panel,
            style="Custom.Calendar",
            selectmode="day",
            date_pattern="yyyy-mm-dd",
            year=self._date.year,
            month=self._date.month,
            day=self._date.day,
            showweeknumbers=False,
            firstweekday="monday",
            font=("Arial", 11),
            cursor="hand2",
        )
        self.cal.pack(padx=8, pady=8)
        self.cal.bind("<<CalendarSelected>>", self._select)

    def toggle(self):
        if self._open:
            self.panel.place_forget()
        else:
            self.panel.lift()
            self.panel.place(in_=self, x=0, y=self.btn.winfo_height() + 2)
        self._open = not self._open

    def _select(self, _event=None):
        selected = self.cal.get_date()
        self._date = datetime.strptime(selected, "%Y-%m-%d").date()
        self.btn.configure(text=self._fmt())
        if self._callback:
            self._callback(self._date)
        self.toggle()

    def on_change(self, callback):
        self._callback = callback

    def get_date(self):
        return self._date

    def set_date(self, d):
        self._date = d
        self.cal.selection_set(d)
        self.btn.configure(text=self._fmt())

    def _fmt(self):
        return f"📅 {self._date.strftime('%Y-%m-%d')}"

class LeaderAttendance(ctk.CTkFrame):
    def __init__(self, parent, user):
        super().__init__(parent, fg_color="transparent")
        self.db = Database()
        self.user = user
        self.team_map = {"All Teams": None}
        self.current_rows = []
        self.metric_colors = {
            "work": "#27AE60",
            "leave": "#3498DB",
            "late": "#E67E22",
            "ot": "#8B5CF6",
        }
        self.month_total_workdays = "-"
        self._detail_mode = False
        self._saved_filters = {}
        self._detail_export_context = None
        self.default_from_date, self.default_to_date = self._default_payroll_range()

        self._build_ui()

        # load data AFTER UI is ready
        self.after(300, self.load_data)

    def _build_ui(self):
        header_f = ctk.CTkFrame(self, fg_color="transparent")
        header_f.pack(fill="x", padx=80, pady=14)

        ctk.CTkLabel(
            header_f,
            text=f"📅 {self._get_team_name()}  Attendance Dashboard",
            font=("Arial", 20, "bold"),
            text_color=("#1A1A1A", "#FFFFFF")
        ).pack(side="left")

        control_card = ctk.CTkFrame(
            self,
            fg_color=("#F6F8FC", "#1F2833"),
            corner_radius=16,
            border_width=1,
            border_color=("#D6DEEB", "#3A4656"),
        )
        control_card.pack(fill="x", padx=80, pady=(0, 5))

        ctk.CTkLabel(control_card, text="Search").grid(row=0, column=0, padx=(20, 6), pady=10, sticky="w")
        self.search_entry = ctk.CTkEntry(
            control_card,
            placeholder_text="Name or ID",
            width=205,
            height=36,
            corner_radius=8,
        )
        self.search_entry.grid(row=0, column=1, padx=6, pady=4)


        ctk.CTkLabel(control_card, text="From").grid(row=0, column=4, padx=(18, 6), pady=8, sticky="w")
        self.from_date_btn = DatePickerButton(control_card, initial_date=self.default_from_date, width=180, height=36, corner_radius=8)
        self.from_date_btn.grid(row=0, column=5, padx=6, pady=8)

        ctk.CTkLabel(control_card, text="To").grid(row=0, column=6, padx=(18, 6), pady=8, sticky="w")
        self.to_date_btn = DatePickerButton(control_card, initial_date=self.default_to_date, width=180, height=36, corner_radius=8)
        self.to_date_btn.grid(row=0, column=7, padx=6, pady=8)

        ctk.CTkButton(
            control_card,
            text="🔍 Filter",
            width=65,
            height=36,
            corner_radius=8,
            fg_color=("#2563EB", "#1D4ED8"),
            hover_color=("#1D4ED8", "#1E40AF"),
             text_color=("#1A1A1A", "#FFFFFF"),
            command=self.load_data,
        ).grid(row=0, column=8, padx=6, pady=8)

        ctk.CTkButton(
            control_card,
            text="✖ Clear",
            width=65,
            height=36,
            corner_radius=8,
            fg_color=("#BABDC1", "#6B7280"),
            hover_color=("#AEB3B8", "#4B5563"),
             text_color=("#1A1A1A", "#FFFFFF"),
            command=self.reset_filters,
        ).grid(row=0, column=9, padx=6, pady=8)

        control_card.grid_columnconfigure(12, weight=1)

        ctk.CTkButton(
            control_card,
            text="📄 Export",
            width=65,
            height=36,
            corner_radius=8,
            fg_color="#DC2626",
            hover_color="#B91C1C",
            text_color=("#1A1A1A", "#FFFFFF"),
            command=self.export_pdf,
        ).grid(row=0, column=10, padx=6, pady=8)

        ctk.CTkButton(
            control_card,
            text="CSV",
            width=65,
            height=36,
            corner_radius=8,
            fg_color="#26DCB5",
            text_color=("#1A1A1A", "#FFFFFF"),
            command=self.export_excel_file,
        ).grid(row=0, column=11, padx=6, pady=8)

        self.result_label = ctk.CTkLabel(self, text="0 records | Month Work: - | Monthly Hours: -", text_color=("#334155", "#CBD5E1"))
        self.result_label.pack(anchor="w", padx=80, pady=(0, 6))

        self._create_table()

    
    def _create_table(self):
    # main container
        self.table_card = ctk.CTkFrame(
            self,
            fg_color=("#FFFFFF", "#1B1F24"),
            corner_radius=16,
            border_width=1,
            border_color=("#D6DEEB", "#2A3442"),
        )
        self.table_card.pack(fill="both", expand=True, padx=80, pady=(0, 20))
        header = ctk.CTkFrame(
            self.table_card,
            fg_color=("#BABDC1", "#6B7280"),
            corner_radius=8,
        )
        header.pack(fill="x", padx=20, pady=(10, 0))

        def h(text, w):
            return ctk.CTkLabel(
                header,
                text=text,
                width=w,
                text_color=("#000000", "#FFFFFF"),
                font=("Arial", 12, "bold")
            )

        h("Emp ID", 60).grid(row=0, column=0, padx=(10, 0), pady=8, sticky="w")
        h("Name", 180).grid(row=0, column=1, padx=0, pady=8, sticky="w")
        h("Work Days", 100).grid(row=0, column=2, padx=0, pady=8, sticky="w")
        h("Working Hours", 120).grid(row=0, column=3, padx=0, pady=8, sticky="w")
        h("Leave Days", 100).grid(row=0, column=4, padx=0, pady=8, sticky="w")
        h("Late Count", 100).grid(row=0, column=5, padx=0, pady=8, sticky="w")
        h("OT Hours", 100).grid(row=0, column=6, padx=(0, 10), pady=8, sticky="w")
        header.grid_columnconfigure(7, weight=1)
        # ・ｽ・ｽ・ｽ讚ｨ SCROLL FRAME (like AdminUsers)
        self.scroll = ctk.CTkScrollableFrame(
            self.table_card,
            fg_color="transparent"
        )
        self.scroll.pack(fill="both", expand=True, padx=10, pady=10)



    @staticmethod
    def _pretty_month_label(ym):
        return datetime.strptime(ym, "%Y-%m").strftime("%b %Y")

    @staticmethod
    def _current_payroll_month_key():
        today = datetime.now()
        if today.day >= 26:
            if today.month == 12:
                return f"{today.year + 1}-01"
            return f"{today.year}-{today.month + 1:02d}"
        return today.strftime("%Y-%m")

    @staticmethod
    def _calculate_month_total_workdays(month_key):
        if not month_key:
            return "-"

        payroll_month = datetime.strptime(month_key, "%Y-%m")
        end_date = payroll_month.replace(day=25).date()
        if payroll_month.month == 1:
            start_date = payroll_month.replace(year=payroll_month.year - 1, month=12, day=26).date()
        else:
            start_date = payroll_month.replace(month=payroll_month.month - 1, day=26).date()

        count = 0
        current = start_date
        while current <= end_date:
            if current.weekday() < 5:
                count += 1
            current += timedelta(days=1)

        return str(count)

    @staticmethod
    def _dash(value):
        return "-" if value in (None, "") else str(value)

    @staticmethod
    def _num_or_dash(value):
        return "-" if value in (None, "") else f"{float(value):g}"

    def _normalize_checkin_time(self, value):
        if value in (None, ""):
            return None
        try:
            return datetime.strptime(str(value), "%H:%M:%S").strftime("%H:%M:%S")
        except Exception:
            try:
                return datetime.strptime(str(value), "%H:%M").strftime("%H:%M:%S")
            except Exception:
                return None

    def _update_attendance_check_in(self, user_id, attendance_date, check_in_time):
        try:
            self.db.ensure_connection()
            if not self.db.conn:
                return False
            sql = "UPDATE attendance SET check_in = %s WHERE user_id = %s AND attendance_date = %s"
            self.db.cursor.execute(sql, (check_in_time, user_id, attendance_date))
            if self.db.conn:
                self.db.conn.commit()
            return True
        except Exception as e:
            messagebox.showerror("Update Error", f"Unable to update check-in: {e}")
            return False

    def _refresh_employee_detail(self):
        context = self._detail_export_context or {}
        if not context:
            return
        self._open_employee_attendance_detail(
            context.get("user_id"),
            context.get("employee_id"),
            context.get("full_name"),
            context.get("team_name"),
            from_date=context.get("from_date"),
            to_date=context.get("to_date"),
        )

    def _open_checkin_editor(self, row, parent_card):
        for widget in parent_card.grid_slaves(row=0, column=1):
            widget.destroy()

        current_value = self._dash(row.get("check_in"))
        edit_frame = ctk.CTkFrame(parent_card, fg_color="transparent")
        edit_frame.grid(row=0, column=1, padx=0, pady=8, sticky="w")

        entry = ctk.CTkEntry(edit_frame, width=130, height=30, corner_radius=10)
        if current_value != "-":
            entry.insert(0, current_value)
        entry.grid(row=0, column=0, sticky="w")
        entry.focus_set()
        entry.select_range(0, "end")
        entry.bind(
            "<Return>",
            lambda event, r=row, w=entry: self._save_check_in_edit(r, w),
        )

        save_btn = ctk.CTkButton(
            edit_frame,
            text="💾",
            width=26,
            height=26,
            corner_radius=8,
            fg_color=("#E0E7FF", "#1F2937"),
            text_color=("#1F2937", "#E0E7FF"),
            hover_color=("#C7D2FE", "#374151"),
            command=lambda r=row, w=entry: self._save_check_in_edit(r, w),
        )
        save_btn.place(in_=entry, relx=1.0, x=-4, rely=0.5, anchor="e")

    def _save_check_in_edit(self, row, entry_widget):
        new_value = entry_widget.get().strip()
        normalized = self._normalize_checkin_time(new_value) if new_value else None

        if new_value and normalized is None:
            messagebox.showerror(
                "Invalid Time",
                "Please enter a valid time in HH:MM or HH:MM:SS format.",
            )
            entry_widget.focus_set()
            return

        current_value = self._dash(row.get("check_in"))
        if (current_value == "-" and not new_value) or (normalized and current_value == normalized):
            self._refresh_employee_detail()
            return

        user_id = self._detail_export_context.get("user_id") if self._detail_export_context else None
        if not user_id:
            messagebox.showerror("Save Error", "Unable to identify the current employee record.")
            return

        updated = self._update_attendance_check_in(
            user_id,
            row.get("attendance_date"),
            normalized,
        )
        if updated:
            try:
                entry_widget.master.destroy()
            except Exception:
                pass
            self.after(50, self._refresh_employee_detail)
        else:
            messagebox.showerror("Update Error", "Unable to save check-in time.")

    @staticmethod
    def _month_hours_from_workdays(workdays):
        try:
            return f"{float(workdays) * 8:g}"
        except (TypeError, ValueError):
            return "-"

    @staticmethod
    def _default_payroll_range():
        today = datetime.now().date()
        if today.month == 1:
            start = today.replace(year=today.year - 1, month=12, day=26)
        else:
            start = today.replace(month=today.month - 1, day=26)
        end = today.replace(day=25)
        return start, end

    def _get_payroll_range(self, from_date, to_date):
        """
        Convert any selected range into payroll-style range (26 → 25 logic)
        """
        start = from_date
        end = to_date
        return start, end

    @staticmethod
    def _fmt_hours(value):
        total_minutes = int(round(max(float(value or 0), 0.0) * 60))
        hh = total_minutes // 60
        mm = total_minutes % 60
        return f"{hh}:{mm:02d}"

    @staticmethod
    def _to_time(value):
        
        if value in (None, ""):
            return None
        if isinstance(value, dt_time):
            return value
        try:
            return datetime.strptime(str(value), "%H:%M:%S").time()
        except Exception:
            try:
                return datetime.strptime(str(value), "%H:%M").time()
            except Exception:
                return None

    def _actual_hours_from_row(self, row):
        check_in = self._to_time(row.get("check_in"))
        check_out = self._effective_checkout_time(row)
        if check_in is None:
            return 0.0

        leave_type = str(row.get("leave_type") or "").lower()
        work_start = max(check_in, dt_time(7, 45))

        if leave_type == "full_day":
            return 0.0
        if leave_type == "morning_half":
            work_end = check_out if check_out is not None else dt_time(11, 45)
        elif leave_type == "evening_half":
            work_end = check_out if check_out is not None else dt_time(12, 30)
        else:
            work_end = check_out

        start_dt = datetime.combine(datetime.today(), work_start)
        end_dt = datetime.combine(datetime.today(), work_end)
        return max((end_dt - start_dt).total_seconds() / 3600.0, 0.0)

    def _effective_checkout_time(self, row):
        check_out = self._to_time(row.get("check_out"))
        try:
            has_accepted_ot_request = float(row.get("ot_hours") or 0) > 0
        except (TypeError, ValueError):
            has_accepted_ot_request = False
        if has_accepted_ot_request and check_out is not None:
            return check_out
        return dt_time(16, 30)

    def _total_hours_between_checkin_checkout(self, row):
        check_in = self._to_time(row.get("check_in"))
        check_out = self._effective_checkout_time(row)
        if check_in is None:
            return 0.0
        start_dt = datetime.combine(datetime.today(), check_in)
        end_dt = datetime.combine(datetime.today(), check_out)
        return max((end_dt - start_dt).total_seconds() / 3600.0, 0.0)

    def _stacked_hours_from_row(self, row):
        total_hours = self._total_hours_between_checkin_checkout(row)
        return max(total_hours - 8.0, 0.0)

    def _ot_hours_from_row(self, row):
        work_date = row.get("attendance_date")
        try:
            work_date_obj = datetime.strptime(str(work_date), "%Y-%m-%d").date() if work_date else None
        except Exception:
            work_date_obj = None
        is_weekend = bool(work_date_obj and work_date_obj.weekday() >= 5)
        has_accepted_ot_request = float(row.get("ot_hours") or 0) > 0
        if is_weekend and has_accepted_ot_request:
            return self._total_hours_between_checkin_checkout(row)
        stacked_hours = self._stacked_hours_from_row(row)
        return max(stacked_hours - 1.0, 0.0)

    def _late_remark_text(self, row):
        check_in = self._to_time(row.get("check_in"))
        late_cutoff = self._to_time("07:45:00")
        if check_in is None or late_cutoff is None or check_in <= late_cutoff:
            return "ON TIME", self.metric_colors["work"]
        late_mins = int((datetime.combine(datetime.today(), check_in) - datetime.combine(datetime.today(), late_cutoff)).total_seconds() // 60)
        if late_mins >= 60:
            h = late_mins // 60
            m = late_mins % 60
            if m == 0:
                late_txt = f"LATE ({h}h)"
            else:
                late_txt = f"LATE ({h}h {m}m)"
        else:
            late_txt = f"LATE ({late_mins}m)"
        return late_txt, self.metric_colors["late"]

    @staticmethod
    def _accepted_leave_days_from_row(row):
        leave_type = str(row.get("leave_type") or "").lower()
        if leave_type == "full_day":
            return 1.0
        if leave_type in ("morning_half", "evening_half"):
            return 0.5
        return 0.0

    def _get_leave_days(self, user_id, from_date, to_date):
        sql = """
            SELECT IFNULL(SUM(
                CASE
                    WHEN total_days IS NOT NULL THEN total_days
                    ELSE CASE
                        WHEN DATEDIFF(end_date, start_date) = 0 THEN
                            CASE
                                WHEN start_shift = 'Full Day' OR end_shift = 'Full Day' THEN 1.0
                                WHEN start_shift = end_shift THEN 0.5
                                ELSE 1.0
                            END
                        ELSE (DATEDIFF(end_date, start_date) + 1)
                             - (CASE WHEN start_shift = 'Full Day' THEN 0 ELSE 0.5 END)
                             - (CASE WHEN end_shift = 'Full Day' THEN 0 ELSE 0.5 END)
                        END
                END
            ), 0) AS leave_days
            FROM leave_requests
            WHERE user_id = %s
              AND status = 'Approved'
              AND start_date <= %s
              AND end_date >= %s
        """
        self.db.cursor.execute(sql, (user_id, to_date, from_date))
        row = self.db.cursor.fetchone() or {"leave_days": 0}
        return float(row["leave_days"] or 0)

    def _get_team_name(self):
        try:
            sql = "SELECT team_name FROM teams WHERE team_id = %s"
            self.db.cursor.execute(sql, (self.user["team_id"],))
            row = self.db.cursor.fetchone()

            if row and row["team_name"]:
                return row["team_name"]

        except Exception as e:
            print("Team name fetch error:", e)

        return "Team"
    def _get_employee_attendance_rows(self, user_id, from_date=None, to_date=None):
        sql = """
        SELECT
            d.work_date AS attendance_date,
            a.check_in,
            a.check_out,
            IFNULL(o.ot_hours, 0) AS ot_hours,
            CASE
                WHEN EXISTS (
                    SELECT 1
                    FROM leave_requests lr
                    WHERE lr.user_id = %s
                      AND lr.status = 'Approved'
                      AND d.work_date BETWEEN lr.start_date AND lr.end_date
                ) THEN 1
                WHEN a.check_in IS NOT NULL OR a.check_out IS NOT NULL THEN 1
                WHEN IFNULL(o.ot_hours, 0) > 0 THEN 1
                ELSE 0
            END AS workdays,
            CASE WHEN a.check_in IS NOT NULL AND a.check_in > '07:45:00' THEN 1 ELSE 0 END AS is_late,
            CASE
                WHEN EXISTS (
                    SELECT 1
                    FROM leave_requests lr
                    WHERE lr.user_id = %s
                      AND lr.status = 'Approved'
                      AND d.work_date BETWEEN lr.start_date AND lr.end_date
                      AND (
                          (lr.start_date = d.work_date AND lr.end_date = d.work_date AND lr.start_shift = 'Morning' AND lr.end_shift = 'Morning')
                          OR (lr.end_date = d.work_date AND lr.end_date > lr.start_date AND lr.end_shift = 'Morning')
                      )
                ) THEN 'morning_half'
                WHEN EXISTS (
                    SELECT 1
                    FROM leave_requests lr
                    WHERE lr.user_id = %s
                      AND lr.status = 'Approved'
                      AND d.work_date BETWEEN lr.start_date AND lr.end_date
                      AND (
                          (lr.start_date = d.work_date AND lr.end_date = d.work_date AND lr.start_shift = 'Evening' AND lr.end_shift = 'Evening')
                          OR (lr.start_date = d.work_date AND lr.end_date > lr.start_date AND lr.start_shift = 'Evening')
                      )
                ) THEN 'evening_half'
                WHEN EXISTS (
                    SELECT 1
                    FROM leave_requests lr
                    WHERE lr.user_id = %s
                      AND lr.status = 'Approved'
                      AND d.work_date BETWEEN lr.start_date AND lr.end_date
                ) THEN 'full_day'
                ELSE ''
            END AS leave_type
        FROM (
            SELECT attendance_date AS work_date
            FROM attendance
            WHERE user_id = %s
            UNION
            SELECT DATE(ot_date) AS work_date
            FROM overtime_requests
            WHERE member_id = %s
            AND status IN ('Accepted', 'Approved')
        ) d
        LEFT JOIN attendance a
            ON a.user_id = %s
            AND a.attendance_date = d.work_date
        LEFT JOIN (
            SELECT DATE(ot_date) AS ot_date, ROUND(SUM(COALESCE(hours, 0)), 2) AS ot_hours
            FROM overtime_requests
            WHERE member_id = %s
            AND status IN ('Accepted', 'Approved')
            GROUP BY DATE(ot_date)
        ) o ON d.work_date = o.ot_date
        WHERE d.work_date IS NOT NULL
        """
        params = [user_id, user_id, user_id, user_id, user_id, user_id, user_id, user_id]

        if from_date and to_date:
            if from_date > to_date:
                from_date, to_date = to_date, from_date
            sql += """
            AND d.work_date BETWEEN %s AND %s
            """
            params.extend([from_date, to_date])

        sql += " ORDER BY d.work_date DESC"
        self.db.cursor.execute(sql, tuple(params))
        return self.db.cursor.fetchall()

    def _open_employee_attendance_detail(self, user_id, employee_id, full_name, team_name=None, from_date=None, to_date=None):
        # allow caller to pass a saved from/to date (used after inline edit refresh)
        if from_date is None or to_date is None:
            try:
                from_date = self.from_date_btn.get_date()
                to_date = self.to_date_btn.get_date()
            except Exception:
                from_date, to_date = self._default_payroll_range()

        if from_date > to_date:
            from_date, to_date = to_date, from_date

        from_date, to_date = self._get_payroll_range(from_date, to_date)
        month_label = f"{from_date} to {to_date}"

        try:
            rows = self._get_employee_attendance_rows(user_id, from_date, to_date)
        except Exception as e:
            messagebox.showerror("Fetch Error", f"Cannot load attendance detail: {e}")
            return
        user_workdays = sum(int(r.get("workdays", 0) or 0) for r in rows)
        total_ot_hours = sum(self._ot_hours_from_row(r) for r in rows)
        actual_hours = sum(self._actual_hours_from_row(r) for r in rows)
        actual_hours_with_ot = actual_hours + total_ot_hours
        late_days = sum(int(r.get("is_late", 0) or 0) for r in rows)
        month_workdays = int(self.month_total_workdays or 0) if str(self.month_total_workdays).isdigit() else 0
        month_hours = float(self._month_hours_from_workdays(month_workdays))
        leave_days = self._get_leave_days(user_id, from_date, to_date)

        search_value = None
        try:
            search_value = self.search_entry.get()
        except Exception:
            search_value = self._saved_filters.get("search", "") if hasattr(self, "_saved_filters") else ""

        self._saved_filters = {
            "search": search_value,
            "month": month_label,
        }
        # store context for refresh after inline edits
        self._detail_export_context = {
            "user_id": user_id,
            "employee_id": employee_id,
            "full_name": full_name,
            "team_name": team_name,
            "month_label": month_label,
            "rows": rows,
            "from_date": from_date,
            "to_date": to_date,
        }
        self._detail_mode = True
        for child in self.winfo_children():
            child.destroy()

        top_bar = ctk.CTkFrame(self, fg_color="transparent")
        top_bar.pack(fill="x", padx=80, pady=(12, 8))
        ctk.CTkButton(
            top_bar,
            text="← Back",
            width=80,
            fg_color=("#DBDBDB", "#333333"),
            text_color=("black", "white"),
            hover_color=("#CFCFCF", "#444444"),
            corner_radius=8,
            command=self._return_from_detail_page,
        ).pack(side="left")
        ctk.CTkButton(
            top_bar,
            text="📄 Export",
            width=85,
            height=36,
            corner_radius=8,
            fg_color="#DC2626",
            hover_color="#B91C1C",
            command=self._export_employee_detail_pdf,
        ).pack(side="right")
        ctk.CTkButton(
            top_bar,
            text="CSV",
            width=78,
            height=36,
            corner_radius=8,
            fg_color="#26DCB5",
            command=self._export_employee_detail_excel,
        ).pack(side="right", padx=(0, 8))

        self._detail_export_context = {
            "user_id": user_id,
            "employee_id": employee_id,
            "full_name": full_name,
            "team_name": team_name,
            "month_label": month_label,
            "rows": rows,
        }

        ctk.CTkLabel(
            self,
            text=f"{full_name} (Employee ID: {employee_id})",
            font=("Arial", 18, "bold"),
        ).pack(anchor="w", padx=80, pady=(4, 4))

        stats_container = ctk.CTkFrame(self, fg_color="transparent")
        stats_container.pack(fill="x", padx=80, pady=(0, 10))

        def stat_card(title, value, color):
            card = ctk.CTkFrame(stats_container, corner_radius=10, width=160)
            card.pack(side="left", padx=(0, 12))
            ctk.CTkLabel(card, text=title).pack(padx=14, pady=(5, 0))
            ctk.CTkLabel(
                card,
                text=value,
                font=("Arial", 18, "bold"),
                text_color=color,
            ).pack(pady=(0, 5))

        stat_card("Month Work", f"{month_workdays:g}", self.metric_colors["work"])
        stat_card("Monthly Hours", f"{month_hours:g}", self.metric_colors["work"])
        stat_card("Actual Hours", f"{actual_hours_with_ot:g}", self.metric_colors["work"])
        stat_card("Working Days", f"{user_workdays:g}", self.metric_colors["work"])
        stat_card("Leave Days", f"{leave_days:g}", self.metric_colors["leave"])
        stat_card("Late Days", f"{late_days:g}", self.metric_colors["late"])
        stat_card("OT Hours", f"{total_ot_hours:g}", self.metric_colors["ot"])

        table_card = ctk.CTkFrame(
            self,
            fg_color=("#FFFFFF", "#1B1F24"),
            corner_radius=12,
            border_width=1,
            border_color=("#D6DEEB", "#2A3442"),
        )
        table_card.pack(fill="both", expand=True, padx=80, pady=(0, 20))

        header = ctk.CTkFrame(table_card, fg_color=("#BABDC1", "#6B7280"), corner_radius=8)
        header.pack(fill="x", padx=20, pady=(10, 0))

        def h(text, w):
            return ctk.CTkLabel(
                header,
                text=text,
                width=w,
                anchor="w",
                text_color=("#000000", "#FFFFFF"),
                font=("Arial", 12, "bold"),
            )

        h("Date", 170).grid(row=0, column=0, padx=(20, 0), pady=8, sticky="w")
        h("Check-In", 140).grid(row=0, column=1, padx=0, pady=8, sticky="w")
        h("Check-Out", 140).grid(row=0, column=2, padx=0, pady=8, sticky="w")
        h("Total Hours", 120).grid(row=0, column=3, padx=0, pady=8, sticky="w")
        h("Stacked Hours", 130).grid(row=0, column=4, padx=0, pady=8, sticky="w")
        h("OT Hours", 120).grid(row=0, column=5, padx=10, pady=8, sticky="w")
        h("Remark", 120).grid(row=0, column=6, padx=(10, 0), pady=8, sticky="w")
        header.grid_columnconfigure(7, weight=1)

        scroll = ctk.CTkScrollableFrame(table_card, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=10, pady=10)

        if not rows:
            ctk.CTkLabel(
                scroll,
                text="No attendance records found for the selected month.",
                text_color=("#64748B", "#94A3B8"),
            ).pack(pady=20)
            return

        for row in rows:
            card = ctk.CTkFrame(scroll, fg_color=("#EEF2F7", "#1F2933"), corner_radius=10)
            card.pack(fill="x", pady=4, padx=4)

            date_value = self._dash(row["attendance_date"])
            check_in_value = self._dash(row["check_in"])
            check_out_value = self._dash(row["check_out"])
            total_hours = self._total_hours_between_checkin_checkout(row)
            stacked_hours = self._stacked_hours_from_row(row)
            ot_value = f"{self._ot_hours_from_row(row):g}"
            remark, remark_color = self._late_remark_text(row)

            ctk.CTkLabel(card, text=date_value, width=170, anchor="w").grid(row=0, column=0, padx=(10, 0), pady=8, sticky="w")
            checkin_widget = ctk.CTkLabel(card, text=check_in_value, width=140, anchor="w")
            checkin_widget.grid(row=0, column=1, padx=0, pady=8, sticky="w")
            checkin_widget.configure(cursor="hand2")
            checkin_widget.bind("<Button-1>", lambda _event, r=row, w=card: self._open_checkin_editor(r, w))
            ctk.CTkLabel(card, text=check_out_value, width=140, anchor="w").grid(row=0, column=2, padx=0, pady=8, sticky="w")
            ctk.CTkLabel(card, text=f"{total_hours:g}", width=120, anchor="w").grid(row=0, column=3, padx=0, pady=8, sticky="w")
            ctk.CTkLabel(card, text=f"{stacked_hours:g}", width=130, anchor="w").grid(row=0, column=4, padx=0, pady=8, sticky="w")
            ctk.CTkLabel(card, text=ot_value, width=120, text_color=self.metric_colors["ot"]).grid(row=0, column=5, padx=0, pady=8, sticky="w")
            ctk.CTkLabel(card, text=remark, width=120, text_color=remark_color).grid(row=0, column=6, padx=(0, 0), pady=8, sticky="w")
            card.grid_columnconfigure(7, weight=1)

    def _return_from_detail_page(self):
        self._detail_mode = False
        for child in self.winfo_children():
            child.destroy()

        self._build_ui()
        if self._saved_filters:
            self.search_entry.delete(0, "end")
            self.search_entry.insert(0, self._saved_filters.get("search", ""))

        self.load_data()

    def _export_employee_detail_pdf(self):
        context = self._detail_export_context or {}
        rows = context.get("rows", [])
        if not rows:
            messagebox.showwarning("No Data", "No detail records to export.")
            return

        full_name = context.get("full_name", "Employee")
        employee_id = context.get("employee_id", "-")
        month_label = context.get("month_label", "Month")
        safe_name = str(full_name).replace(" ", "_")
        safe_month = str(month_label).replace(" ", "_")
        file_name = f"AttendanceDetail_{safe_name}_{safe_month}.pdf"

        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            initialfile=file_name,
            filetypes=[("PDF files", "*.pdf")],
        )
        if not file_path:
            return

        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        except Exception as e:
            messagebox.showerror("Missing Dependency", f"Cannot export PDF: {e}")
            return

        try:
            doc = SimpleDocTemplate(file_path, pagesize=letter, leftMargin=28, rightMargin=28, topMargin=28, bottomMargin=28)
            styles = getSampleStyleSheet()
            elements = []

            title_style = ParagraphStyle("DetailTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=20, leading=24, textColor=colors.HexColor("#0B1220"), alignment=1, spaceAfter=6)
            meta_style = ParagraphStyle("DetailMeta", parent=styles["Normal"], fontName="Helvetica", fontSize=10, leading=13, textColor=colors.HexColor("#334155"))

            user_workdays = sum(int(r.get("workdays", 0) or 0) for r in rows)
            total_ot_hours = sum(self._ot_hours_from_row(r) for r in rows)
            actual_hours = sum(self._actual_hours_from_row(r) for r in rows)
            actual_hours_with_ot = actual_hours + total_ot_hours
            team_label = context.get("team_name") or self._leader_team_label()
            safe_team_label = escape(str(team_label))
            safe_month_label = escape(str(month_label))
            month_workdays = int(self.month_total_workdays or 0) if str(self.month_total_workdays).isdigit() else 0
            from_date = context.get("from_date", self.from_date_btn.get_date())
            to_date = context.get("to_date", self.to_date_btn.get_date())
            if from_date and to_date and from_date > to_date:
                from_date, to_date = to_date, from_date
            leave_days = self._get_leave_days(context.get("user_id"), from_date, to_date)
            safe_month_work = escape(str(month_workdays))
            safe_actual_work = escape(f"{user_workdays:g}")
            safe_leave = escape(str(leave_days))
            safe_month_hours = escape(str(self._month_hours_from_workdays(month_workdays)))
            safe_actual_hours = escape(f"{actual_hours_with_ot:g}")
            total_ot_hours = sum(self._ot_hours_from_row(r) for r in rows)
            safe_ot_hours=escape(self._fmt_hours(total_ot_hours))

            elements.append(Paragraph("Employee Attendance Detail Report", title_style))
            elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", meta_style))
            elements.append(Spacer(1, 10))

            info_table = Table(
                [[
                    Paragraph(
                        f"<b><font color='#1D4ED8'>Month:</font></b> {safe_month_label}<br/>"
                        f"<b><font color='#1D4ED8'>Team:</font></b> {safe_team_label}<br/>"
                        f"<b><font color='#1D4ED8'>Employee:</font></b> {escape(str(full_name))}<br/>"
                        f"<b><font color='#1D4ED8'>Employee ID:</font></b> {escape(str(employee_id))}",
                        meta_style
                    ),
                    Paragraph(
                        f"<b><font color='#1D4ED8'>Month Work:</font></b> {safe_month_work}<br/>"
                        f"<b><font color='#1D4ED8'>Actual Work:</font></b> {safe_actual_work}<br/>"
                        f"<b><font color='#1D4ED8'>Leave:</font></b> {safe_leave}",
                        meta_style
                    ),
                    Paragraph(
                        f"<b><font color='#1D4ED8'>Month Hours:</font></b> {safe_month_hours}<br/>"
                        f"<b><font color='#1D4ED8'>Actual Hours:</font></b> {safe_actual_hours}<br/>"
                        f"<b><font color='#1D4ED8'>OT Hours:</font></b> {safe_ot_hours}",
                        meta_style
                    ),
                ]],
                colWidths=[2.45 * inch, 2.35 * inch, 2.2 * inch],
            )
            info_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#CBD5E1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#E2E8F0")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 10))

            table_data = [["Date", "Check-In", "Check-Out", "OT Hours", "Remark"]]
            for row in rows:
                is_late = bool(row["is_late"])
                remark = "LATE" if is_late else "ON TIME"
                remark_color = "#EA580C" if remark == "LATE" else "#16A34A"
                table_data.append([
                    self._dash(row["attendance_date"]),
                    self._dash(row["check_in"]),
                    self._dash(row["check_out"]),
                    f"{self._ot_hours_from_row(row):g}",
                    Paragraph(f"<font color='{remark_color}'>{remark}</font>", meta_style),
                ])

            table = Table(table_data, repeatRows=1, colWidths=[1.7 * inch, 1.45 * inch, 1.45 * inch, 1.0 * inch, 1.6 * inch])
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#FFFFFF"), colors.HexColor("#F8FAFC")]),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ("TEXTCOLOR", (3, 1), (3, -1), colors.HexColor("#7C3AED")),
            ]))
            elements.append(table)

            doc.build(elements)
            messagebox.showinfo("Export Successful", "Detail attendance PDF exported successfully.")
        except Exception as e:
            messagebox.showerror("Export Error", f"PDF export failed: {e}")

    def _leader_team_label(self):
        for row in self.current_rows:
            team_name = row.get("team_name")
            if team_name not in (None, "", "-"):
                return str(team_name)

        if self.user.get("team_name"):
            return str(self.user["team_name"])
        if self.user.get("team_id") is not None:
            return f"Team_{self.user['team_id']}"
        return "Team"
    def reset_filters(self):
        self.search_entry.delete(0, "end")
        from_date, to_date = self._default_payroll_range()
        self.from_date_btn.set_date(from_date)
        self.to_date_btn.set_date(to_date)
        self.load_data()

    def export_excel_file(self):
        if not self.current_rows:
            messagebox.showwarning("No Data", "No records to export.")
            return

        selected_month = f"{self.from_date_btn.get_date()}_{self.to_date_btn.get_date()}"
        selected_team = self._leader_team_label().replace(" ", "_")
        file_name = f"Attendance_{selected_team}_{selected_month}.xlsx"

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=file_name,
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not file_path:
            return
        if not file_path.lower().endswith(".xlsx"):
            file_path = f"{file_path}.xlsx"

        self._export_excel(file_path)

    def load_data(self):
        # clear old rows (scroll frame)
        for w in self.scroll.winfo_children():
            w.destroy()

        search_query = self.search_entry.get().strip()
        from_date = self.from_date_btn.get_date()
        to_date = self.to_date_btn.get_date()
        if from_date > to_date:
            from_date, to_date = to_date, from_date

        from_date, to_date = self._get_payroll_range(from_date, to_date)
        self.month_total_workdays = str(
            sum(
                1
                for i in range((to_date - from_date).days + 1)
                if (from_date + timedelta(days=i)).weekday() < 5
            )
        )

        att_where = "WHERE attendance_date BETWEEN %s AND %s"
        leave_where = "WHERE status = 'Approved' AND start_date <= %s AND end_date >= %s"
        ot_where = "WHERE status IN ('Accepted', 'Approved') AND DATE(ot_date) BETWEEN %s AND %s"

        att_params = [from_date, to_date]
        leave_params = [to_date, from_date]
        ot_params = [from_date, to_date]

        # ================= SEARCH FILTER =================
        outer_where = "WHERE u.team_id = %s AND (u.full_name LIKE %s OR CAST(u.employee_id AS CHAR) LIKE %s)"
        outer_params = [
            self.user["team_id"],
            f"%{search_query}%",
            f"%{search_query}%",
        ]

        sql = f"""
            SELECT
                u.id AS user_id,
                u.employee_id,
                u.full_name,
                IFNULL(t.team_name, '-') AS team_name,
                IFNULL(a.workdays, 0) AS workdays,
                IFNULL(lr.leaveday, 0) AS leaveday,
                IFNULL(a.latecount, 0) AS latecount,
                IFNULL(o.overtimehour, 0) AS overtimehour
            FROM users u
            LEFT JOIN teams t ON u.team_id = t.team_id

            LEFT JOIN (
                SELECT
                    user_id,
                    MAX(attendance_date) AS latest_date,
                    COUNT(DISTINCT attendance_date) AS workdays,
                    SUM(CASE WHEN check_in > '07:45:00' THEN 1 ELSE 0 END) AS latecount
                FROM attendance
                {att_where}
                GROUP BY user_id
            ) a ON u.id = a.user_id

            LEFT JOIN (
                SELECT
                    user_id,
                    SUM(
                        CASE
                            WHEN total_days IS NOT NULL THEN total_days
                            ELSE CASE
                                WHEN DATEDIFF(end_date, start_date) = 0 THEN
                                    CASE
                                        WHEN start_shift = 'Full Day' OR end_shift = 'Full Day' THEN 1.0
                                        WHEN start_shift = end_shift THEN 0.5
                                        ELSE 1.0
                                    END
                                ELSE (DATEDIFF(end_date, start_date) + 1)
                                     - (CASE WHEN start_shift = 'Full Day' THEN 0 ELSE 0.5 END)
                                     - (CASE WHEN end_shift = 'Full Day' THEN 0 ELSE 0.5 END)
                                END
                        END
                    ) AS leaveday
                FROM leave_requests
                {leave_where}
                GROUP BY user_id
            ) lr ON u.id = lr.user_id

            LEFT JOIN (
                SELECT
                    member_id,
                    ROUND(SUM(COALESCE(hours, 0)), 2) AS overtimehour
                FROM overtime_requests
                {ot_where}
                GROUP BY member_id
            ) o ON u.id = o.member_id

            WHERE u.team_id = %s
            AND (a.latest_date IS NOT NULL OR o.overtimehour IS NOT NULL)
            AND (u.full_name LIKE %s OR CAST(u.employee_id AS CHAR) LIKE %s)

            ORDER BY u.employee_id
        """

        params = att_params + leave_params + ot_params + outer_params

        try:
            self.db.cursor.execute(sql, tuple(params))
            self.current_rows = self.db.cursor.fetchall()
            if not self.current_rows:
                empty_frame = ctk.CTkFrame(
                    self.scroll,
                    fg_color="transparent"
                )

                empty_frame.pack(
                    fill="both",
                    expand=True,
                    pady=40
                )

                ctk.CTkLabel(
                    empty_frame,
                    text="There is no Records",
                    font=("Arial", 18, "bold"),
                    text_color=("#6B7280", "#9CA3AF")
                ).pack(pady=(20, 5))

                ctk.CTkLabel(
                    empty_frame,
                    text="Try changing the filter date range.",
                    font=("Arial", 12),
                    text_color=("#9CA3AF", "#6B7280")
                ).pack()

                return

            # ================= CREATE CARDS =================
            for row in self.current_rows:
                detail_rows = self._get_employee_attendance_rows(row["user_id"], from_date, to_date)
                row_workdays_with_ot = sum(
                    int(r.get("workdays", 0) or 0)
                    for r in detail_rows
                )
                row_working_hours = sum(
                    self._actual_hours_from_row(r)
                    for r in detail_rows
                )
                row_ot_hours = sum(self._ot_hours_from_row(r) for r in detail_rows)

                card = ctk.CTkFrame(
                    self.scroll,
                    fg_color=("#EEF2F7", "#1F2933"),
                    corner_radius=10,
                )
                card.pack(fill="x", pady=4, padx=5)

                # row content
                ctk.CTkLabel(card, text=self._dash(row["employee_id"]), width=60, anchor="w").grid(row=0, column=0, padx=(10, 0), pady=8, sticky="w")
                ctk.CTkButton(
                    card,
                    text=self._dash(row["full_name"]),
                    width=180,
                    anchor="w",
                    fg_color="transparent",
                    hover_color=("#DBEAFE", "#1E3A5F"),
                    text_color=("#1D4ED8", "#60A5FA"),
                    font=("Arial", 12),
                    cursor="hand2",
                    command=lambda user_id=row["user_id"], employee_id=row["employee_id"], full_name=row["full_name"], team_name=row.get("team_name"):
                        self._open_employee_attendance_detail(user_id, employee_id, full_name, team_name),
                ).grid(row=0, column=1, padx=0, pady=8, sticky="w")
                ctk.CTkLabel(card, text=f"{row_workdays_with_ot:g}", width=80, text_color=self.metric_colors["work"]).grid(row=0, column=2, padx=0, pady=8, sticky="w")
                ctk.CTkLabel(card, text=f"{row_working_hours:g}", width=100, text_color=self.metric_colors["work"]).grid(row=0, column=3, padx=0, pady=8, sticky="w")
                ctk.CTkLabel(card, text=self._num_or_dash(row["leaveday"]), width=80, text_color=self.metric_colors["leave"]).grid(row=0, column=4, padx=0, pady=8, sticky="w")
                ctk.CTkLabel(card, text=self._dash(row["latecount"]), width=80, text_color=self.metric_colors["late"]).grid(row=0, column=5, padx=30, pady=8, sticky="w")
                ctk.CTkLabel(card, text=self._num_or_dash(row_ot_hours), width=100, text_color=self.metric_colors["ot"]).grid(row=0, column=6, padx=(20, 0), pady=8, sticky="w")
                card.grid_columnconfigure(7, weight=1)

                # bottom border line (like AdminUsers style)
                border = ctk.CTkFrame(card, height=1, fg_color=("#D6DEEB", "#334155"))
                border.grid(row=1, column=0, columnspan=8, sticky="ew", pady=(6, 0))

            month_hours = self._month_hours_from_workdays(self.month_total_workdays)
            self.result_label.configure(text=f"{len(self.current_rows)} records | Month Work: {self.month_total_workdays} | Monthly Hours: {month_hours}")

        except Exception as e:
            print("Fetch Error:", e)
            self.current_rows = []
            month_hours = self._month_hours_from_workdays(self.month_total_workdays)
            self.result_label.configure(text=f"0 records | Month Work: {self.month_total_workdays} | Monthly Hours: {month_hours}")

    def export_pdf(self):
        if not self.current_rows:
            messagebox.showwarning("No Data", "No records to export.")
            return

        team_label = self._leader_team_label()
        selected_month = f"{self.from_date_btn.get_date()}_{self.to_date_btn.get_date()}"
        selected_team = team_label.replace(" ", "_")
        file_name = f"Attendance_{selected_team}_{selected_month}.pdf"

        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            initialfile=file_name,
            filetypes=[("PDF files", "*.pdf"), ("Excel files", "*.xlsx")],
        )
        if not file_path:
            return
        if file_path.lower().endswith(".xlsx"):
            self._export_excel(file_path)
            return

        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        except Exception as e:
            messagebox.showerror("Missing Dependency", f"Cannot export PDF: {e}")
            return

        try:
            doc = SimpleDocTemplate(
                file_path,
                pagesize=letter,
                leftMargin=28,
                rightMargin=28,
                topMargin=28,
                bottomMargin=28,
            )
            styles = getSampleStyleSheet()
            elements = []
            safe_team_label = escape(team_label)
            safe_month_label = escape(f"{self.from_date_btn.get_date()} to {self.to_date_btn.get_date()}")
            month_workdays = int(self.month_total_workdays or 0) if str(self.month_total_workdays).isdigit() else 0
            month_hours = float(self._month_hours_from_workdays(month_workdays))
            safe_month_work = escape(str(month_workdays))
            safe_month_hours = escape(f"{month_hours:g}")

            title_style = ParagraphStyle(
                "LeaderTitle",
                parent=styles["Title"],
                fontName="Helvetica-Bold",
                fontSize=22,
                leading=26,
                textColor=colors.HexColor("#0B1220"),
                alignment=1,
                spaceAfter=6,
            )
            meta_style = ParagraphStyle(
                "LeaderMeta",
                parent=styles["Normal"],
                fontName="Helvetica",
                fontSize=10,
                leading=13,
                textColor=colors.HexColor("#334155"),
            )

            elements.append(Paragraph("Employee Attendance Report", title_style))
            elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", meta_style))
            elements.append(Spacer(1, 10))

            from_date = self.from_date_btn.get_date()
            to_date = self.to_date_btn.get_date()
            if from_date > to_date:
                from_date, to_date = to_date, from_date
            from_date, to_date = self._get_payroll_range(from_date, to_date)

            info_table = Table(
                [[
                    Paragraph(
                        f"<b><font color='#1D4ED8'>Month:</font></b> {safe_month_label}<br/>"
                        f"<b><font color='#1D4ED8'>Team:</font></b> {safe_team_label}<br/>"
                        f"<b><font color='#1D4ED8'>Employee:</font></b> All",
                        meta_style
                    ),
                    Paragraph(
                        f"<b><font color='#1D4ED8'>Month Work:</font></b> {safe_month_work}<br/>"
                        f"<b><font color='#1D4ED8'>Month Hours:</font></b> {safe_month_hours}",
                        meta_style
                    ),
                ]],
                colWidths=[3.5 * inch, 3.5 * inch],
            )
            info_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.darkblue),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#CBD5E1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#E2E8F0")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 18))

            table_data = [[
                "Emp ID",
                "Employee",
                "Work Days",
                "Working Hours",
                "Leave Days",
                "Late Count",
                "OT Hours",
            ]]
            for row in self.current_rows:
                detail_rows = self._get_employee_attendance_rows(row["user_id"], from_date, to_date)
                row_workdays_with_ot = sum(int(r.get("workdays", 0) or 0) for r in detail_rows)
                row_working_hours = sum(self._actual_hours_from_row(r) for r in detail_rows)
                row_ot_hours = sum(self._ot_hours_from_row(r) for r in detail_rows)
                table_data.append([
                    self._dash(row["employee_id"]),
                    self._dash(row["full_name"]),
                    f"{row_workdays_with_ot:g}",
                    f"{row_working_hours:g}",
                    self._num_or_dash(row["leaveday"]),
                    self._dash(row["latecount"]),
                    self._num_or_dash(row_ot_hours),
                ])

            table = Table(
                table_data,
                repeatRows=1,
                colWidths=[0.9 * inch, 2.4 * inch, 1.0 * inch, 1.2 * inch, 1.05 * inch, 1.0 * inch, 1.0 * inch],
            )
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("ALIGN", (1, 1), (1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#FFFFFF"), colors.HexColor("#F8FAFC")]),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ("TEXTCOLOR", (2, 1), (2, -1), colors.HexColor("#16A34A")),
                ("TEXTCOLOR", (3, 1), (3, -1), colors.HexColor("#0284C7")),
                ("TEXTCOLOR", (4, 1), (4, -1), colors.HexColor("#EA580C")),
                ("TEXTCOLOR", (5, 1), (5, -1), colors.HexColor("#7C3AED")),
            ]))
            elements.append(table)

            doc.build(elements)
            messagebox.showinfo("Export Successful", "Attendance PDF exported successfully.")
        except Exception as e:
            messagebox.showerror("Export Error", f"PDF export failed: {e}")

    def _export_employee_detail_excel(self):

        context = self._detail_export_context or {}
        rows = context.get("rows", [])

        if not rows:
            messagebox.showwarning(
                "No Data",
                "No detail records to export."
            )
            return

        full_name = context.get("full_name", "Employee")
        month_label = context.get("month_label", "Month")

        safe_name = str(full_name).replace(" ", "_")
        safe_month = str(month_label).replace(" ", "_")

        file_name = f"AttendanceDetail_{safe_name}_{safe_month}.xlsx"

        team_label = (
            context.get("team_name")
            or self._leader_team_label()
        )

        employee_id = context.get("employee_id", "-")

        working_days = sum(
            int(r.get("workdays", 0) or 0)
            for r in rows
        )

        working_hours = sum(
            self._actual_hours_from_row(r)
            for r in rows
        )

        ot_hours = sum(
            self._ot_hours_from_row(r)
            for r in rows
        )

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=file_name,
            filetypes=[("Excel files", "*.xlsx")],
        )

        if not file_path:
            return

        if not file_path.lower().endswith(".xlsx"):
            file_path = f"{file_path}.xlsx"

        try:
            from openpyxl import Workbook

            from openpyxl.styles import (
                Font,
                PatternFill,
                Alignment,
                Border,
                Side,
            )

            from openpyxl.utils import get_column_letter

        except Exception as e:

            messagebox.showerror(
                "Missing Dependency",
                f"Cannot export Excel: {e}"
            )

            return

        try:

            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance Detail"

            # ================= REPORT INFO =================
            info_row = 1

            label_font = Font(
                bold=True,
                color="0F172A"
            )

            info_fill = PatternFill(
                start_color="E2E8F0",
                end_color="E2E8F0",
                fill_type="solid"
            )

            info_cells = [
                (f"A{info_row}", "Team"),
                (f"A{info_row+1}", "Employee ID"),
                (f"A{info_row+2}", "Employee"),
                (f"D{info_row}", "Working Days"),
                (f"D{info_row+1}", "Working Hours"),
                (f"D{info_row+2}", "OT Hours"),
            ]

            for cell_ref, value in info_cells:

                cell = ws[cell_ref]

                cell.value = value

                cell.font = label_font

                cell.fill = info_fill

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            ws[f"B{info_row}"] = team_label
            ws[f"B{info_row+1}"] = str(employee_id)
            ws[f"B{info_row+2}"] = full_name

            ws[f"E{info_row}"] = working_days
            ws[f"E{info_row+1}"] = self._fmt_hours(working_hours)
            ws[f"E{info_row+2}"] = self._fmt_hours(ot_hours)

            # ================= TABLE HEADER =================
            header_row = 6

            headers = [
                "Date",
                "Check-In",
                "Check-Out",
                "OT Hours",
                "Remark",
            ]

            for col_num, header in enumerate(headers, 1):

                cell = ws.cell(
                    row=header_row,
                    column=col_num
                )

                cell.value = header

                cell.font = Font(
                    bold=True,
                    color="FFFFFF"
                )

                cell.fill = PatternFill(
                    start_color="0F172A",
                    end_color="0F172A",
                    fill_type="solid"
                )

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            # ================= DATA ROWS =================
            thin = Side(
                border_style="thin",
                color="D1D5DB"
            )

            data_start = header_row + 1

            for idx, row in enumerate(rows, start=data_start):

                is_late = bool(row.get("is_late"))

                values = [
                    self._dash(row.get("attendance_date")),
                    self._dash(row.get("check_in")),
                    self._dash(row.get("check_out")),
                    f"{self._ot_hours_from_row(row):g}",
                    "LATE" if is_late else "ON TIME",
                ]

                for col_num, value in enumerate(values, 1):

                    cell = ws.cell(
                        row=idx,
                        column=col_num
                    )

                    cell.value = value

                    cell.border = Border(
                        left=thin,
                        right=thin,
                        top=thin,
                        bottom=thin
                    )

                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )

                    # alternating row color
                    if idx % 2 == 0:

                        cell.fill = PatternFill(
                            start_color="F8FAFC",
                            end_color="F8FAFC",
                            fill_type="solid"
                        )

            # ================= COLUMN WIDTH =================
            widths = {
                1: 18,
                2: 18,
                3: 18,
                4: 15,
                5: 18,
            }

            for col_num, width in widths.items():

                ws.column_dimensions[
                    get_column_letter(col_num)
                ].width = width

            # ================= ROW HEIGHT =================
            ws.row_dimensions[1].height = 30

            # ================= SAVE =================
            wb.save(file_path)

            messagebox.showinfo(
                "Export Successful",
                "Detail attendance Excel exported successfully."
            )

        except Exception as e:

            messagebox.showerror(
                "Export Error",
                f"Excel export failed: {e}"
            )

    def _export_excel(self, file_path):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except Exception as e:
            messagebox.showerror("Missing Dependency", f"Cannot export Excel: {e}")
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance Report"

            from_date = self.from_date_btn.get_date()
            to_date = self.to_date_btn.get_date()
            if from_date > to_date:
                from_date, to_date = to_date, from_date
            from_date, to_date = self._get_payroll_range(from_date, to_date)

            # ================= REPORT INFO =================
            info_row = 1
            ws[f"A{info_row}"] = "Generated"
            ws[f"B{info_row}"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            ws[f"D{info_row}"] = "Team"
            ws[f"E{info_row}"] = self._leader_team_label()

            # ================= TABLE HEADER =================
            header_row = 6
            headers = ["Emp ID", "Employee", "Work Days", "Working Hours", "Leave Days", "Late Count", "OT Hours"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # ================= DATA ROWS =================
            thin = Side(border_style="thin", color="D1D5DB")
            data_start = header_row + 1
            for idx, row in enumerate(self.current_rows, start=data_start):
                detail_rows = self._get_employee_attendance_rows(row["user_id"], from_date, to_date)
                row_workdays_with_ot = sum(int(r.get("workdays", 0) or 0) for r in detail_rows)
                row_working_hours = sum(self._actual_hours_from_row(r) for r in detail_rows)
                row_ot_hours = sum(self._ot_hours_from_row(r) for r in detail_rows)
                values = [
                    self._dash(row["employee_id"]),
                    self._dash(row["full_name"]),
                    f"{row_workdays_with_ot:g}",
                    f"{row_working_hours:g}",
                    self._num_or_dash(row["leaveday"]),
                    self._dash(row["latecount"]),
                    self._num_or_dash(row_ot_hours),
                ]
                for col_num, value in enumerate(values, 1):
                    cell = ws.cell(row=idx, column=col_num)
                    cell.value = value
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

            # ================= COLUMN WIDTH =================
            widths = {1: 15, 2: 28, 3: 15, 4: 15, 5: 15, 6: 15, 7: 15}
            for col_num, width in widths.items():
                ws.column_dimensions[get_column_letter(col_num)].width = width

            # ================= ROW HEIGHT =================
            ws.row_dimensions[1].height = 30

            wb.save(file_path)
            messagebox.showinfo("Export Successful", "Attendance Excel exported successfully.")
        except Exception as e:
            messagebox.showerror("Export Error", f"Excel export failed: {e}")




