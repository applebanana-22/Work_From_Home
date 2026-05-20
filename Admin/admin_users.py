import re
import customtkinter as ctk
from database import Database
from tkinter import filedialog
from tkinter import messagebox
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import letter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.units import inch

class UserRegisterFrame(ctk.CTkFrame):
    """The Registration Form UI - Handles dynamic visibility and auto-increment ID"""
    def __init__(self, parent, db, back_callback):
        super().__init__(parent, fg_color="transparent")
        self.db = db
        self.back_callback = back_callback
        
        # --- Top Navigation ---
        self.nav_bar = ctk.CTkFrame(self, fg_color="transparent")
        self.nav_bar.pack(fill="x", padx=20, pady=10)
        
        self.back_btn = ctk.CTkButton(
            self.nav_bar, text="← Back to Users", width=100,
            fg_color="#4A4A4A", hover_color="#333333",
            command=self.back_callback
        )
        self.back_btn.pack(side="left")

        ctk.CTkLabel(self, text="Register New Account", font=("Arial", 22, "bold")).pack(pady=15)

        # Form Container
        self.main_container = ctk.CTkFrame(
            self, 
            fg_color=("#F9F9F9", "#252525"), # <--- Light mode: Light Grey | Dark mode: Dark Grey
            border_width=1, 
            border_color=("#DBDBDB", "#333333"), # <--- Added borders so it stands out in Light Mode
            corner_radius=10
        )
        # Slightly tighter padding so the form fits better on smaller screens
        self.main_container.pack(fill="both", expand=True, padx=60, pady=(8, 20))

        # --- The Grid Form ---
        self.f = ctk.CTkFrame(self.main_container, fg_color="transparent")
        # Match update form padding and spacing
        self.f.pack(expand=True, padx=30, pady=22)
        # Consistent entry width and padding
        entry_w = 340
        entry_padx = 14
        entry_pady = 12
        self._entry_width = entry_w
        self._entry_padx = entry_padx
        self._entry_pady = entry_pady
        
        self.f.grid_columnconfigure(0, weight=0, minsize=120)
        self.f.grid_columnconfigure(1, weight=1, minsize=self._entry_width)
        self.f.grid_columnconfigure(2, weight=0, minsize=120)
        self.f.grid_columnconfigure(3, weight=1, minsize=self._entry_width)

        # 1. Employee ID (Auto-generated & Disabled)
        ctk.CTkLabel(self.f, text="Employee ID:", font=("Arial", 13, "bold")).grid(row=0, column=0, padx=15, pady=10, sticky="e")
        self.emp_id = ctk.CTkEntry(self.f, height=38, width=self._entry_width, fg_color="#a59d9d")
        self.emp_id.grid(row=0, column=1, padx=self._entry_padx, pady=(10, self._entry_pady), sticky="w")

        # 2. Full Name
        name_label_frame = ctk.CTkFrame(self.f, fg_color="transparent")
        name_label_frame.grid(row=0, column=2, padx=15, pady=10, sticky="e")
        ctk.CTkLabel(name_label_frame, text="Full Name:", font=("Arial", 13, "bold")).pack(side="left")
        ctk.CTkLabel(name_label_frame, text=" *", font=("Arial", 20), text_color="#D32F2F").pack(side="left")
        self.name = ctk.CTkEntry(self.f, height=36, width=self._entry_width)
        self.name.grid(row=0, column=3, padx=self._entry_padx, pady=(10, self._entry_pady), sticky="w")
        self.name_err = ctk.CTkLabel(self.f, text="", text_color="#D32F2F", font=("Arial", 10))
        self.name_err.grid(row=1, column=3, padx=self._entry_padx, pady=(0, 10), sticky="w")

        next_id = self.get_next_emp_id()
        self.emp_id.insert(0, str(next_id))
        self.emp_id.configure(state="disabled")

        # 3. Username
        username_label_frame = ctk.CTkFrame(self.f, fg_color="transparent")
        username_label_frame.grid(row=2, column=0, padx=15, pady=10, sticky="e")
        ctk.CTkLabel(username_label_frame, text="Username:", font=("Arial", 13, "bold")).pack(side="left")
        ctk.CTkLabel(username_label_frame, text=" *", font=("Arial", 20, "bold"), text_color="#D32F2F").pack(side="left")
        self.uname = ctk.CTkEntry(self.f, height=36, width=self._entry_width)
        self.uname.grid(row=2, column=1, padx=self._entry_padx, pady=self._entry_pady, sticky="w")
        self.uname_err = ctk.CTkLabel(self.f, text="", text_color="#D32F2F", font=("Arial", 10))
        self.uname_err.grid(row=3, column=1, padx=self._entry_padx, pady=(0, 10), sticky="w")

        # 4. Password
        password_label_frame = ctk.CTkFrame(self.f, fg_color="transparent")
        password_label_frame.grid(row=2, column=2, padx=15, pady=10, sticky="e")
        ctk.CTkLabel(password_label_frame, text="Password:", font=("Arial", 13, "bold")).pack(side="left")
        ctk.CTkLabel(password_label_frame, text=" *", font=("Arial", 20, "bold"), text_color="#D32F2F").pack(side="left")
        self.pwd = ctk.CTkEntry(self.f, height=36, width=self._entry_width, show="*")
        self.pwd.grid(row=2, column=3, padx=self._entry_padx, pady=self._entry_pady, sticky="w")
        self.pwd_err = ctk.CTkLabel(self.f, text="", text_color="#D32F2F", font=("Arial", 10))
        self.pwd_err.grid(row=3, column=3, padx=self._entry_padx, pady=(0, 10), sticky="w")

        # 5. Role Selection
        role_label_frame = ctk.CTkFrame(self.f, fg_color="transparent")
        role_label_frame.grid(row=4, column=0, padx=15, pady=10, sticky="e")
        ctk.CTkLabel(role_label_frame, text="User Role:", font=("Arial", 13, "bold")).pack(side="left")
        ctk.CTkLabel(role_label_frame, text=" *", font=("Arial", 20, "bold"), text_color="#D32F2F").pack(side="left")
        self.role = ctk.CTkOptionMenu(
            self.f, 
            values=["Choose Role", "admin", "leader", "member"],
            height=36, width=self._entry_width,
            command=self.update_ui_by_role
        )
        self.role.grid(row=4, column=1, padx=self._entry_padx, pady=self._entry_pady, sticky="w")
        self.role.set("Choose Role")
        self.role_err = ctk.CTkLabel(self.f, text="", text_color="#D32F2F", font=("Arial", 10))
        self.role_err.grid(row=5, column=1, padx=self._entry_padx, pady=(0, 10), sticky="w")

        # 6. Team Selection (Hidden by default until role is chosen)
        team_label_frame = ctk.CTkFrame(self.f, fg_color="transparent")
        self.team_label = team_label_frame
        ctk.CTkLabel(team_label_frame, text="Assign Team:", font=("Arial", 13, "bold")).pack(side="left")
        ctk.CTkLabel(team_label_frame, text=" *", font=("Arial", 20, "bold"), text_color="#D32F2F").pack(side="left")
        
        teams_data = self.db.get_all_teams()
        self.team_map = {t['team_name']: t['team_id'] for t in teams_data}
        team_options = ["Choose Team"] + list(self.team_map.keys()) if self.team_map else ["Choose Team", "No Teams Found"]
        self.team_dropdown = ctk.CTkOptionMenu(self.f, values=team_options, height=36, width=self._entry_width, command=lambda v: self.clear_error())
        self.team_dropdown.set("Choose Team")
        self.team_err = ctk.CTkLabel(self.f, text="", text_color="#D32F2F", font=("Arial", 10))
        # Note: team_err will be gridded when team_dropdown is shown

        # 7. Batch Selection (Hidden by default)
        batch_label_frame = ctk.CTkFrame(self.f, fg_color="transparent")
        self.batch_label = batch_label_frame
        ctk.CTkLabel(batch_label_frame, text="Batch Number:", font=("Arial", 13, "bold")).pack(side="left")
        ctk.CTkLabel(batch_label_frame, text=" *", font=("Arial", 20, "bold"), text_color="#D32F2F").pack(side="left")
        self.batch = ctk.CTkEntry(self.f, height=36, width=self._entry_width)
        self.batch_err = ctk.CTkLabel(self.f, text="", text_color="#D32F2F", font=("Arial", 10))

        # 8. Save Button
        # Inline global error label (hidden when empty)
        self.error_label = ctk.CTkLabel(self.f, text="", text_color="#D32F2F", font=("Arial", 12, "bold"))
        # place the global error aligned with inputs
        self.error_label.grid(row=8, column=0, columnspan=4, padx=self._entry_padx, pady=(0, 10), sticky="w")

        # Clear error when user types in any core field
        # Clear errors when typing in entries
        for widget in (self.name, self.uname, self.pwd, self.batch):
            try:
                widget.bind("<KeyRelease>", lambda e: self.clear_error())
            except Exception:
                pass

        # Place Register button centered below the form and aligned with input padding
        self.save_btn = ctk.CTkButton(self.f, text="Register", width=100, height=36, fg_color="#10B981",
        hover_color="#0E9A6B", command=self.save_user)
        self.save_btn.grid(row=9, column=0, columnspan=4, padx=self._entry_padx, pady=16)

    def get_next_emp_id(self):
        try:
            self.db.cursor.execute("SELECT MAX(CAST(employee_id AS UNSIGNED)) as max_id FROM users")
            result = self.db.cursor.fetchone()
            return (int(result['max_id']) + 1) if result and result['max_id'] else 1001
        except:
            return 1001

    def update_ui_by_role(self, choice):
        """Hides/Shows fields based on Role"""
        # clear any previous field errors when role changes
        try:
            self.clear_error()
        except Exception:
            pass
        # Hide everything dynamic first
        self.batch_label.grid_forget()
        self.batch.grid_forget()
        self.team_label.grid_forget()
        self.team_dropdown.grid_forget()
        try:
            self.team_err.grid_forget()
        except Exception:
            pass
        try:
            self.batch_err.grid_forget()
        except Exception:
            pass

        if choice == "leader":
            self.team_label.grid(row=4, column=2, padx=15, pady=10, sticky="e")
            self.team_dropdown.grid(row=4, column=3, padx=self._entry_padx, pady=self._entry_pady, sticky="w")
            self.team_err.grid(row=5, column=3, padx=self._entry_padx, pady=(0, 10), sticky="w")
        elif choice == "member":
            self.team_label.grid(row=4, column=2, padx=15, pady=10, sticky="e")
            self.team_dropdown.grid(row=4, column=3, padx=self._entry_padx, pady=self._entry_pady, sticky="w")
            self.team_err.grid(row=5, column=3, padx=self._entry_padx, pady=(0, 10), sticky="w")

            self.batch_label.grid(row=6, column=0, padx=15, pady=10, sticky="e")
            self.batch.grid(row=6, column=1, padx=self._entry_padx, pady=self._entry_pady, sticky="w")
            self.batch_err.grid(row=7, column=1, padx=self._entry_padx, pady=(0, 10), sticky="w")
        # 'admin' remains with them hidden

    def validate_inputs(self):
        full_name = self.name.get().strip()
        username = self.uname.get().strip()
        password = self.pwd.get().strip()
        role_val = self.role.get()
        batch_val = self.batch.get().strip()

        if not all([full_name, username, password]): 
            return False, "All core fields are required!"
        if full_name.isdigit():
            return False, "Full Name cannot be numeric."
        if len(full_name) < 2:
            return False, "Full Name is too short."
        if username.isdigit():
            return False, "User Name cannot be numeric."
        if " " in username:
            return False, "Username cannot contain spaces."
        if len(username) < 4:
            return False, "Username must be at least 4 characters."
        if len(password) < 6:
            return False, "Password must be at least 6 characters long."
        if not re.search(r"[A-Z]", password) or not re.search(r"[a-z]", password) or not re.search(r"\d", password):
            return False, "Password must contain uppercase, lowercase, and a number."
        if len(password) > 20:
            return False, "Password is too long."
        if role_val == "Choose Role":
            return False, "Please select a valid User Role."
        if role_val != "admin":
            if self.team_dropdown.get() == "Choose Team":
                return False, "Please select a valid Team."
       
        batch_val = self.batch.get().strip()
 
        # 1. Basic empty check for core fields
        if not all([full_name, username, password]):          
            return False, "All fields are required!"
 
        # ... (other validations)
 
        # 2. Batch format validation
        if batch_val and not batch_val.isalnum():
            return False, "Batch can only contain letters and numbers."

        # 3. Required check for members
        if role_val == "member" and not batch_val:
            return False, "Batch Number is required for members."
        return True, "Success"

    def save_user(self):
        is_valid, msg = self.validate_inputs()

        # reset any previous error visuals
        self.clear_error()

        default_border = ["#979DA2", "#565B5E"]

        if not is_valid:
            # Handle generic required-fields case by showing per-field inline errors
            low = msg.lower()
            full_name = self.name.get().strip()
            username = self.uname.get().strip()
            password = self.pwd.get().strip()
            role_val = self.role.get()

            if "all fields" in low or "core fields" in low or "required" in low:
                if not full_name:
                    self.name_err.configure(text="Full Name is required.")
                    try: self.name.configure(border_color="red")
                    except Exception: pass
                if not username:
                    self.uname_err.configure(text="Username is required.")
                    try: self.uname.configure(border_color="red")
                    except Exception: pass
                if not password:
                    self.pwd_err.configure(text="Password is required.")
                    try: self.pwd.configure(border_color="red")
                    except Exception: pass
                if role_val == "Choose Role":
                    self.role_err.configure(text="Please select a valid User Role.")
                    try: self.role.configure(border_color="red")
                    except Exception: pass
                if role_val != "admin" and self.team_dropdown.get() == "Choose Team":
                    self.team_err.configure(text="Please select a valid Team.")
                    try: self.team_dropdown.configure(border_color="red")
                    except Exception: pass
                if role_val == "member" and not self.batch.get().strip():
                    self.batch_err.configure(text="Batch Number is required for members.")
                    try: self.batch.configure(border_color="red")
                    except Exception: pass
                return

            # Specific validation messages -> attach to appropriate field
            if "full name" in low:
                self.name_err.configure(text=msg)
                try: self.name.configure(border_color="red")
                except Exception: pass
                return
            if "user name" in low or "username" in low:
                self.uname_err.configure(text=msg)
                try: self.uname.configure(border_color="red")
                except Exception: pass
                return
            if "password" in low:
                self.pwd_err.configure(text=msg)
                try: self.pwd.configure(border_color="red")
                except Exception: pass
                return
            if "team" in low:
                self.team_err.configure(text=msg)
                try: self.team_dropdown.configure(border_color="red")
                except Exception: pass
                return
            if "batch" in low:
                self.batch_err.configure(text=msg)
                try: self.batch.configure(border_color="red")
                except Exception: pass
                return

            # fallback to global message
            self.error_label.configure(text=msg)
            return

        role_val = self.role.get()
        # Admin gets no team, others get selected team
        team_id = self.team_map.get(self.team_dropdown.get()) if role_val != "admin" else None
        
        # Batch formatting: numeric values get Batch prefix, non-numeric values stored raw
        batch_input = self.batch.get().strip()
        if role_val == "member" and batch_input:
            formatted_batch = f"Batch {batch_input}" if batch_input.isdigit() else batch_input
        else:
            formatted_batch = "N/A"

        try:
            sql = """INSERT INTO users (employee_id, full_name, username, password, role, team_id, batch) 
                     VALUES (%s, %s, %s, %s, %s, %s, %s)"""
            data = (
                self.emp_id.get(), 
                self.name.get().strip(), 
                self.uname.get().strip(), 
                self.pwd.get(), 
                role_val, 
                team_id, 
                formatted_batch
            )
            self.db.cursor.execute(sql, data)
            self.db.conn.commit()
            messagebox.showinfo("Success", f"User {self.name.get()} created successfully!")
            self.back_callback()
        except Exception as e:
            messagebox.showerror("Database Error", str(e))

    def clear_error(self):
        default_border = ["#979DA2", "#565B5E"]
        try:
            self.error_label.configure(text="")
        except Exception:
            pass
        try:
            self.name_err.configure(text="")
        except Exception:
            pass
        try:
            self.uname_err.configure(text="")
        except Exception:
            pass
        try:
            self.pwd_err.configure(text="")
        except Exception:
            pass
        try:
            self.team_err.configure(text="")
        except Exception:
            pass
        try:
            self.batch_err.configure(text="")
        except Exception:
            pass
        try:
            self.role_err.configure(text="")
        except Exception:
            pass

        # Reset borders
        try: self.name.configure(border_color=default_border)
        except Exception: pass
        try: self.uname.configure(border_color=default_border)
        except Exception: pass
        try: self.pwd.configure(border_color=default_border)
        except Exception: pass
        try: self.batch.configure(border_color=default_border)
        except Exception: pass
        try: self.team_dropdown.configure(border_color=default_border)
        except Exception: pass
        try: self.role.configure(border_color=default_border)
        except Exception: pass
            
            
class UserUpdateFrame(ctk.CTkFrame):
    """The Update Form UI - Pre-fills data and updates existing records"""
    def __init__(self, parent, db, user_id, back_callback):
        super().__init__(parent, fg_color="transparent")
        self.db = db
        self.user_id = user_id
        self.back_callback = back_callback
        
        # --- UI Setup ---
        self.nav_bar = ctk.CTkFrame(self, fg_color="transparent")
        self.nav_bar.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkButton(self.nav_bar, text="← Back", width=100,
                      fg_color="#4A4A4A", command=self.back_callback).pack(side="left")

        ctk.CTkLabel(self, text="Update User Account",
                     font=("Arial", 22, "bold")).pack(pady=15)

        self.main_container = ctk.CTkFrame(
            self, 
            fg_color=("#F9F9F9", "#252525"),
            border_width=1, 
            border_color=("#DBDBDB", "#333333"),
            corner_radius=10
        )
        self.main_container.pack(fill="both", expand=True, padx=70, pady=(10, 20))

        self.f = ctk.CTkFrame(self.main_container, fg_color="transparent")
        self.f.pack(expand=True, padx=30, pady=22)

        self._entry_width = 340
        self._entry_padx = 14
        self._entry_pady = 12

        self.f.grid_columnconfigure(0, weight=0, minsize=120)
        self.f.grid_columnconfigure(1, weight=1, minsize=self._entry_width)
        self.f.grid_columnconfigure(2, weight=0, minsize=120)
        self.f.grid_columnconfigure(3, weight=1, minsize=self._entry_width)

        ctk.CTkLabel(self.f, text="Full Name:", font=("Arial", 13, "bold")).grid(row=0, column=0, padx=15, pady=10, sticky="e")
        self.name = ctk.CTkEntry(self.f, height=38, width=self._entry_width)
        self.name.grid(row=0, column=1, padx=self._entry_padx, pady=(10, self._entry_pady), sticky="w")
        self.name_err = ctk.CTkLabel(self.f, text="", text_color="#D32F2F", font=("Arial", 10))
        self.name_err.grid(row=1, column=1, padx=self._entry_padx, pady=(0, 10), sticky="w")

        ctk.CTkLabel(self.f, text="Username:", font=("Arial", 13, "bold")).grid(row=0, column=2, padx=15, pady=10, sticky="e")
        self.uname = ctk.CTkEntry(self.f, height=38, width=self._entry_width)
        self.uname.grid(row=0, column=3, padx=self._entry_padx, pady=(10, self._entry_pady), sticky="w")
        self.uname_err = ctk.CTkLabel(self.f, text="", text_color="#D32F2F", font=("Arial", 10))
        self.uname_err.grid(row=1, column=3, padx=self._entry_padx, pady=(0, 10), sticky="w")

        ctk.CTkLabel(self.f, text="User Role:", font=("Arial", 13, "bold")).grid(row=2, column=0, padx=15, pady=10, sticky="e")
        self.role = ctk.CTkOptionMenu(
            self.f,
            values=["admin", "leader", "member"],
            height=38,
            width=self._entry_width,
            command=self.update_ui_by_role
        )
        self.role.grid(row=2, column=1, padx=self._entry_padx, pady=self._entry_pady, sticky="w")
        self.role.set("admin")
        self.role_err = ctk.CTkLabel(self.f, text="", text_color="#D32F2F", font=("Arial", 10))
        self.role_err.grid(row=3, column=1, padx=self._entry_padx, pady=(0, 10), sticky="w")

        self.team_label = ctk.CTkLabel(self.f, text="Team:", font=("Arial", 13, "bold"))
        self.team_dropdown = ctk.CTkOptionMenu(self.f, values=[], height=38, width=self._entry_width)
        self.team_err = ctk.CTkLabel(self.f, text="", text_color="#D32F2F", font=("Arial", 10))

        self.batch_label = ctk.CTkLabel(self.f, text="Batch:", font=("Arial", 13, "bold"))
        self.batch = ctk.CTkEntry(self.f, height=38, width=self._entry_width)
        self.batch_err = ctk.CTkLabel(self.f, text="", text_color="#D32F2F", font=("Arial", 10))

        self.error_label = ctk.CTkLabel(self.f, text="", text_color="#D32F2F", font=("Arial", 12, "bold"))
        self.error_label.grid(row=6, column=0, columnspan=4, padx=self._entry_padx, pady=(0, 10), sticky="w")

        teams_data = self.db.get_all_teams()
        self.team_map = {t['team_name']: t['team_id'] for t in teams_data}
        team_options = ["Select Team"] + list(self.team_map.keys()) if self.team_map else ["No Teams"]
        self.team_dropdown.configure(values=team_options)
        self.team_dropdown.set("Select Team")

        self.update_btn = ctk.CTkButton(
            self.f, text="Submit",
            width=100,
            height=36,
            fg_color="#10B981", hover_color="#059669",
            command=self.perform_update
        )
        self.update_btn.grid(row=7, column=0, columnspan=4, padx=self._entry_padx, pady=24)

        self.load_user_data()

    def update_ui_by_role(self, role):
        """Show/hide fields dynamically"""
        self.team_label.grid_forget()
        self.team_dropdown.grid_forget()
        self.team_err.grid_forget()
        self.batch_label.grid_forget()
        self.batch.grid_forget()
        self.batch_err.grid_forget()

        if role == "leader":
            self.team_label.grid(row=2, column=2, padx=15, pady=10, sticky="e")
            self.team_dropdown.grid(row=2, column=3, padx=self._entry_padx, pady=self._entry_pady, sticky="w")
            self.team_err.grid(row=3, column=3, padx=self._entry_padx, pady=(0, 10), sticky="w")
        elif role == "member":
            self.team_label.grid(row=2, column=2, padx=15, pady=10, sticky="e")
            self.team_dropdown.grid(row=2, column=3, padx=self._entry_padx, pady=self._entry_pady, sticky="w")
            self.team_err.grid(row=3, column=3, padx=self._entry_padx, pady=(0, 10), sticky="w")
            self.batch_label.grid(row=4, column=0, padx=15, pady=10, sticky="e")
            self.batch.grid(row=4, column=1, padx=self._entry_padx, pady=self._entry_pady, sticky="w")
            self.batch_err.grid(row=5, column=1, padx=self._entry_padx, pady=(0, 10), sticky="w")

    def clear_error(self):
        default_border = ["#979DA2", "#565B5E"]
        try:
            self.error_label.configure(text="")
        except Exception:
            pass
        try:
            self.name_err.configure(text="")
        except Exception:
            pass
        try:
            self.uname_err.configure(text="")
        except Exception:
            pass
        try:
            self.role_err.configure(text="")
        except Exception:
            pass
        try:
            self.team_err.configure(text="")
        except Exception:
            pass
        try:
            self.batch_err.configure(text="")
        except Exception:
            pass

        try: self.name.configure(border_color=default_border)
        except Exception: pass
        try: self.uname.configure(border_color=default_border)
        except Exception: pass
        try: self.batch.configure(border_color=default_border)
        except Exception: pass
        try: self.team_dropdown.configure(border_color=default_border)
        except Exception: pass

    def load_user_data(self):
        self.db.cursor.execute("SELECT * FROM users WHERE id=%s", (self.user_id,))
        user = self.db.cursor.fetchone()

        if user:
            self.name.insert(0, user['full_name'])
            self.uname.insert(0, user['username'])
            self.role.set(user['role'])

            self.update_ui_by_role(user['role'])

            if user['team_id']:
                for name, tid in self.team_map.items():
                    if tid == user['team_id']:
                        self.team_dropdown.set(name)

            if user['batch'] and user['batch'] != "N/A":
                self.batch.insert(0, user['batch'].replace("Batch ", ""))

    def perform_update(self):
        try:
            role = self.role.get()
            team_name = self.team_dropdown.get()
            batch_input = self.batch.get().strip()
            full_name = self.name.get().strip()
            username = self.uname.get().strip()

            self.clear_error()

            if not full_name:
                self.name_err.configure(text="Full Name is required.")
                try: self.name.configure(border_color="red")
                except Exception: pass
                return
            if full_name.isdigit():
                self.name_err.configure(text="Full Name cannot be numeric.")
                try: self.name.configure(border_color="red" ,)
                except Exception: pass
                return
            if len(full_name) < 2:
                self.name_err.configure(text="Full Name is too short.")
                try: self.name.configure(border_color="red")
                except Exception: pass
                return

            if not username:
                self.uname_err.configure(text="Username is required.")
                try: self.uname.configure(border_color="red")
                except Exception: pass
                return
            if username.isdigit():
                self.uname_err.configure(text="Username cannot be numeric.")
                try: self.uname.configure(border_color="red")
                except Exception: pass
                return
            if " " in username:
                self.uname_err.configure(text="Username cannot contain spaces.")
                try: self.uname.configure(border_color="red")
                except Exception: pass
                return
            if len(username) < 4:
                self.uname_err.configure(text="Username must be at least 4 characters.")
                try: self.uname.configure(border_color="red")
                except Exception: pass
                return

            if role not in ["admin", "leader", "member"]:
                self.role_err.configure(text="Please select a valid User Role.")
                try: self.role.configure(border_color="red")
                except Exception: pass
                return

            if role != "admin":
                if team_name == "Select Team" or team_name not in self.team_map:
                    self.team_err.configure(text="Please select a valid Team.")
                    try: self.team_dropdown.configure(border_color="red")
                    except Exception: pass
                    return

            if role == "member":
                if not batch_input:
                    self.batch_err.configure(text="Batch Number is required for members.")
                    try: self.batch.configure(border_color="red")
                    except Exception: pass
                    return
            if batch_input and not batch_input.isalnum():
                self.batch_err.configure(text="Batch can only contain letters and numbers.")
                try: self.batch.configure(border_color="red")
                except Exception: pass
                return

            team_id = self.team_map.get(team_name) if role != "admin" else None
            if role == "member" and batch_input:
                batch_val = f"Batch {batch_input}" if batch_input.isdigit() else batch_input
            else:
                batch_val = "N/A"
            sql = """
            UPDATE users 
            SET full_name=%s, username=%s, role=%s, team_id=%s, batch=%s 
            WHERE id=%s
            """

            val = (
                full_name,
                username,
                role,
                team_id,
                batch_val,
                self.user_id
            )

            self.db.cursor.execute(sql, val)
            self.db.conn.commit()

            messagebox.showinfo("Success", "User updated successfully!")
            self.back_callback()

        except Exception as e:
            messagebox.showerror("Error", str(e))
            
            
class AdminUsers(ctk.CTkFrame):
    """Main User Management View"""
    def __init__(self, master, user_data):
        super().__init__(master, fg_color="transparent")
        self.db = Database()
        self.user = user_data
        self.container = ctk.CTkFrame(self, fg_color="transparent")
        self.container.pack(fill="both", expand=True)
        self.show_list_view()

    def clear_container(self):
        for widget in self.container.winfo_children():
            widget.destroy()

    def show_list_view(self):
        self.clear_container()
        header = ctk.CTkFrame(self.container, fg_color="transparent")
        header.pack(fill="x", padx=80, pady=20)
        
        ctk.CTkLabel(header, text="User Management", font=("Arial", 24, "bold")).pack(side="left")
        
        # --- FILTER BAR ---
        filter_frame = ctk.CTkFrame(self.container, fg_color="transparent")
        filter_frame.pack(fill="x", padx=80, pady=(0, 10))

        self.search_id = ctk.CTkEntry(filter_frame, placeholder_text="Employee ID", width=100, height=36)
        self.search_id.pack(side="left", padx=5)

        self.search_name = ctk.CTkEntry(filter_frame, placeholder_text="Full Name", width=100, height=36)
        self.search_name.pack(side="left", padx=5)

        self.search_role = ctk.CTkOptionMenu(
            filter_frame,
            values=["Roles", "admin", "leader", "member"],
            width=120
        )
        self.search_role.set("Roles")
        self.search_role.pack(side="left", padx=5)

        teams = ["Teams"] + self.get_team_names()
        self.search_team = ctk.CTkOptionMenu(filter_frame, values=teams, width=120)
        self.search_team.set("Teams")
        self.search_team.pack(side="left", padx=5)

        self.search_batch = ctk.CTkEntry(filter_frame, placeholder_text="Batch", width=100)
        self.search_batch.pack(side="left", padx=5)

        ctk.CTkButton(filter_frame, text="🔍 Filter",fg_color="#2471A3", hover_color="#1A5276", width=90, command=self.filter_users).pack(side="left", padx=5)
        ctk.CTkButton(filter_frame, text="✖ Clear", fg_color="#566573", hover_color="#424949", width=90, command=self.reset_filters).pack(side="left", padx=5)
        ctk.CTkButton(filter_frame, text="📄PDF", fg_color="#AA4242", hover_color="#B62525", width=90, command=self.export_users).pack(side="left", padx=5)
        ctk.CTkButton(filter_frame, text="📥CSV", fg_color="#16A085", hover_color="#107863", width=90, command=self.csvexport_user).pack(side="left", padx=5)
        
        
        ctk.CTkButton(
            header, text="+ Create Account", 
            fg_color="#10B981",
            hover_color="#0E9A6B", height=38, 
            command=self.show_register_view
        ).pack(side="right")
        
        self.list_frame = ctk.CTkScrollableFrame(
            self.container,
            label_text="Our Members",
            label_font=("Arial", 18, "bold"),
            fg_color=("#FFFFFF", "#181818"),
            border_width=1,
            border_color=("#E0E0E0", "#2B2B2B"),
            corner_radius=10
        )

        self.list_frame.pack(
            fill="both",
            expand=True,
            padx=80,
            pady=10
        )
        self.refresh_list()

    def show_register_view(self):
        self.clear_container()
        reg_form = UserRegisterFrame(self.container, self.db, self.show_list_view)
        reg_form.pack(fill="both", expand=True)
        
    def get_team_names(self):
        self.db.cursor.execute("SELECT team_name FROM teams")
        return [t['team_name'] for t in self.db.cursor.fetchall()]

    def refresh_list(self):
        for w in self.list_frame.winfo_children():
            w.destroy()

        sql = """
        SELECT u.*, t.team_name
        FROM users u
        LEFT JOIN teams t ON u.team_id = t.team_id
        ORDER BY u.id DESC
        """
        self.db.cursor.execute(sql)
        rows = self.db.cursor.fetchall()

        self.render_table(rows)
    def filter_users(self):
        
        if self.search_id.get() and not self.search_id.get().isdigit():
            messagebox.showwarning("Invalid Input", "Employee ID filter must contain only numbers.")
            self.search_id.focus()
            return
        
        for w in self.list_frame.winfo_children():
            w.destroy()

        conditions = []
        params = []

        if self.search_id.get():
            conditions.append("u.employee_id LIKE %s")
            params.append(f"%{self.search_id.get()}%")

        if self.search_name.get():
            if self.search_name.get().isdigit():
                messagebox.showwarning("Invalid Input", "Full Name filter cannot be numeric.")
                self.search_name.focus()
                return
            else:
                conditions.append("u.full_name LIKE %s")
                params.append(f"%{self.search_name.get()}%")

        if self.search_role.get() != "Roles":
            conditions.append("u.role = %s")
            params.append(self.search_role.get())

        if self.search_team.get() != "Teams":
            conditions.append("t.team_name = %s")
            params.append(self.search_team.get())

        if self.search_batch.get():
            if not self.search_batch.get().isdigit():
                messagebox.showwarning("Invalid Input", "Batch filter must contain only numbers.")
                self.search_batch.focus()
                return
            else:
                conditions.append("u.batch LIKE %s")
                params.append(f"%{self.search_batch.get()}%")

        where_clause = " AND ".join(conditions)
        if where_clause:
            where_clause = "WHERE " + where_clause

        sql = f"""
        SELECT u.*, t.team_name 
        FROM users u
        LEFT JOIN teams t ON u.team_id = t.team_id
        {where_clause}
        ORDER BY u.id DESC
        """

        self.db.cursor.execute(sql, tuple(params))
        rows = self.db.cursor.fetchall()

        # ✅ MUST BE INSIDE FUNCTION (same indentation level)
        if not rows:
            ctk.CTkLabel(
                self.list_frame,
                text="This member is not exist.",
                font=("Arial", 16, "bold"),
                text_color="red"
            ).pack(pady=50)
            return  # ✅ NOW CORRECT

        self.render_table(rows)
    def reset_filters(self):
        # Clear text fields
        self.search_id.delete(0, "end" )
        self.search_name.delete(0, "end" )
        self.search_batch.delete(0, "end")

        # Reset dropdowns
        self.search_role.set("Roles")
        self.search_team.set("Teams")

        self.search_id.focus()
        self.search_name.focus()
        self.search_batch.focus()
        self.focus()  # Set focus back to the main frame

        # Reload full list
        self.refresh_list()
        
    def csvexport_user(self):
        try:
            conditions = []
            params = []

            # Employee ID Filter
            if self.search_id.get():
                conditions.append("u.employee_id LIKE %s")
                params.append(f"%{self.search_id.get()}%")

            # Full Name Filter
            if self.search_name.get():
                conditions.append("u.full_name LIKE %s")
                params.append(f"%{self.search_name.get()}%")

            # Role Filter
            if self.search_role.get() != "Roles":
                conditions.append("u.role = %s")
                params.append(self.search_role.get())

            # Team Filter
            if self.search_team.get() != "Teams":
                conditions.append("t.team_name = %s")
                params.append(self.search_team.get())

            # Batch Filter
            if self.search_batch.get():
                conditions.append("u.batch LIKE %s")
                params.append(f"%{self.search_batch.get()}%")

            # WHERE Clause
            where_clause = ""
            if conditions:
                where_clause = "WHERE " + " AND ".join(conditions)

            # SQL Query
            sql = f"""
            SELECT
                u.employee_id,
                u.full_name,
                u.username,
                u.role,
                IFNULL(t.team_name, 'No Team') AS team_name,
                u.batch
            FROM users u
            LEFT JOIN teams t ON u.team_id = t.team_id
            {where_clause}
            ORDER BY u.id DESC
            """

            # Execute Query
            self.db.cursor.execute(sql, tuple(params))
            rows = self.db.cursor.fetchall()

            # No Data
            if not rows:
                messagebox.showwarning(
                    "No Data",
                    "No matching users found to export."
                )
                return

            # Current Date
            current_date = datetime.now().strftime("%Y-%m-%d")

            # Save File Dialog
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile=f"GIC_allmembers_list_{current_date}.xlsx",
                title="Save Excel Report"
            )

            if not file_path:
                return

            # Create Workbook
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Users"

            # Header Columns
            headers = [
                "Employee ID",
                "Full Name",
                "Username",
                "Role",
                "Team",
                "Batch"
            ]

            # Add Header Row
            sheet.append(headers)

            # Add Data Rows
            for row in rows:
                sheet.append([
                    row['employee_id'],
                    row['full_name'],
                    row['username'],
                    row['role'],
                    row['team_name'],
                    row['batch']
                ])

            # =========================
            # Styles
            # =========================

            # Header Background Color
            header_fill = PatternFill(
            start_color="D3D3D3",
            end_color="D3D3D3",
            fill_type="solid"
            )

            # Header Font
            header_font = Font(
                bold=True,
                color="000000",
                size=12
            )

            # Body Font
            body_font = Font(
                size=11
            )

            # Alignment
            center_alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            # Border Style
            thin_border = Border(
                left=Side(style="thin", color="808080"),
                right=Side(style="thin", color="808080"),
                top=Side(style="thin", color="808080"),
                bottom=Side(style="thin", color="808080")
            )

            # =========================
            # Header Style
            # =========================
            for col_num in range(1, len(headers) + 1):

                cell = sheet.cell(row=1, column=col_num)

                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border

            # =========================
            # Body Style
            # =========================
            for row in sheet.iter_rows(
                min_row=2,
                max_row=sheet.max_row,
                min_col=1,
                max_col=len(headers)
            ):

                for cell in row:

                    cell.font = body_font
                    cell.alignment = center_alignment
                    cell.border = thin_border

            # =========================
            # Column Widths
            # =========================
            widths = {
                "A": 18,
                "B": 28,
                "C": 22,
                "D": 15,
                "E": 20,
                "F": 18
            }

            for col, width in widths.items():
                sheet.column_dimensions[col].width = width

            # =========================
            # Row Heights
            # =========================
            for row in range(1, sheet.max_row + 1):
                sheet.row_dimensions[row].height = 28

            # Freeze Header Row
            # sheet.freeze_panes = "A2"

            # Save Workbook
            workbook.save(file_path)

            # Success Message
            self.show_success_toast("Excel File exported successfully!")

        except Exception as e:

            messagebox.showerror(
                "Export Error",
                str(e)
            )    
    def show_success_toast(self, message):

        toast = ctk.CTkFrame(
            self,
            fg_color="#22C55E",
            corner_radius=8
        )

        toast.place(
            relx=1.0,
            y=20,
            anchor="ne"
        )

        label = ctk.CTkLabel(
            toast,
            text=message,
            text_color="white",
            font=("Arial", 12, "bold")
        )

        label.pack(
            padx=20,
            pady=10
        )

        # Auto hide after 3 seconds
        self.after(
            3000,
            toast.destroy
        )
        
    def export_users(self):
        try:
            conditions = []
            params = []

            # Employee ID Filter
            if self.search_id.get():
                conditions.append("u.employee_id LIKE %s")
                params.append(f"%{self.search_id.get()}%")

            # Full Name Filter
            if self.search_name.get():
                conditions.append("u.full_name LIKE %s")
                params.append(f"%{self.search_name.get()}%")

            # Role Filter
            if self.search_role.get() != "Roles":
                conditions.append("u.role = %s")
                params.append(self.search_role.get())

            # Team Filter
            if self.search_team.get() != "Teams":
                conditions.append("t.team_name = %s")
                params.append(self.search_team.get())

            # Batch Filter
            if self.search_batch.get():
                conditions.append("u.batch LIKE %s")
                params.append(f"%{self.search_batch.get()}%")

            where_clause = ""
            if conditions:
                where_clause = "WHERE " + " AND ".join(conditions)

            sql = f"""
            SELECT
                u.employee_id,
                u.full_name,
                u.username,
                u.role,
                IFNULL(t.team_name, 'No Team') AS team_name,
                u.batch
            FROM users u
            LEFT JOIN teams t ON u.team_id = t.team_id
            {where_clause}
            ORDER BY u.id DESC
            """

            self.db.cursor.execute(sql, tuple(params))
            rows = self.db.cursor.fetchall()

            if not rows:
                self.show_success_toast("No matching users found to export.")
                return

            # Save PDF
            current_date = datetime.now().strftime("%Y-%m-%d")

            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF Files", "*.pdf")],
                initialfile=f"GIC_allmember_list_{current_date}.pdf",
                title="Save User Report"
            )

            if not file_path:
                return

            # Create PDF
            doc = SimpleDocTemplate(
                file_path,
                pagesize=letter,
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=20
            )

            elements = []

            styles = getSampleStyleSheet()

            # Center Title Style
            title_style = ParagraphStyle(
                name="TitleStyle",
                parent=styles['Heading1'],
                alignment=TA_CENTER,
                fontSize=24,
                leading=30,
                spaceAfter=20,
                textColor=colors.black
            )

            normal_style = ParagraphStyle(
                name="NormalStyle",
                parent=styles['Normal'],
                fontSize=12,
                leading=18
            )

            # Title
            title = Paragraph(
                "User Management Report",
                title_style
            )

            elements.append(title)

            # Date
            report_date = Paragraph(
                f"Generated Date: {current_date}",
                normal_style
            )

            elements.append(report_date)
            elements.append(Spacer(1, 20))

            # Table Data
            data = [[
                "Employee ID",
                "Full Name",
                "Username",
                "Role",
                "Team",
                "Batch"
            ]]

            for row in rows:
                data.append([
                    row['employee_id'],
                    row['full_name'],
                    row['username'],
                    row['role'],
                    row['team_name'],
                    row['batch']
                ])

            # Column Widths
            col_widths = [
                1.2 * inch,
                1.7 * inch,
                1.4 * inch,
                1.0 * inch,
                1.5 * inch,
                1.0 * inch
            ]

            # Create Table
            table = Table(
                data,
                colWidths=col_widths,
                repeatRows=1
            )

            table.setStyle(TableStyle([

            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),

            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),

            # Body Rows
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),

            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 11),

            # Grid
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),

            # Alignment
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

            # Padding
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),

        ]))

            elements.append(table)

            # Build PDF
            doc.build(elements)

            self.show_success_toast(
                "PDF File exported successfully!"
            )

        except Exception as e:
            messagebox.showerror(
                "Export Error",
                str(e)
            )
        
    def render_table(self, rows):

        # ===== FULL WIDTH CONFIG =====
        columns = 6

        for col in range(columns):
            self.list_frame.grid_columnconfigure(col, weight=1)

        headers = ["Employee ID", "Full Name", "Role", "Team", "Batch", "Actions"]

        # ===== TABLE HEADER =====
        for col, header in enumerate(headers):
            ctk.CTkLabel(
                self.list_frame,
                text=header,
                font=("Arial", 15, "bold"),
                anchor="center"
            ).grid(
                row=0,
                column=col,
                padx=10,
                pady=(10, 15),
                sticky="ew"
            )

        # ===== TABLE ROWS =====
        for i, row in enumerate(rows, start=1):

            values = [
                row['employee_id'],
                row['full_name'],
                row['role'],
                row['team_name'] if row['team_name'] else "No Team",
                row['batch'] if row['batch'] else "N/A"
            ]

            for col, value in enumerate(values):
                ctk.CTkLabel(
                    self.list_frame,
                    text=value,
                    font=("Arial", 13),
                    height=40,
                    anchor="center"
                ).grid(
                    row=i,
                    column=col,
                    padx=10,
                    pady=5,
                    sticky="ew"
                )

            # ===== ACTION BUTTONS =====
            btn_frame = ctk.CTkFrame(
                self.list_frame,
                fg_color="transparent"
            )

            btn_frame.grid(
                row=i,
                column=5,
                padx=10,
                pady=5,
                
            )

            btn_frame.grid_columnconfigure((0, 1), weight=1)

            ctk.CTkButton(
                btn_frame,
                text="Edit",
                width=70,
                height=30,
                fg_color="#F39C12",
                hover_color="#D68910",
                command=lambda uid=row['id']: self.handle_update(uid)
            ).grid(row=0, column=0, padx=5)

            ctk.CTkButton(
                btn_frame,
                text="Delete",
                width=70,
                height=30,
                fg_color="#E74C3C",
                hover_color="#C0392B",
                command=lambda uid=row['id']: self.handle_delete(uid)
            ).grid(row=0, column=1, padx=5)
        
    def handle_update(self, uid):
        self.clear_container()

        update_form = UserUpdateFrame(
            self.container,
            self.db,
            uid,
            self.show_list_view
        )

        update_form.pack(fill="both", expand=True)

    def handle_delete(self, uid):
        if uid == self.user['id']:
            messagebox.showwarning("Denied", "You cannot delete yourself!")
            return
        if messagebox.askyesno("Confirm", "Are you sure you want to delete this user?"):
            self.db.cursor.execute("DELETE FROM users WHERE id=%s", (uid,))
            self.db.conn.commit()
            self.refresh_list()