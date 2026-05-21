import customtkinter as ctk
from database import Database
from tkinter import messagebox, filedialog, ttk
from datetime import datetime
import tkinter as tk
from fpdf import FPDF
from tkcalendar import Calendar
from collections import defaultdict
import os

# ── Reusable modern date-picker ──────────────────────────────────────────────
class DatePickerButton(ctk.CTkFrame):
    def __init__(self, master, initial_date=None):
        super().__init__(master, fg_color="transparent")

        self._date = initial_date or datetime.today().date()
        self._open = False

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Custom.Calendar",
            background="#1A1A2E",
            foreground="white",
            headersbackground="#16213E",
            headersforeground="#4FC3F7",
            selectbackground="#3498DB",
            selectforeground="white",
            normalbackground="#1A1A2E",
            normalforeground="#CCCCCC",
            weekendbackground="#1A1A2E",
            weekendforeground="#F39C12",
            othermonthforeground="#555555",
            bordercolor="#2A2A4A",
            relief="flat"
        )

        self.btn = ctk.CTkButton(
            self,
            text=self._fmt(),
            width=140,
            height=36,
            corner_radius=8,
            hover_color=("#EBEBEB", "#2A2A2A"),
            border_width=1,
            fg_color=("#F9F9F9", "#1E1E1E"),
            border_color=("#DBDBDB", "#2C2C2C"),
            text_color=("black", "white"),
            anchor="w",
            command=self.toggle
        )
        self.btn.pack()

        self.panel = ctk.CTkFrame(
            self.winfo_toplevel(),
            fg_color=("#FFFFFF", "#141E2B"),
            corner_radius=12,
            border_width=1,
            border_color=("#CCCCCC", "#2A3A4A")
        )

        self.cal = Calendar(
            self.panel,
            style="Custom.Calendar",
            selectmode="day",
            date_pattern="yyyy-mm-dd",
            year=self._date.year,
            month=self._date.month,
            day=self._date.day,
            maxdate=datetime.today().date(),
            showweeknumbers=False,
            firstweekday="monday",
            font=("Arial", 11),
            cursor="hand2"
        )
        self.cal.pack(padx=8, pady=8)
        self.cal.bind("<<CalendarSelected>>", self._select)

    def toggle(self):
        if self._open:
            self.panel.place_forget()
        else:
            self.panel.lift()
            self.panel.place(in_=self, x=0, y=self.btn.winfo_height() + 2)
        self._open = not self._open

    def _select(self, event):
        selected = self.cal.get_date()
        self._date = datetime.strptime(selected, "%Y-%m-%d").date()
        self.btn.configure(text=self._fmt())
        self.toggle()

    def get_date(self):
        return self._date

    def set_date(self, d):
        self._date = d
        self.cal.selection_set(d)
        self.btn.configure(text=self._fmt())

    def _fmt(self):
        return f"  📅  {self._date.strftime('%Y-%m-%d')}"

class DailyReportFrame(ctk.CTkFrame):
    def __init__(self, master, user, **kwargs):
        super().__init__(master, **kwargs)
        self.db = Database()
        self.user = user
        self.configure(fg_color="transparent")

        today = datetime.today().date()
        self.start_date = today
        self.end_date = today
        self.member_search = ""
        self.selected_team = "All Teams"          # bottom report filter
        self.count_selected_team = "All Teams"   # top count filter
        self.count_filter_date = today

        self.team_options = ["All Teams"]
        self.team_map = {"All Teams": None}
        self.load_team_options()

        self.container = ctk.CTkFrame(self, fg_color="transparent")
        self.container.pack(fill="both", expand=True)

        self.show_reports_list()

    def _show_message(self, message, message_type="info", duration=3000):
        if message_type == "error":
            bg_color = "#E74C3C"
        elif message_type == "warning":
            bg_color = "#F39C12"
        elif message_type == "success":
            bg_color = "#27AE60"
        else:
            bg_color = "#3498DB"

        message_frame = ctk.CTkFrame(
            self.winfo_toplevel(),
            fg_color=bg_color,
            corner_radius=8
        )
        message_frame.place(relx=1.0, rely=0, x=-20, y=20, anchor="ne")

        ctk.CTkLabel(
            message_frame,
            text=message,
            text_color="white",
            font=("Arial", 12, "bold"),
            wraplength=250
        ).pack(padx=15, pady=10)

        self.after(duration, message_frame.destroy)

    def load_team_options(self):
        try:
            teams = self.db.get_all_teams() or []
            for team in teams:
                team_name = team.get('team_name')
                team_id = team.get('team_id')
                if team_name:
                    self.team_options.append(team_name)
                    self.team_map[team_name] = team_id
        except Exception:
            pass

    def clear_container(self):
        for widget in self.container.winfo_children():
            widget.destroy()

    def show_reports_list(self):
        content_padx = 80
        self.clear_container()

        header = ctk.CTkFrame(self.container, fg_color="transparent")
        header.pack(fill="x", padx=content_padx, pady=14)

        ctk.CTkLabel(
            header,
            text="📊 Daily Report",
            font=("Arial", 20, "bold"),
            text_color=("#333333", "#FFFFFF")
        ).pack(side="left")

        counts_bar = ctk.CTkFrame(
            self.container,
            corner_radius=14,
            fg_color=("#F9F9F9", "#1E1E1E"),
            border_width=1,
            border_color=("#E0E0E0", "#2C2C2C")
        )
        counts_bar.pack(fill="x", padx=content_padx, pady=(0, 10))

        left_wrap = ctk.CTkFrame(counts_bar, fg_color="transparent")
        left_wrap.pack(anchor="w", padx=18, pady=14)

        top_row = ctk.CTkFrame(left_wrap, fg_color="transparent")
        top_row.pack(anchor="w", pady=(0, 6))

        self.count_date = DatePickerButton(
            top_row,
            initial_date=getattr(
                self,
                "count_filter_date",
                datetime.today().date()
            )
        )

        self.count_date.pack(side="left", padx=(0, 8))

        # keep first date picker selectable
        self.count_date.btn.configure(
            command=self.count_date.toggle
        )

        # Team selector next to the counts date
        self.count_team_menu = ctk.CTkOptionMenu(
            top_row,
            values=self.team_options,
            command=self.set_count_team_filter,
            width=120,
            height=28,
            corner_radius=8,
            fg_color=("#FFFFFF", "#2E2E2E"),
            button_color=("#FFFFFF", "#2E2E2E"),
            text_color=("#000000", "#FFFFFF"),
            button_hover_color=("#E5E5E5", "#2A2A2A")
        )
        self.count_team_menu.set(self.count_selected_team)
        self.count_team_menu.pack(side="left", padx=(8, 8))

        # Filter and Clear buttons
        self.count_filter_btn = ctk.CTkButton(
            top_row,
            text="🔍 Filter",
            width=60,
            height=36,
            corner_radius=8,
            fg_color="#2471A3",
            hover_color="#1A5276",
            font=("Arial", 11, "bold"),
            command=self.apply_counts_filter
        )
        self.count_filter_btn.pack(side="left", padx=(6, 4))

        self.count_clear_btn = ctk.CTkButton(
            top_row,
            text="✖ Clear",
            width=60,
            height=36,
            corner_radius=8,
            fg_color="#566573",
            hover_color="#424949",
            font=("Arial", 11, "bold"),
            command=self.clear_counts
        )
        self.count_clear_btn.pack(side="left")
        def create_small_card(parent, title, color, command=None):

            card = ctk.CTkFrame(
                parent,
                width=140,
                height=72,
                corner_radius=15,
                fg_color=("#FFFFFF", "#2B2B2B"),
                border_width=1,
                border_color=("#DDDDDD", "#3A3A3A")
            )

            card.pack(side="left", padx=(0, 16))
            card.pack_propagate(False)

            title_lbl = ctk.CTkLabel(
                card,
                text=title,
                font=("Arial", 12),
                text_color=("#666666", "#888888")
            )

            title_lbl.pack(pady=(10, 0))

            value_lbl = ctk.CTkLabel(
                card,
                text="0",
                font=("Arial", 20, "bold"),
                text_color=color,
                cursor="hand2" if command else ""
            )

            value_lbl.pack(pady=(2, 0))

            if command:
                title_lbl.bind("<Button-1>", lambda e: command())
                value_lbl.bind("<Button-1>", lambda e: command())
                card.bind("<Button-1>", lambda e: command())

            return value_lbl

        stats_wrap = ctk.CTkFrame(left_wrap, fg_color="transparent")
        stats_wrap.pack(anchor="w")

        self.total_lbl = create_small_card(stats_wrap, "Total Members", "#3498DB")
        self.submitted_lbl = create_small_card(stats_wrap, "Submitted", "#27AE60", lambda: self.show_member_list(True))
        self.not_submitted_lbl = create_small_card(stats_wrap, "Not Submitted", "#C0392B", lambda: self.show_member_list(False))

        filter_card = ctk.CTkFrame(
            self.container,
            corner_radius=14,
            fg_color=("#F9F9F9", "#1E1E1E"),
            border_width=1,
            border_color=("#E0E0E0", "#2C2C2C")
        )
        filter_card.pack(fill="x", padx=content_padx, pady=(4, 10))
        inner = ctk.CTkFrame(filter_card, fg_color="transparent")
        inner.pack(fill="x", padx=8, pady=8)

        def _lbl(parent, text):
            ctk.CTkLabel(parent, text=text,
                         font=("Arial", 11, "bold"),
                         text_color=("#555555", "#FFFFFF")).pack(side="left", padx=(0, 4))

        _lbl(inner, "From")
        self.start_cal = DatePickerButton(inner, initial_date=self.start_date)
        self.start_cal.pack(side="left", padx=(0, 8))

        _lbl(inner, "To")
        self.end_cal = DatePickerButton(inner, initial_date=self.end_date)
        self.end_cal.pack(side="left", padx=(0, 12))

        # team dropdown (no label to save space)
        self.team_menu = ctk.CTkOptionMenu(
            inner,
            values=self.team_options,
            command=self.set_report_team_filter,
            width=120,
            height=28,
            corner_radius=8,
            fg_color=("#FFFFFF", "#2E2E2E"),
            button_color=("#FFFFFF", "#2E2E2E"),
            text_color=("#000000", "#FFFFFF"),
            button_hover_color=("#E5E5E5", "#2A2A2A")
        )
        self.team_menu.set(self.selected_team)
        self.team_menu.pack(side="left", padx=(0, 8))

        _lbl(inner, "Member")
        self.member_entry = ctk.CTkEntry(
            inner,
            placeholder_text="Search name...",
            width=100,
            height=36,
            corner_radius=8,
            border_color=("#CCCCCC", "#2C2C2C"),
            border_width=1,
            fg_color=("#FFFFFF", "#1E1E1E"),
            text_color=("#000000", "#FFFFFF"),
            font=("Arial", 12)
        )
        self.member_entry.pack(side="left", padx=(0, 10), pady=2)

        def _btn(parent, text, color, hover, cmd, width=70):
            b = ctk.CTkButton(
                parent, text=text, width=width, height=36,
                corner_radius=9, font=("Arial", 11, "bold"),
                fg_color=color, hover_color=hover, command=cmd
            )
            b.pack(side="left", padx=4)
            return b

        _btn(inner, "🔍 Filter", "#2471A3", "#1A5276", self.refresh_view, width=60)
        _btn(inner, "✖ Clear", "#566573", "#424949", self.clear_filters, width=60)
        _btn(inner, "📄 PDF", "#C0392B", "#922B21", self.export_to_pdf, width=60)
        _btn(inner, "📥 Excel", "#16A085", "#117A65", self.export_to_csv, width=60)

        list_header = ctk.CTkFrame(self.container, fg_color="transparent")
        list_header.pack(fill="x", padx=content_padx, pady=(4, 2))
        ctk.CTkLabel(list_header, text="Report List", font=("Arial", 14, "bold"), text_color=("#333333", "#FFFFFF")).pack(side="left", padx=4)
        self.count_lbl = ctk.CTkLabel(list_header, text="", font=("Arial", 12), text_color="#566573")
        self.count_lbl.pack(side="left", padx=8)

        self.scroll = ctk.CTkScrollableFrame(
            self.container,
            fg_color=("#EEEEEE", "#1A1A1A"),
            corner_radius=14,
            border_width=1,
            border_color=("#DDDDDD", "#2C2C2C")
        )
        self.scroll.pack(fill="both", expand=True, padx=content_padx, pady=(0, 10))

        self.refresh_view()
        self.refresh_counts()

    def set_report_team_filter(self, value):
        self.selected_team = value

    def set_count_team_filter(self, value):
        self.count_selected_team = value

    def refresh_view(self):
        for w in self.scroll.winfo_children():
            w.destroy()

        start = self.start_cal.get_date()
        end = self.end_cal.get_date()
        if start > end:
            self._show_message("Start date cannot be later than End date.", "error")
            return

        member_search = self.member_entry.get().strip()
        team_id = self.team_map.get(self.selected_team)

        query = '''SELECT u.full_name, dr.report_date, dr.today_work,
                dr.tomorrow_work, dr.problems_issues, dr.shared_matters,
                t.team_name
            FROM users u
            JOIN daily_reports dr ON u.id = dr.user_id
            LEFT JOIN teams t ON u.team_id = t.team_id
            WHERE u.role='member' AND dr.report_date BETWEEN %s AND %s'''
        params = [start, end]

        if team_id is not None:
            query += " AND u.team_id = %s"
            params.append(team_id)

        if member_search:
            query += " AND u.full_name LIKE %s"
            params.append(f"%{member_search}%")

        query += " ORDER BY dr.report_date DESC, u.full_name ASC, dr.created_at DESC"

        try:
            self.db.cursor.execute(query, tuple(params))
            rows = self.db.cursor.fetchall()

            grouped = defaultdict(list)
            for r in rows:
                key = (r['report_date'], r['full_name'])
                grouped[key].append(r)

            self.count_lbl.configure(text=f"({len(grouped)} records)")
            if not grouped:
                ctk.CTkLabel(
                    self.scroll,
                    text="No records found.",
                    font=("Arial", 15, "bold"),
                    text_color=("#666666", "#AAAAAA")
                ).pack(
                    pady=20
                )
                return

            last_date = None
            today_date = datetime.today().date()
            for (date_val, member), tasks in grouped.items():
                if last_date != date_val:
                    date_sep = ctk.CTkFrame(self.scroll, fg_color="transparent")
                    date_sep.pack(fill="x", padx=8, pady=(8, 1))
                    day_str = date_val.strftime("%A")
                    date_str = date_val.strftime("%d %b %Y")
                    badge_color = "#E67E22" if date_val == today_date else ("#BDC3C7", "#2C3E50")
                    badge_text = f"  {date_str}  ·  {day_str}{'  ← TODAY' if date_val == today_date else ''}  "
                    ctk.CTkLabel(
                        date_sep, text=badge_text,
                        font=("Arial", 11, "bold"),
                        text_color=("black", "white"),
                        fg_color=badge_color, corner_radius=6,
                        padx=6, pady=2
                    ).pack(side="left")
                    last_date = date_val

                team_name = tasks[0].get('team_name') or "Unknown Team"
                card = ctk.CTkFrame(self.scroll,
                    corner_radius=12,
                    fg_color=("#FFFFFF", "#1E1E1E"),
                    border_width=1,
                    border_color=("#E0E0E0", "#2C2C2C")
                )
                card.pack(fill="x", pady=1, padx=10)
                info = ctk.CTkFrame(card, fg_color="transparent")
                info.pack(side="left", fill="both", expand=True, padx=10, pady=6)

                ctk.CTkLabel(
                    info,
                    text=f"{member} ({team_name})",
                    font=("Arial", 12, "bold"),
                    text_color=("#333333", "#E8EDF2")
                ).pack(anchor="w", pady=(0, 4))

                preview = tasks[0].get('today_work', "") if tasks else ""
                preview = preview or "No details yet."
                lines = preview.splitlines()
                if len(lines) > 2:
                    preview = "\n".join(lines[:2]) + " ..."
                elif len(preview) > 120:
                    preview = preview[:120] + "..."

                ctk.CTkLabel(
                    info,
                    text=preview,
                    wraplength=650,
                    justify="left",
                    anchor="w",
                    text_color=("#555555", "#AAB7C4")
                ).pack(anchor="w", fill="x", pady=(4, 0))

                btn_frame = ctk.CTkFrame(card, fg_color="transparent")
                btn_frame.pack(side="right", padx=12, pady=6)

                ctk.CTkButton(
                    btn_frame,
                    text="View",
                    width=60,
                    height=30,
                    corner_radius=8,
                    font=("Arial", 12, "bold"),
                    fg_color="#2E5B9A",
                    hover_color="#1F4E8C",
                    command=lambda tlist=tasks, m=member, d=date_val: self.open_detail_with_state(m, d, tlist)
                ).pack()
        except Exception as e:
            self._show_message(str(e, "error"))

    def refresh_counts(self):
        try:
            sel_date = self.count_date.get_date()
        except Exception:
            sel_date = datetime.today().date()

        try:

            # Apply team filter if selected
            team_id = self.team_map.get(self.count_selected_team)

            if team_id is None:
                q_total = "SELECT COUNT(*) AS cnt FROM users WHERE role='member'"
                self.db.cursor.execute(q_total)
                total = self.db.cursor.fetchone().get('cnt', 0) or 0

                q_sub = '''SELECT COUNT(DISTINCT u.id) AS cnt FROM users u
                           JOIN daily_reports dr ON u.id=dr.user_id
                           WHERE u.role='member' AND dr.report_date=%s'''
                self.db.cursor.execute(q_sub, (sel_date,))
                submitted = self.db.cursor.fetchone().get('cnt', 0) or 0
            else:
                q_total = "SELECT COUNT(*) AS cnt FROM users WHERE role='member' AND team_id=%s"
                self.db.cursor.execute(q_total, (team_id,))
                total = self.db.cursor.fetchone().get('cnt', 0) or 0

                q_sub = '''SELECT COUNT(DISTINCT u.id) AS cnt FROM users u
                           JOIN daily_reports dr ON u.id=dr.user_id
                           WHERE u.role='member' AND dr.report_date=%s AND u.team_id=%s'''
                self.db.cursor.execute(q_sub, (sel_date, team_id))
                submitted = self.db.cursor.fetchone().get('cnt', 0) or 0

            not_sub = max(0, total - submitted)
            self.total_lbl.configure(text=str(total))
            self.submitted_lbl.configure(text=str(submitted))
            self.not_submitted_lbl.configure(text=str(not_sub))
        except Exception as e:
            print("Error refreshing counts:", e)

    def apply_counts_filter(self):
        self.count_filter_date = self.count_date.get_date()
        self.refresh_counts()

    def show_member_list(self, submitted=True):
        try:
            sel_date = self.count_date.get_date()
        except Exception:
            sel_date = datetime.today().date()

        team_id = self.team_map.get(self.count_selected_team)

        if submitted:
            q = '''SELECT u.full_name, t.team_name FROM users u
                   LEFT JOIN teams t ON u.team_id=t.team_id
                   JOIN daily_reports dr ON u.id=dr.user_id
                   WHERE u.role='member' AND dr.report_date=%s'''
            params = [sel_date]
            if team_id is not None:
                q += " AND u.team_id=%s"
                params.append(team_id)
            q += " ORDER BY u.full_name ASC"
            title = f"Members who submitted on {sel_date}"
        else:
            q = '''SELECT u.full_name, t.team_name FROM users u
                   LEFT JOIN teams t ON u.team_id=t.team_id
                   WHERE u.role='member' AND u.id NOT IN
                   (SELECT user_id FROM daily_reports WHERE report_date=%s)'''
            params = [sel_date]
            if team_id is not None:
                q += " AND u.team_id=%s"
                params.append(team_id)
            q += " ORDER BY u.full_name ASC"
            title = f"Members who not Submitted on {sel_date}"

        try:
            self.db.cursor.execute(q, params)
            rows = self.db.cursor.fetchall() or []

            self.clear_container()
            content_padx = 80

            top = ctk.CTkFrame(self.container, fg_color="transparent")
            top.pack(fill="x", padx=content_padx, pady=20)

            ctk.CTkButton(
                top,
                text="← Back",
                width=80,
                height = 36,
                fg_color=("#DBDBDB", "#333333"),
                text_color=("black", "white"),
                hover_color=("#CFCFCF", "#444444"),
                corner_radius=8,
                command=self.back_to_list
            ).pack(side="left")

            ctk.CTkLabel(
                top,
                text=title,
                font=("Arial", 20, "bold"),
                text_color=("black", "white")
            ).pack(side="left", padx=20)

            # MAIN CARD
            main_card = ctk.CTkFrame(
                self.container,
                corner_radius=14,
                fg_color=("#F0F0F0", "#252525"),
                border_width=1,
                border_color=("#DDDDDD", "#333333")
            )

            main_card.pack(
                fill="both",
                expand=True,
                padx=content_padx,
                pady=(0, 20)
            )

            # INNER LIST BOX
            list_frame = ctk.CTkScrollableFrame(
                main_card,
                fg_color=("#FFFFFF", "#1E1E26"),
                corner_radius=12,
                border_width=1,
                border_color=("#DDDDDD", "#333333")
            )

            list_frame.pack(
                fill="both",
                expand=True,
                padx=20,
                pady=20
            )

            if not rows:
                ctk.CTkLabel(list_frame, text="No members found.", font=("Arial", 12)).pack(anchor="w", padx=8, pady=6)
            else:
                for i, row in enumerate(rows, start=1):
                    team_name = row.get('team_name') or "Unknown Team"
                    ctk.CTkLabel(
                        list_frame,
                        text=f"{i}. {row.get('full_name')} ({team_name})",
                        font=("Arial", 12)
                    ).pack(anchor="w", padx=8, pady=4)

            self.refresh_counts()
        except Exception as e:
            self._show_message(str(e, "error"))

    def open_detail_with_state(self, member, date_val, tasks):
        self.start_date = self.start_cal.get_date()
        self.end_date = self.end_cal.get_date()
        self.member_search = self.member_entry.get()
        self.selected_team = self.team_menu.get()
        team_name = tasks[0].get("team_name", "No Team") if tasks else "No Team"
        self.show_member_detail(member, team_name, date_val, tasks)

    def show_member_detail(self, member_name, team_name, date_val, tasks):
        self.clear_container()

        content_padx = 80
        top = ctk.CTkFrame(self.container, fg_color="transparent")
        top.pack(fill="x", padx=content_padx, pady=20)

        ctk.CTkButton(
            top,
            text="← Back",
            width=80,
            height = 36,
            fg_color=("#DBDBDB", "#333333"),
            text_color=("black", "white"),
            hover_color=("#CFCFCF", "#444444"),
            corner_radius=8,
            command=self.back_to_list
        ).pack(side="left")

        ctk.CTkLabel(
            top,
            text=f"{member_name} ({team_name}) - {date_val}",
            font=("Arial", 20, "bold"),
            text_color=("black", "white")
        ).pack(side="left", padx=20)

        self.form_scroll = ctk.CTkScrollableFrame(self.container, fg_color="transparent")
        self.form_scroll.pack(fill="both", expand=True, padx=content_padx, pady=10)

        COLOR_CONTAINER_BG = ("#F0F0F0", "#252525")
        COLOR_TEXT_MAIN = ("#1A1A1A", "#E8EDF2")

        for r in tasks:
            def add_readonly_section(title, content):
                content = content or "-"
                lines = content.split("\n")
                total_lines = 0
                for line in lines:
                    wrapped = max(1, (len(line) // 85) + 1)
                    total_lines += wrapped
                textbox_height = max(45, total_lines * 24)
                section = ctk.CTkFrame(self.form_scroll, fg_color=COLOR_CONTAINER_BG, corner_radius=10)
                section.pack(fill="x", pady=8, padx=5)
                ctk.CTkLabel(section, text=title, font=("Arial", 14, "bold"), text_color=COLOR_TEXT_MAIN).pack(anchor="w", padx=15, pady=(12, 6))
                text_box = ctk.CTkTextbox(section, height=int(textbox_height), fg_color=("#FFFFFF", "#1e1e26"), border_width=1, border_color=("#DBDBDB", "#333333"), text_color=("black", "white"), wrap="word")
                text_box.pack(fill="both", expand=True, padx=15, pady=(0, 15))
                text_box.insert("1.0", content)
                text_box.configure(state="disabled")

            add_readonly_section("Today's Work Detail", r.get('today_work', ""))
            add_readonly_section("Tomorrow's Work Schedule", r.get('tomorrow_work', ""))
            add_readonly_section("Problems / Issues", r.get('problems_issues', ""))
            add_readonly_section("Shared Matters", r.get('shared_matters', ""))

    def back_to_list(self):
        self.show_reports_list()
        self.start_cal.set_date(self.start_date)
        self.end_cal.set_date(self.end_date)
        self.team_menu.set(self.selected_team)
        self.member_entry.delete(0, "end")
        self.member_entry.insert(0, self.member_search)
        self.refresh_view()

    def clear_filters(self):
        today = datetime.today().date()

        # reset ONLY lower filters
        self.start_date = today
        self.end_date = today
        self.member_search = ""
        self.selected_team = "All Teams"

        self.start_cal.set_date(today)
        self.end_cal.set_date(today)

        self.team_menu.set("All Teams")

        self.member_entry.delete(0, "end")

        # refresh ONLY report list
        self.refresh_view()

    def clear_counts(self):
        today = datetime.today().date()
        self.count_filter_date = today
        try:
            if hasattr(self, 'count_date'):
                self.count_date.set_date(today)

            self.count_selected_team = "All Teams"

            if hasattr(self, 'count_team_menu'):
                self.count_team_menu.set("All Teams")

        except Exception:
            pass
        # ONLY refresh top counts
        self.refresh_counts()
    
    def export_to_pdf(self):
        start = self.start_cal.get_date()
        end = self.end_cal.get_date()
        member_search = self.member_entry.get().strip()
        team_id = self.team_map.get(self.selected_team)

        query = '''
        SELECT
            dr.report_date,
            u.full_name,
            t.team_name,
            dr.today_work,
            dr.tomorrow_work,
            dr.problems_issues,
            dr.shared_matters
            FROM users u
            JOIN daily_reports dr ON u.id = dr.user_id
            LEFT JOIN teams t ON u.team_id = t.team_id
            WHERE u.role='member'
            AND dr.report_date BETWEEN %s AND %s
        '''
        params = [start, end]

        if team_id is not None:
            query += " AND u.team_id = %s"
            params.append(team_id)

        if member_search:
            query += " AND u.full_name LIKE %s"
            params.append(f"%{member_search}%")

        query += '''
            ORDER BY
            dr.report_date DESC,
            u.full_name ASC,
            dr.created_at ASC
        '''

        try:
            self.db.cursor.execute(query, tuple(params))
            rows = self.db.cursor.fetchall()
            grouped = defaultdict(list)

            for r in rows:
                key = (r['report_date'], r['full_name'])
                grouped[key].append(r)

            rows = []

            for (report_date, member_name), reports in grouped.items():

                first = reports[0]

                rows.append({
                    "report_date": report_date,
                    "full_name": member_name,
                    "team_name": first.get("team_name", "-"),

                    "today_work": "\n".join(
                        r.get("today_work", "")
                        for r in reports
                        if r.get("today_work")
                    ),

                    "tomorrow_work": "\n".join(
                        r.get("tomorrow_work", "")
                        for r in reports
                        if r.get("tomorrow_work")
                    ),

                    "problems_issues": "\n".join(
                        r.get("problems_issues", "")
                        for r in reports
                        if r.get("problems_issues")
                    ),

                    "shared_matters": "\n".join(
                        r.get("shared_matters", "")
                        for r in reports
                        if r.get("shared_matters")
                    )
                })
            if not rows:
                self._show_message("No reports found.", "info")
                return

            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                initialfile=f"Admin_Daily_Reports.pdf"
            )
            if not file_path:
                return

            class PDF(FPDF):
                def header(self):
                    self.set_draw_color(169, 169, 169)
                    self.set_fill_color(0, 0, 139)
                    self.set_text_color(0, 0, 0)
                    self.set_font("Arial", "B", 20)
                    self.cell(0, 12, "Admin Daily Reports", ln=True, align="C")
                    self.ln(4)
                    self.set_font("Arial", "B", 11)
                    self.set_text_color(0, 0, 0)
                    self.cell(14, 8, "Date :", ln=0)
                    self.set_font("Arial", "", 11)
                    self.cell(0, 8, f"{self.start_date} ~ {self.end_date}", ln=True)
                    self.ln(4)
                    self.set_font("Arial", "B", 10)
                    headers = ["Date", "Name", "Team", "Today's Work Detail", "Tomorrow's Work Schedule", "Problems/Issues", "Shared Matters"]
                    widths = self.col_widths
                    self.set_draw_color(169, 169, 169)
                    self.set_fill_color(0, 0, 139)
                    self.set_text_color(245, 245, 245)
                    for i, header in enumerate(headers):
                        self.cell(widths[i], 12, header, border=1, align="C", fill=True)
                    self.ln()
                    self.set_text_color(0, 0, 0)
                    self.set_fill_color(255, 255, 255)

            pdf = PDF("L", "mm", "A4")
            pdf.start_date = start
            pdf.end_date = end
            pdf.col_widths = [28, 32, 28, 52, 52, 38, 38]
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()
            body_font_size = 8
            pdf.set_font("Arial", "", body_font_size)
            line_height = 5

            def split_text(text, width):
                if not text:
                    return ["-"]
                pdf.set_font("Arial", "", body_font_size)
                lines = []
                for paragraph in str(text).split("\n"):
                    words = paragraph.split(" ")
                    current_line = ""
                    for word in words:
                        test_line = (current_line + " " + word).strip()
                        if pdf.get_string_width(test_line) <= width - 2:
                            current_line = test_line
                        else:
                            if current_line:
                                lines.append(current_line)
                            current_line = word
                    if current_line:
                        lines.append(current_line)
                return lines if lines else ["-"]

            for r in rows:
                cells = [
                    str(r['report_date']),
                    str(r.get('full_name') or "-"),
                    str(r.get('team_name') or "-"),
                    str(r.get('today_work') or "-"),
                    str(r.get('tomorrow_work') or "-"),
                    str(r.get('problems_issues') or "-"),
                    str(r.get('shared_matters') or "-")
                ]
                split_cells = []
                max_lines = 1
                for i, text in enumerate(cells):
                    lines = split_text(text, pdf.col_widths[i])
                    split_cells.append(lines)
                    max_lines = max(max_lines, len(lines))
                row_height = (max_lines * line_height) + 3
                if pdf.get_y() + row_height > 190:
                    pdf.add_page()
                    pdf.set_font("Arial", "", body_font_size)
                pdf.set_draw_color(169, 169, 169)
                x = pdf.get_x(); y = pdf.get_y()
                for width in pdf.col_widths:
                    pdf.rect(x, y, width, row_height)
                    x += width
                x = pdf.l_margin
                for i, lines in enumerate(split_cells):
                    current_y = y + 2
                    for line in lines:
                        pdf.set_xy(x + 2, current_y)
                        pdf.cell(pdf.col_widths[i] - 4, line_height, line, border=0)
                        current_y += line_height
                    x += pdf.col_widths[i]
                pdf.set_y(y + row_height)

            pdf.output(file_path)
            self._show_message("PDF exported successfully!", "success")
        except Exception as e:
            self._show_message(str(e, "error"))

    def export_to_csv(self):
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment
        except ImportError:
            self._show_message("openpyxl is required to export Excel files. Install it with:\n\npip install openpyxl", "error")
            return

        start = self.start_cal.get_date()
        end = self.end_cal.get_date()
        member_search = self.member_entry.get().strip()
        team_id = self.team_map.get(self.selected_team)

        query = '''
        SELECT
            dr.report_date,
            u.full_name,
            t.team_name,
            dr.today_work,
            dr.tomorrow_work,
            dr.problems_issues,
            dr.shared_matters
            FROM users u
            JOIN daily_reports dr ON u.id = dr.user_id
            LEFT JOIN teams t ON u.team_id = t.team_id
            WHERE u.role='member'
            AND dr.report_date BETWEEN %s AND %s
        '''
        params = [start, end]

        if team_id is not None:
            query += " AND u.team_id = %s"
            params.append(team_id)

        if member_search:
            query += " AND u.full_name LIKE %s"
            params.append(f"%{member_search}%")

        query += '''
            ORDER BY
            dr.report_date DESC,
            u.full_name ASC,
            dr.created_at ASC
        '''

        try:
            self.db.cursor.execute(query, tuple(params))
            rows = self.db.cursor.fetchall()
            grouped = defaultdict(list)

            for r in rows:
                key = (r['report_date'], r['full_name'])
                grouped[key].append(r)

            rows = []

            for (report_date, member_name), reports in grouped.items():

                first = reports[0]

                rows.append({
                    "report_date": report_date,
                    "full_name": member_name,
                    "team_name": first.get("team_name", "-"),

                    "today_work": "\n".join(
                        r.get("today_work", "")
                        for r in reports
                        if r.get("today_work")
                    ),

                    "tomorrow_work": "\n".join(
                        r.get("tomorrow_work", "")
                        for r in reports
                        if r.get("tomorrow_work")
                    ),

                    "problems_issues": "\n".join(
                        r.get("problems_issues", "")
                        for r in reports
                        if r.get("problems_issues")
                    ),

                    "shared_matters": "\n".join(
                        r.get("shared_matters", "")
                        for r in reports
                        if r.get("shared_matters")
                    )
                })
            if not rows:
                self._show_message("No reports found.", "info")
                return

            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"Admin_Daily_Reports.xlsx"
            )
            if not file_path:
                return
            base, ext = os.path.splitext(file_path)
            if ext.lower() != ".xlsx":
                file_path = base + ".xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Reports"

            # Title and date rows (mimic PDF)
            ncols = 7
            title = "Admin Daily Reports"
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            tcell = sheet.cell(row=1, column=1, value=title)
            tcell.font = tcell.font.copy(bold=True, size=18, color="000000")
            tcell.alignment = tcell.alignment.copy(horizontal='center', vertical='center')
            sheet.row_dimensions[1].height = 26

            # Date row
            date_text = f"Date : {self.start_date} ~ {self.end_date}"
            sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
            dcell = sheet.cell(row=2, column=1, value=date_text)
            dcell.font = dcell.font.copy(bold=False, size=11, color="000000")
            dcell.alignment = dcell.alignment.copy(horizontal='left', vertical='center')
            sheet.row_dimensions[2].height = 18

            headers = [
                "Date",
                "Name",
                "Team",
                "Today's Work Detail",
                "Tomorrow's Work Schedule",
                "Problems/Issues",
                "Shared Matters"
            ]

            # Header row index
            hdr_row = 4

            from openpyxl.styles import (
                PatternFill,
                Border,
                Side,
                Font,
                Alignment
            )

            header_fill = PatternFill(
                fill_type="solid",
                start_color="D9D9D9",
                end_color="D9D9D9"
            )

            side = Side(
                border_style="thin",
                color="000000"
            )

            for c, h in enumerate(headers, start=1):

                cell = sheet.cell(
                    row=hdr_row,
                    column=c,
                    value=h
                )

                cell.font = Font(
                    bold=True,
                    color="000000"
                )

                cell.fill = header_fill

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

                cell.border = Border(
                    left=side,
                    right=side,
                    top=side,
                    bottom=side
                )

            sheet.row_dimensions[hdr_row].height = 25

            # Append data rows starting after header
            for r in rows:
                values = [
                    r['report_date'],
                    r.get('full_name') or "-",
                    r.get('team_name') or "-",
                    r.get('today_work') or "-",
                    r.get('tomorrow_work') or "-",
                    r.get('problems_issues') or "-",
                    r.get('shared_matters') or "-"
                ]
                row_idx = sheet.max_row + 1
                for col_idx, value in enumerate(values, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    # apply border only if data exists
                    side = Side(
                        border_style="thin",
                        color="A9A9A9"
                    )

                    cell.border = Border(
                        left=side,
                        right=side,
                        top=side,
                        bottom=side
                    )

            # Auto-size columns based on longest line
            widths = []
            for col_idx in range(1, ncols + 1):
                max_width = 0
                for row_idx in range(1, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value or ""
                    for line in str(cell_value).splitlines():
                        max_width = max(max_width, len(line))
                widths.append(max(max_width + 2, 15))
                sheet.column_dimensions[get_column_letter(col_idx)].width = widths[-1]

            def estimate_row_height(row_idx):
                max_lines = 1
                for col_idx, width in enumerate(widths, start=1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value or ""
                    lines = []
                    for paragraph in str(cell_value).splitlines():
                        current = ""
                        for word in paragraph.split(" "):
                            test_line = (current + " " + word).strip()
                            if len(test_line) <= width:
                                current = test_line
                            else:
                                if current:
                                    lines.append(current)
                                current = word
                        if current:
                            lines.append(current)
                    max_lines = max(max_lines, len(lines))
                return max(18, max_lines * 15)

            for row_idx in range(3, sheet.max_row + 1):
                sheet.row_dimensions[row_idx].height = estimate_row_height(row_idx)
                
            from openpyxl.styles import Border, Side

            side = Side(
                border_style="thin",
                color="A9A9A9"
            )

            for row in sheet.iter_rows(
                min_row=4,              # header row
                max_row=sheet.max_row,
                min_col=1,
                max_col=7               # Date -> Shared Matters
            ):
                for cell in row:
                    cell.border = Border(
                        left=side,
                        right=side,
                        top=side,
                        bottom=side
                    )

            workbook.save(file_path)
            self._show_message("Excel exported successfully!", "success")
        except Exception as e:
            self._show_message(str(e, "error"))
